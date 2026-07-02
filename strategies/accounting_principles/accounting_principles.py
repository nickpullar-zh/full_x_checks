"""
Accounting Principles strategy.

Compares the severity letter recorded for each X-Check on the EBX 'cross
checks all' sheet against the W/E recorded in the 'FIP Methods Rules and
Condition' sheet (the VALMSG dump). The Validation Methods xlsx supplies
the candidate methods per Validation Event and their expected severity.
"""
from __future__ import annotations

import pandas as pd

from strategies.base_strategy import BaseStrategy, UploadTaskConfig
from .validation_methods import parse_method_bindings
from .compare import compare_with_bindings


# Default subset of Validation Events used on first run (before the user's
# choices have been persisted to %APPDATA%/X-Checks/accounting_principles.json).
# Names must match the row-1 headers of the validation methods file exactly.
DEFAULT_EVENTS = [
    "IFRS New RFD", "IFRS New SFD", "IFRS New CFD",
    "Stammhaus SLST RFD", "Stammhaus SLST SFD", "Stammhaus SLST CFD",
    "SST RFD", "SST SFD", "SST CFD",
    "SII RFD", "SII SFD", "SII CFD",
    "RI Assets IFRSN RFD", "RI Assets IFRSN SFD", "RI Assets IFRSN CFD",
    "Equity IFRSN RFD", "Equity IFRSN SFD", "Equity IFRSN CFD",
    "I/C IFRSN RFD", "I/C IFRSN SFD", "I/C IFRSN CFD",
    "Tax IFRSN RFD", "Tax IFRSN SFD", "Tax IFRSN CFD",
    "DE-GAAP RFD", "DE-GAAP SFD", "DE-GAAP CFD",
]

OUTPUT_COLUMNS = [
    "X-Check No.", "Event", "Expected", "FIP", "Actual", "Method", "Match",
]


class AccountingPrinciples(BaseStrategy):

    def __init__(self, config: UploadTaskConfig):
        super().__init__(config)

    def process(self, loaded_files: dict, files: dict):
        self.log_step(self.log, "System", "Starting Accounting Principles processing", 0)

        vm_path = files["files"].get("Validation Methods File")
        if not vm_path:
            self.log_step(self.log, "System",
                          "Missing required file: Validation Methods File", 0)
            return

        # 1. Validation Methods: parse method bindings (one per cell occurrence)
        subset = files.get("validation_events_subset") or DEFAULT_EVENTS
        self.log_step(self.log, "Validation Methods",
                      "Parsing validation methods file...", len(subset))
        bindings = parse_method_bindings(vm_path, subset)
        self.log_step(self.log, "Validation Methods",
                      "Method bindings extracted", len(bindings))

        # 2. Cross Checks All — already loaded by BaseStrategy._load_files,
        #    using header_signals to detect the right header row automatically.
        cc_df = loaded_files.get("X-Checks Publication File")
        if cc_df is None:
            self.log_step(self.log, "System",
                          "Missing required file: X-Checks Publication File", 0)
            return

        # 3. FIP Methods Rules and Condition — also already loaded.
        fip_df = loaded_files.get("FIP File (VALMSG)")
        if fip_df is None:
            self.log_step(self.log, "System",
                          "Missing required file: FIP File (VALMSG)", 0)
            return

        # 4. In-scope X-Checks. When 'Process only differences' is on, use the
        # same pipeline as v0.4.1 X-Check No Selection (drop INACTIVE rows,
        # keep non-blank Type of Change, drop Exclude Z-Core = X, drop yellow
        # Category). Otherwise take every unique non-blank X-Check No.
        x_check_col = "X-Check No."
        if x_check_col not in cc_df.columns:
            self.log_step(self.log, "EBX",
                          f"Required column '{x_check_col}' not found — aborting", 0)
            return

        if files.get("process_only_differences", False):
            ebx_path  = files["files"].get("X-Checks Publication File")
            ebx_sheet = files["sheet_names"].get("X-Checks Publication File",
                                                 "cross checks all")
            xchecks = self._select_in_scope_x_checks(cc_df, ebx_path, ebx_sheet)
            self.log_step(self.log, "EBX",
                          "In-scope X-Check Nos (filtered)", len(xchecks))
        else:
            xchecks: list[str] = []
            seen: set = set()
            for v in cc_df[x_check_col].tolist():
                s = str(v).strip()
                if s in ("", "nan", "None") or s in seen:
                    continue
                seen.add(s)
                xchecks.append(s)
            self.log_step(self.log, "EBX", "In-scope X-Check Nos", len(xchecks))

        # 5. Compare
        rows = compare_with_bindings(bindings, cc_df, xchecks, fip_df)
        if not rows:
            self.log_step(self.log, "Compare",
                          "No comparable rows produced — aborting output", 0)
            return
        df_compare = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)
        self.log_step(self.log, "Compare", "Comparison rows produced", len(df_compare))

        # 5b. Apply known exceptions (annotation only — Match column unchanged)
        _AP_FINGERPRINT = ["X-Check No.", "Event", "Expected", "FIP", "Actual", "Method"]
        result = self._annotate_known_exceptions(
            df_compare, files["files"].get("Known Exception List"),
            sheet_name="Accounting Principles", fingerprint_columns=_AP_FINGERPRINT
        )
        if result is False:
            return
        df_compare = result

        # 6. EBX sheet: cross-checks-all rows for in-scope X-Checks only
        in_scope_set = set(xchecks)
        ebx_mask = cc_df[x_check_col].astype(str).str.strip().isin(in_scope_set)
        df_ebx = cc_df.loc[ebx_mask].reset_index(drop=True)
        self.log_step(self.log, "EBX", "Rows kept for output sheet", len(df_ebx))

        # 7. FIP sheet: only rows whose V-code is in the validation-methods subset
        known_methods = {b.method for b in bindings}
        def _v_code(key: str) -> str:
            s = str(key)
            return s.split("|", 1)[0].strip() if "|" in s else ""
        fip_mask = fip_df["Key"].astype(str).map(_v_code).isin(known_methods)
        df_fip = fip_df.loc[fip_mask].reset_index(drop=True)
        self.log_step(self.log, "FIP", "Rows kept for output sheet", len(df_fip))

        # 8. Write output. Sheet order: EBX, FIP, Comparison.
        out_path = self.build_output_path(
            files["output_directory"],
            "Accounting Principles Comparison",
            files["timestamp"],
        )
        self.write_excel_output(
            output_path=out_path,
            sheets={
                "EBX":        df_ebx,
                "FIP":        df_fip,
                "Comparison": df_compare,
            },
            log=self.log,
        )
        return True

    # Column names used by the 'Process only differences' filter, matched
    # case-insensitively against the cross-checks-all header row.
    _COL_STATUS    = "Status"
    _COL_TYPE_CHG  = "Type of change"
    _COL_EXCL_ZC   = "Exclude Z-Core"
    _COL_CATEGORY  = "Category"
    _COL_X_CHECK   = "X-Check No."
    _YELLOW_RGB    = "FFFF00"

    def _select_in_scope_x_checks(self, cc_df, ebx_path: str, ebx_sheet: str) -> list[str]:
        """
        Mirrors v0.4.1 select_x_check_nos: drop INACTIVE rows, keep non-blank
        Type of change, drop X-Checks where Exclude Z-Core = X, drop X-Checks
        whose Category cell is filled with standard Excel yellow (#FFFF00).
        Returns unique X-Check Nos in order of first appearance.
        """
        import openpyxl
        # Resolve actual column names case-insensitively
        def _resolve(name: str) -> str | None:
            target = name.casefold()
            for c in cc_df.columns:
                if str(c).casefold() == target:
                    return c
            return None

        col_x       = _resolve(self._COL_X_CHECK)
        col_status  = _resolve(self._COL_STATUS)
        col_type    = _resolve(self._COL_TYPE_CHG)
        col_excl_zc = _resolve(self._COL_EXCL_ZC)
        col_cat     = _resolve(self._COL_CATEGORY)
        if col_x is None:
            return []

        df = cc_df.copy()
        if col_status is not None:
            mask = df[col_status].astype(str).str.strip().str.upper() != "INACTIVE"
            df = df[mask]
        if col_type is not None:
            toc = df[col_type].astype(str).str.strip()
            df = df[~toc.isin(("", "nan", "None"))]

        # Z-Core exclusion: drop X-Checks with any 'X' on this column
        excluded_x: set = set()
        if col_excl_zc is not None:
            zc = df[col_excl_zc].astype(str).str.strip().str.upper()
            for _, r in df[zc == "X"].iterrows():
                excluded_x.add(str(r[col_x]).strip())

        # Yellow Category exclusion (read original sheet for fill colour)
        yellow_x: set = set()
        if col_cat is not None and ebx_path and ebx_sheet:
            try:
                wb = openpyxl.load_workbook(ebx_path, data_only=True)
                ws = wb[ebx_sheet]
                # Find the Excel column index for Category from the header row.
                # Header may be row 1 or row 2 — auto-detect by scanning rows 1-6
                # for a row that contains 'X-Check No.' / 'Status' / 'Type of change'.
                hdr_row = None
                wanted = {self._COL_X_CHECK.casefold(),
                          self._COL_STATUS.casefold(),
                          self._COL_TYPE_CHG.casefold()}
                for r_idx in range(1, 7):
                    cells = {str(ws.cell(r_idx, c).value).strip().casefold()
                             for c in range(1, ws.max_column + 1)
                             if ws.cell(r_idx, c).value is not None}
                    if wanted.issubset(cells):
                        hdr_row = r_idx
                        break
                if hdr_row is not None:
                    cat_col = None
                    xc_col = None
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(hdr_row, c).value
                        if v is None: continue
                        s = str(v).strip().casefold()
                        if s == self._COL_CATEGORY.casefold(): cat_col = c
                        elif s == self._COL_X_CHECK.casefold(): xc_col = c
                    if cat_col is not None and xc_col is not None:
                        for r in range(hdr_row + 1, ws.max_row + 1):
                            cell = ws.cell(r, cat_col)
                            fill = cell.fill
                            if (fill and fill.fill_type and fill.fill_type != "none"
                                and fill.fgColor and fill.fgColor.type == "rgb"):
                                rgb = str(fill.fgColor.rgb).upper()
                                if rgb[-6:] == self._YELLOW_RGB:
                                    xv = ws.cell(r, xc_col).value
                                    if xv is not None:
                                        yellow_x.add(str(xv).strip())
                wb.close()
            except Exception as e:
                self.log_step(self.log, "EBX",
                              f"Could not read Category fill colour: {e}", 0)

        out: list[str] = []
        seen: set = set()
        for v in df[col_x].tolist():
            s = str(v).strip()
            if s in ("", "nan", "None") or s in seen:
                continue
            if s in excluded_x or s in yellow_x:
                continue
            seen.add(s)
            out.append(s)
        return out

    def apply_output_formatting(self, workbook):
        if "Comparison" not in workbook.sheetnames:
            return
        self.apply_conditional_formatting(
            worksheet=workbook["Comparison"],
            column_name="Match",
            rules={
                "Match":    (self.FILL_GREEN, self.FONT_GREEN),
                "MisMatch": (self.FILL_RED,   self.FONT_RED),
            },
        )
