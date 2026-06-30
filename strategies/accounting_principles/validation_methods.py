"""
Parses the Validation Methods xlsx (sheet 'Validation Methods').

Layout (Q1 2026 file):
  - Row 1, cols C..BM: Validation Event names (the 'Accounting Principles').
  - Rows 4-6 hold the current period block ('from 2023'):
      * row 4 = Warning row
      * rows 5, 6 = Error rows
  - A cell merged across rows 4-5 (or 4-6) for a single column means that
    event accepts EITHER 'w' or 'e' on a cross-checks-all row, i.e. severity = Both.
  - Method cells contain newline-separated entries like
        "V900A - Part A (2023 onwards)"
    The leading code before " - " is the method ID we keep.
  - "-" or empty cells mean 'no entry' for that severity.
  - Cells using grey font (theme=1, tint > 0) are 'reference copies' — they
    bind the same V-code to a second event but only count when no higher-priority
    binding produces a non-empty actual letter on cross-checks-all.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
from openpyxl.cell.cell import MergedCell


SHEET_NAME = "Validation Methods"
EVENT_ROW = 1
WARNING_ROW = 4
ERROR_ROWS = (5, 6)
EVENT_COL_START = 3   # column C


@dataclass
class EventDefinition:
    """
    Severity declared for a Validation Event in the current-period block,
    plus the method codes the user should expect to see for each severity.

    Kept for backwards compatibility with existing tests. The comparator
    now consumes MethodBinding instead.
    """
    event: str
    severity: str            # "Warning" | "Error" | "Both"
    methods_w: list[str] = field(default_factory=list)
    methods_e: list[str] = field(default_factory=list)


@dataclass
class MethodBinding:
    """
    One (V-code, event, severity, font-colour, column) binding gleaned from
    a single cell in the validation methods file. The comparator orders these
    per V-code by (font_priority, column) and walks them in priority order
    when looking up actual letters on cross-checks-all rows.
    """
    method:    str            # e.g. 'V900W'
    event:     str            # e.g. 'IFRS New RFD'
    severity:  str            # 'Warning' | 'Error' | 'Both'
    font:      str            # 'black' | 'grey'
    column:    int            # 1-based Excel column index (for left-to-right ordering)


_METHOD_LINE_RE = re.compile(r"^\s*([A-Za-z0-9]+)")


def _extract_method_codes(cell_value) -> list[str]:
    """
    Splits a cell's content on newlines and pulls the leading code from each line.
    'V900A - Part A...' -> 'V900A'
    Empty / dash / whitespace-only lines are dropped.
    """
    if cell_value is None:
        return []
    text = str(cell_value).strip()
    if not text or text == "-":
        return []
    codes: list[str] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or line == "-":
            continue
        m = _METHOD_LINE_RE.match(line)
        if m:
            codes.append(m.group(1))
    return codes


def _is_blank(cell_value) -> bool:
    """Empty, whitespace-only, or '-' cells count as blank."""
    if cell_value is None:
        return True
    s = str(cell_value).strip()
    return s == "" or s == "-"


def _font_kind(cell) -> str:
    """
    Returns 'grey' if the cell's font is theme=1 with tint > 0 (the canonical
    'reference copy' style in this file), else 'black'.
    """
    f = cell.font
    if f is None or f.color is None:
        return "black"
    c = f.color
    if c.type == "theme" and c.theme == 1 and c.tint and c.tint > 0:
        return "grey"
    return "black"


def _column_is_in_warning_merge_with_error(ws, col: int) -> bool:
    """True if a single merged region in this column spans the Warning row AND
    at least one Error row (i.e. the cell's content covers Warning+Error =
    severity 'Both' for this column)."""
    for rng in ws.merged_cells.ranges:
        if rng.min_col <= col <= rng.max_col:
            covers_warning = rng.min_row <= WARNING_ROW <= rng.max_row
            covers_error   = any(rng.min_row <= er <= rng.max_row for er in ERROR_ROWS)
            if covers_warning and covers_error:
                return True
    return False


def _merge_origin(ws, row: int, col: int) -> tuple[int, int]:
    """Returns the (origin_row, origin_col) of the merged range covering (row, col),
    or (row, col) if not in a merge."""
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return (row, col)
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return (rng.min_row, rng.min_col)
    return (row, col)


def _read_cell_through_merges(ws, row: int, col: int):
    """Returns the value at (row,col), following merged-cell semantics so
    every cell in a merged range yields the merged value (not None)."""
    o_row, o_col = _merge_origin(ws, row, col)
    return ws.cell(row=o_row, column=o_col).value


def parse_validation_methods(filepath: str, subset: list[str]) -> list[EventDefinition]:
    """
    Returns a list of EventDefinition records, one per (event, severity)
    combination with non-empty methods. Kept for tests and backward compat;
    the comparator uses parse_method_bindings() instead.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in {filepath}")
    ws = wb[SHEET_NAME]

    subset_set = set(subset)
    col_for_event: dict[str, int] = {}
    for col in range(EVENT_COL_START, ws.max_column + 1):
        v = ws.cell(row=EVENT_ROW, column=col).value
        if v is None:
            continue
        name = str(v).strip()
        if name in subset_set and name not in col_for_event:
            col_for_event[name] = col

    results: list[EventDefinition] = []

    for event in subset:
        col = col_for_event.get(event)
        if col is None:
            continue

        warning_val = _read_cell_through_merges(ws, WARNING_ROW, col)
        error_vals  = [_read_cell_through_merges(ws, r, col) for r in ERROR_ROWS]

        warning_methods = _extract_method_codes(warning_val)
        error_methods: list[str] = []
        for v in error_vals:
            error_methods.extend(_extract_method_codes(v))

        is_both = _column_is_in_warning_merge_with_error(ws, col)

        if is_both:
            merged_methods = list(dict.fromkeys(warning_methods + error_methods))
            if not merged_methods:
                continue
            results.append(EventDefinition(
                event=event, severity="Both",
                methods_w=merged_methods, methods_e=merged_methods,
            ))
            continue

        if warning_methods and error_methods:
            results.append(EventDefinition(
                event=event, severity="Warning",
                methods_w=warning_methods, methods_e=[],
            ))
            results.append(EventDefinition(
                event=event, severity="Error",
                methods_w=[], methods_e=error_methods,
            ))
        elif warning_methods:
            results.append(EventDefinition(
                event=event, severity="Warning",
                methods_w=warning_methods, methods_e=[],
            ))
        elif error_methods:
            results.append(EventDefinition(
                event=event, severity="Error",
                methods_w=[], methods_e=error_methods,
            ))

    wb.close()
    return results


def parse_method_bindings(filepath: str, subset: list[str]) -> list[MethodBinding]:
    """
    Reads the Validation Methods sheet and returns one MethodBinding per
    (V-code, event, severity, font-colour, column) cell occurrence. Order in
    the returned list reflects file order; callers sort/filter as needed.

    For black font we trust every binding. For grey we still emit the binding
    but mark font='grey' so the comparator can decide whether to use it.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in {filepath}")
    ws = wb[SHEET_NAME]

    subset_set = set(subset)
    bindings: list[MethodBinding] = []

    for col in range(EVENT_COL_START, ws.max_column + 1):
        h = ws.cell(row=EVENT_ROW, column=col).value
        if h is None:
            continue
        event = str(h).strip()
        if event not in subset_set:
            continue

        is_both = _column_is_in_warning_merge_with_error(ws, col)
        if is_both:
            # All methods on the merged cell apply for both 'w' and 'e' actuals.
            origin_row, origin_col = _merge_origin(ws, WARNING_ROW, col)
            cell = ws.cell(row=origin_row, column=origin_col)
            font = _font_kind(cell)
            for m in _extract_method_codes(cell.value):
                bindings.append(MethodBinding(
                    method=m, event=event, severity="Both",
                    font=font, column=col,
                ))
            continue

        # Independent Warning / Error rows
        for r, sev in ((WARNING_ROW, "Warning"),
                       (ERROR_ROWS[0], "Error"),
                       (ERROR_ROWS[1], "Error")):
            origin_row, origin_col = _merge_origin(ws, r, col)
            cell = ws.cell(row=origin_row, column=origin_col)
            if _is_blank(cell.value):
                continue
            font = _font_kind(cell)
            for m in _extract_method_codes(cell.value):
                bindings.append(MethodBinding(
                    method=m, event=event, severity=sev,
                    font=font, column=col,
                ))

    wb.close()
    return bindings


def list_all_event_names(filepath: str) -> list[str]:
    """Returns every Validation Event name in row 1 (cols C..BM), in column order."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[SHEET_NAME]
    names: list[str] = []
    for col in range(EVENT_COL_START, ws.max_column + 1):
        v = ws.cell(row=EVENT_ROW, column=col).value
        if v is not None:
            n = str(v).strip()
            if n:
                names.append(n)
    wb.close()
    return names
