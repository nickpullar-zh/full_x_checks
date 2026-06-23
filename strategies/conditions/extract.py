"""
Extract condition data from the X-Checks Publication file.

Scans the 'cross checks all' sheet for yellow and green cells in the 5
condition columns, collects the associated X-Check No. values, deduplicates
them, and returns a working DataFrame ready for comparison.

Yellow rule: every row where the condition cell is yellow → collect X-Check No.
Green rule:  rows where column-A cell is green → collect X-Check No. for every
             non-blank value in the 5 condition columns.
"""

import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string

CONDITION_COLS = [
    "Reference  X-Check (Condition)",   # two spaces — matches the actual publication file header
    "Applicable Quarters",
    "Included RUs",
    "Excluded RUs",
    "Reference X-Check (Limit, %)",
]

# Hex colours Excel uses for its standard yellow and green highlights.
# Values stored in openpyxl as 8-char ARGB strings (FF prefix = fully opaque).
_YELLOW_RGBS = {"FFFFFF00", "FFFFC000", "FFFFEB9C"}   # yellow / dark-yellow / light-yellow
_GREEN_RGBS  = {"FF92D050", "FF00B050", "FFC6EFCE",   # standard greens
                "FF70AD47", "FF548235"}


def _resolve_rgb(cell, workbook) -> str | None:
    """
    Return the 8-char ARGB hex string for a cell's foreground fill colour,
    or None if there is no solid fill.

    openpyxl represents colours in three ways:
      - rgb:     direct ARGB hex string  (most common for manually-filled cells)
      - indexed: palette index           (legacy Excel colour table)
      - theme:   theme colour index      (rarely used for highlights)
    """
    fill = cell.fill
    if fill is None or fill.fill_type not in ("solid", "patternFill") and fill.patternType not in (None, "solid"):
        # PatternFill with patternType="solid" is the standard highlight fill.
        # fill_type is used on newer openpyxl; patternType on the raw XML object.
        pass

    fg = getattr(fill, "fgColor", None)
    if fg is None:
        return None

    if fg.type == "rgb":
        rgb = fg.rgb  # e.g. "FFFF0000"
        if rgb and rgb != "00000000":
            return rgb.upper()

    elif fg.type == "indexed":
        # Indexed colour table (0-63). Map the small subset we care about.
        _INDEXED_YELLOW = {13, 27, 36}   # Excel palette entries for yellow variants
        _INDEXED_GREEN  = {10, 17, 35, 42, 50}
        idx = fg.indexed
        if idx in _INDEXED_YELLOW:
            return "FFFFFF00"
        if idx in _INDEXED_GREEN:
            return "FF92D050"

    elif fg.type == "theme":
        # Theme colours are not easy to resolve without full theme XML parsing.
        # In practice, manual highlights are never theme-based; skip gracefully.
        pass

    return None


def _is_yellow(cell, workbook) -> bool:
    rgb = _resolve_rgb(cell, workbook)
    return rgb in _YELLOW_RGBS if rgb else False


def _is_green(cell, workbook) -> bool:
    rgb = _resolve_rgb(cell, workbook)
    return rgb in _GREEN_RGBS if rgb else False


def extract_conditions(
    pub_path: str,
    sheet_name: str,
    process_only_differences: bool = True,
) -> tuple[pd.DataFrame, list[str]]:
    """
    Load the publication workbook and extract condition rows.

    Args:
        process_only_differences: when True (default / checkbox checked), collect
            only condition cells that are themselves yellow or green — the changed/
            new rows (~20 rows). When False (checkbox unchecked), collect every
            non-blank condition cell regardless of colour — the full file.

    Returns:
        working_df  — columns: ["X-Check No."] + CONDITION_COLS +
                       [col + " (Concat)" for col in CONDITION_COLS]
        warnings    — list of non-fatal warning strings
    """
    warnings: list[str] = []

    wb = openpyxl.load_workbook(pub_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {pub_path}")
    ws = wb[sheet_name]

    # ------------------------------------------------------------------
    # Locate header row and map column names → column indices (1-based)
    # ------------------------------------------------------------------
    header_row_idx = None
    col_index: dict[str, int] = {}  # column name → 1-based column index

    for row in ws.iter_rows(min_row=1, max_row=20):
        for cell in row:
            if str(cell.value).strip() == "X-Check No.":
                header_row_idx = cell.row
                break
        if header_row_idx:
            break

    if header_row_idx is None:
        raise ValueError(f"Could not find 'X-Check No.' header in sheet '{sheet_name}'")

    header_row = ws[header_row_idx]
    for cell in header_row:
        if cell.value is not None:
            col_index[str(cell.value).strip()] = cell.column

    missing = [c for c in CONDITION_COLS if c not in col_index]
    if missing:
        warnings.append(f"Condition columns not found in sheet and will be skipped: {missing}")

    active_cols = [c for c in CONDITION_COLS if c in col_index]

    if "X-Check No." not in col_index:
        raise ValueError("'X-Check No.' column not found in header row")

    xcno_col = col_index["X-Check No."]

    # ------------------------------------------------------------------
    # Collect (x_check_no, condition_col_name, cell_value) tuples
    # ------------------------------------------------------------------
    # Structure: { xcheck_no: { condition_col: value_or_None } }
    collected: dict[str, dict[str, str | None]] = {}

    def _record(xcheck_no, cond_col, value):
        xcheck_no = str(xcheck_no).strip()
        if not xcheck_no:
            return
        if xcheck_no not in collected:
            collected[xcheck_no] = {c: None for c in active_cols}
        if value is not None and str(value).strip():
            collected[xcheck_no][cond_col] = str(value).strip()

    for row in ws.iter_rows(min_row=header_row_idx + 1):
        row_dict = {cell.column: cell for cell in row}

        xcno_cell = row_dict.get(xcno_col)
        if xcno_cell is None or xcno_cell.value is None:
            continue

        xcheck_no = str(xcno_cell.value).strip()
        if not xcheck_no:
            continue

        for cond_col in active_cols:
            cond_cell = row_dict.get(col_index[cond_col])
            if cond_cell is None:
                continue

            cell_val = cond_cell.value
            if cell_val is None or not str(cell_val).strip():
                continue  # blank condition cell — never collect regardless of colour

            if process_only_differences:
                # Checked: only yellow or green condition cells (changed/new rows)
                if _is_yellow(cond_cell, wb) or _is_green(cond_cell, wb):
                    _record(xcheck_no, cond_col, cell_val)
            else:
                # Unchecked: every non-blank condition cell regardless of colour
                _record(xcheck_no, cond_col, cell_val)

    wb.close()

    if not collected:
        warnings.append("No yellow or green condition cells found — output will be empty.")

    # ------------------------------------------------------------------
    # Build working DataFrame
    # ------------------------------------------------------------------
    rows = []
    for xcheck_no, cond_vals in sorted(collected.items()):
        row: dict[str, object] = {"X-Check No.": xcheck_no}
        for cond_col in active_cols:
            val = cond_vals.get(cond_col)
            row[cond_col] = val if val is not None else ""
        # Concatenated columns: "XCheck|value" or "" if no value
        for cond_col in active_cols:
            val = cond_vals.get(cond_col)
            concat_col = cond_col + " (Concat)"
            if val and str(val).strip():
                row[concat_col] = f"{xcheck_no}|{val}"
            else:
                row[concat_col] = ""
        rows.append(row)

    # Ensure all 5 condition cols present even if some were missing from file
    all_cond_cols = CONDITION_COLS
    all_concat_cols = [c + " (Concat)" for c in all_cond_cols]
    working_df = pd.DataFrame(rows, columns=["X-Check No."] + all_cond_cols + all_concat_cols)

    return working_df, warnings
