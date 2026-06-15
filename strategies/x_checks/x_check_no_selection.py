"""
X-Check No Selection — produces the list of X-Check Nos in scope for a FIP run.

Pipeline (operates on a copy of the EBX 'cross checks all' sheet, no mutation
of the original DataFrame):
  1. Drop rows where Status = INACTIVE.
  2. From the survivors, keep only rows where 'Type of Change' is non-blank.
  3. Take the unique X-Check Nos from those rows  → in-scope set.
  4. From the FULL surviving (ACTIVE) data, drop any X-Check whose row(s)
     contain Exclude Z-Core = X.
  5. From the FULL surviving (ACTIVE) data, drop any X-Check whose Category
     cell is filled with standard Excel yellow (#FFFF00). Detected via openpyxl.
  6. Return the X-Check Nos remaining after steps 4 and 5 are applied to the
     in-scope set.

All column lookups are case-insensitive on the column header.
"""
from __future__ import annotations

import os
from typing import Iterable

import openpyxl
import pandas as pd


YELLOW_RGB = "FFFF00"   # openpyxl returns colours as 8-hex (alpha+RGB) — compare RGB only


def _resolve_col(df: pd.DataFrame, name: str) -> str | None:
    """Returns the actual column header in df matching `name` case-insensitively, or None."""
    target = name.casefold()
    for c in df.columns:
        if str(c).casefold() == target:
            return c
    return None


def select_x_check_nos(
    df: pd.DataFrame,
    filepath: str,
    sheet_name: str,
) -> list[str]:
    """
    Runs the full selection pipeline and returns the unique X-Check Nos in scope,
    in order of first appearance in `df`. See module docstring for the rules.
    """
    df_copy = df.copy()

    col_status   = _resolve_col(df_copy, "Status")
    col_type     = _resolve_col(df_copy, "Type of Change")
    col_xcheck   = _resolve_col(df_copy, "X-Check No.")
    col_excl_zc  = _resolve_col(df_copy, "Exclude Z-Core")
    col_category = _resolve_col(df_copy, "Category")

    if col_xcheck is None:
        return []

    # Step 1 — drop INACTIVE rows
    if col_status is not None:
        status = df_copy[col_status].astype(str).str.strip().str.upper()
        df_copy = df_copy[status != "INACTIVE"]

    # Step 2 — only rows with a non-blank Type of Change become candidates
    if col_type is None:
        in_scope_xchecks = _ordered_unique_str(df_copy[col_xcheck])
    else:
        toc = df_copy[col_type].astype(str).str.strip()
        candidates = df_copy[~toc.isin(("", "nan", "None"))]
        in_scope_xchecks = _ordered_unique_str(candidates[col_xcheck])

    # Step 4 — drop any X-Check that has an 'Exclude Z-Core' = X anywhere
    if col_excl_zc is not None:
        zc = df_copy[col_excl_zc].astype(str).str.strip().str.upper()
        excluded = set(_ordered_unique_str(df_copy.loc[zc == "X", col_xcheck]))
        in_scope_xchecks = [x for x in in_scope_xchecks if x not in excluded]

    # Step 5 — drop any X-Check whose Category cell is yellow (#FFFF00)
    if col_category is not None:
        yellow_xchecks = _x_checks_with_yellow_category(
            filepath, sheet_name, df_copy, col_category, col_xcheck
        )
        in_scope_xchecks = [x for x in in_scope_xchecks if x not in yellow_xchecks]

    return in_scope_xchecks


def _ordered_unique_str(series: pd.Series) -> list[str]:
    """Stringify, strip, drop nan/blank, dedupe preserving order of first occurrence."""
    seen: set = set()
    out: list[str] = []
    for raw in series.tolist():
        v = str(raw).strip()
        if v in ("", "nan", "None") or v in seen:
            continue
        seen.add(v)
        out.append(v)
    return out


def _x_checks_with_yellow_category(
    filepath: str,
    sheet_name: str,
    df_active: pd.DataFrame,
    col_category: str,
    col_xcheck: str,
) -> set[str]:
    """
    Returns the set of X-Check Nos whose Category cell is filled with standard
    Excel yellow (#FFFF00) on the original sheet.

    `df_active` must be the post-INACTIVE-drop DataFrame; its index values are
    used as 0-based row offsets into the Excel data (i.e. Excel row = index+2,
    accounting for the 1-based header row).
    """
    if os.path.splitext(filepath)[1].lower() != ".xlsx":
        return set()

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)
        ws = wb[sheet_name]
    except Exception:
        return set()

    # Find the Excel column index for Category from the header row
    cat_excel_col = None
    for cell in ws[1]:
        if cell.value is not None and str(cell.value).casefold() == col_category.casefold():
            cat_excel_col = cell.column
            break
    if cat_excel_col is None:
        wb.close()
        return set()

    yellow: set[str] = set()
    for df_idx, x_val in zip(df_active.index, df_active[col_xcheck].astype(str)):
        excel_row = df_idx + 2
        cell = ws.cell(row=excel_row, column=cat_excel_col)
        if _is_yellow(cell):
            v = x_val.strip()
            if v not in ("", "nan", "None"):
                yellow.add(v)
    wb.close()
    return yellow


def _is_yellow(cell) -> bool:
    """True if the cell is filled with standard Excel yellow #FFFF00."""
    fill = cell.fill
    if fill is None or fill.fill_type in (None, "none"):
        return False
    fg = fill.fgColor
    if fg is None or fg.type != "rgb":
        return False
    rgb = str(fg.rgb).upper()
    # openpyxl returns 8-hex (AARRGGBB) or 6-hex (RRGGBB); take the trailing 6
    return rgb[-6:] == YELLOW_RGB
