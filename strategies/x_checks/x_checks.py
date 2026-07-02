import os
import pandas as pd
from strategies.base_strategy import BaseStrategy, UploadTaskConfig
from .ebx_extraction import extract_ebx
from .fip_extraction import extract_fip
from .compare import compare
from .x_check_no_selection import select_x_check_nos


class XChecks(BaseStrategy):

    def __init__(self, config: UploadTaskConfig):
        super().__init__(config)

    def process(self, loaded_files: dict, files: dict):
        self.log_step(self.log, "System", "Starting X-Checks processing", 0)

        # 0. X-Check No Selection — write the list of changed X-Check Nos to a .txt
        #    file the user can paste into FIP. Only runs when "Process only differences"
        #    is enabled, since otherwise every X-Check would be listed.
        if files.get("process_only_differences", False):
            self._write_x_check_no_list(loaded_files, files)

        # 1. Load GCoA QU accounts (optional)
        qu_accounts: set = set()
        gcoa_df = loaded_files.get("GCoA Publication File")
        if gcoa_df is not None:
            qu_mask = gcoa_df["Data type"].astype(str).str.strip() == "QU"
            qu_accounts = set(gcoa_df.loc[qu_mask, "Account ID"].astype(str).str.strip())
            self.log_step(self.log, "GCoA", "QU accounts loaded", len(qu_accounts))
        else:
            self.log_step(self.log, "GCoA", "No GCoA file provided — QU_YTD substitution skipped", 0)

        # 2. Extract EBX
        self.log_step(self.log, "EBX", "Extracting from publication file...", 0)
        ebx_results = extract_ebx(
            loaded_files["X-Checks Publication File"],
            qu_accounts=qu_accounts,
            apply_version_spanning=files.get("apply_version_spanning", False),
            apply_prior_year_balance=files.get("apply_prior_year_balance", False),
        )
        self.log_step(self.log, "EBX", "X-Checks extracted", len(ebx_results))

        # 3. Extract FIP — x_check_list from all unique X-Check No. values in the raw file,
        #    matching old FIPExtraction.py which used the EBX file directly rather than
        #    extraction results (ensures X-Checks with no Account No. rows are still searched in FIP)
        ebx_df = loaded_files["X-Checks Publication File"]
        if "X-Check No." not in ebx_df.columns:
            self.log_step(self.log, "EBX", "Required column 'X-Check No.' not found — aborting", 0)
            return
        x_check_list = sorted(set(
            str(x) for x in ebx_df["X-Check No."].tolist()
            if str(x) not in ("nan", "", "NaN", "None")
        ))
        self.log_step(self.log, "FIP", "Extracting from FIP text...", len(x_check_list))
        fip_results = extract_fip(loaded_files["FIP File"], x_check_list)
        self.log_step(self.log, "FIP", "X-Checks extracted", len(fip_results))

        # 4. Compare and sort — matches old Compare_Files.py "All Data" sheet sort order
        self.log_step(self.log, "Comparison", "Comparing EBX and FIP...", 0)
        comparison_rows = compare(ebx_results, fip_results)
        if not comparison_rows:
            self.log_step(self.log, "Comparison", "No X-Checks to compare — aborting output", 0)
            return
        df_comparison = pd.DataFrame(comparison_rows)
        df_comparison = df_comparison.sort_values("X-Check No.").reset_index(drop=True)

        # 5. Apply known exceptions if file was provided
        _XC_FINGERPRINT = [
            "X-Check No.", "EBX Formula", "FIP Formula",
            "EBX Formula (Excl)", "FIP Formula (Excl)",
            "EBX Variables", "FIP Variables", "FIP Variable (Builder)",
        ]
        exc_path = files["files"].get("Known Exception List")
        if exc_path:
            try:
                known_exceptions = self._load_known_exceptions(
                    exc_path, sheet_name="X-Checks", fingerprint_columns=_XC_FINGERPRINT
                )
            except (ValueError, KeyError) as e:
                self.log_step(self.log, "Exceptions", f"Known Exception List is invalid — aborting: {e}", 0)
                return
            self.log_step(self.log, "Exceptions", "Known exceptions loaded", len(known_exceptions))

            def _xc_key(row):
                parts = [str(row[c]).strip() if pd.notna(row.get(c)) else "" for c in _XC_FINGERPRINT]
                return tuple(parts)

            df_comparison["Known Exception"] = df_comparison.apply(
                lambda row: known_exceptions.get(_xc_key(row), ""), axis=1
            )
        else:
            self.log_step(self.log, "Exceptions", "No Known Exception List provided — skipping", 0)

        # 6. Write Excel output — no summary, headers at row 1
        self.write_excel_output(
            output_path=self.build_output_path(
                files["output_directory"], "Comparison", files["timestamp"]
            ),
            sheets={"Comparison": df_comparison},
            log=self.log,
        )

    def _write_x_check_no_list(self, loaded_files: dict, files: dict) -> None:
        """
        Runs the full Cross Checks All selection pipeline (Status / Type of Change /
        Exclude Z-Core / yellow Category) and writes the surviving X-Check Nos to
        <timestamp>_X-Check_Nos.txt in the output directory, one per line.

        Operates on a copy of the EBX DataFrame; the loaded df is unchanged so
        downstream comparison still runs against the full dataset.
        """
        ebx_df = loaded_files.get("X-Checks Publication File")
        if ebx_df is None:
            self.log_step(self.log, "X-Check No Selection", "Skipped — EBX not loaded", 0)
            return

        ebx_path = files["files"].get("X-Checks Publication File")
        ebx_sheet = files["sheet_names"].get("X-Checks Publication File")
        if not ebx_path or not ebx_sheet:
            self.log_step(self.log, "X-Check No Selection",
                          "Skipped — EBX file path or sheet not provided", 0)
            return

        x_check_nos = select_x_check_nos(ebx_df, ebx_path, ebx_sheet)
        if not x_check_nos:
            self.log_step(self.log, "X-Check No Selection",
                          "No X-Check Nos in scope after pipeline", 0)
            return

        out_path = self.build_output_path(
            files["output_directory"], "X-Check_Nos", files["timestamp"], extension=".txt"
        )
        with open(out_path, "w", encoding="utf-8") as f:
            f.write("\n".join(x_check_nos))
        self.log_step(self.log, "X-Check No Selection",
                      f"Wrote {os.path.basename(out_path)}", len(x_check_nos),
                      notes=out_path)

    def apply_output_formatting(self, workbook):
        from openpyxl.styles import PatternFill, Font

        if "Comparison" not in workbook.sheetnames:
            return

        green_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font  = Font(color="276221")
        red_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font    = Font(color="9C0006")
        orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        orange_font = Font(color="9C6500")

        # Zurich brand: Light Blue #91BFE3 fill, Dark Blue #23366F font
        blue_fill = PatternFill(start_color="91BFE3", end_color="91BFE3", fill_type="solid")
        blue_font = Font(color="23366F")

        ws = workbook["Comparison"]
        for col in ("Formula Match", "Formula Match (Excl)", "Variables Match", "Variables Match (Builder)"):
            self.apply_conditional_formatting(
                worksheet=ws,
                column_name=col,
                rules={
                    "Match":                      (green_fill,  green_font),
                    "MisMatch":                   (red_fill,    red_font),
                    "Not Found":                  (orange_fill, orange_font),
                    "Mismatch - Known Exception": (blue_fill,   blue_font),
                }
            )

        # Highlight known exceptions in blue — applied to every non-blank cell in the column
        header_values = [cell.value for cell in ws[1]]
        if "Known Exception" in header_values:
            col_idx = header_values.index("Known Exception") + 1
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                if cell.value and str(cell.value).strip() not in ("", "nan"):
                    cell.fill = blue_fill
                    cell.font = blue_font
