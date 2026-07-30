"""
Grouping By strategy — compares X-Check grouping data from the publication
file against the FIP ZQ9_VALFLDGR extract.

Output workbook (6 sheets + Processing Log):
  Mapping File    — the CSV mapping loaded as a DataFrame
  FIP - Original  — raw FIP data as loaded
  FIP - Processed — FIP with EBX Item mapped + Key column built
  EBX - Original  — raw EBX (publication) data as loaded
  EBX - Processed — EBX filtered, split, stacked, with Key column
  Compare         — EBX Key vs FIP lookup: Matched / Not in FIP
  Processing Log  — standard BaseStrategy log (always written automatically)
"""

import pandas as pd
import openpyxl
from collections import OrderedDict

from strategies.base_strategy import BaseStrategy
from task_configs import GROUPING_BY_UPLOAD_CONFIG

# Colours that indicate a row is in scope for "Process only differences".
# Yellow = Changed, Green = New x-check or association.
_DIFF_YELLOW_RGBS = {"FFFFFF00", "FFFFC000", "FFFFEB9C"}
_DIFF_GREEN_RGBS  = {"FF92D050", "FF00B050", "FFC6EFCE", "FF70AD47", "FF548235"}


def _cell_rgb(cell) -> str | None:
    """Return 8-char ARGB hex for a cell's solid fill, or None if no fill."""
    fg = getattr(getattr(cell, "fill", None), "fgColor", None)
    if fg is None:
        return None
    if fg.type == "rgb":
        rgb = str(fg.rgb).upper()
        return rgb if rgb and rgb != "00000000" else None
    if fg.type == "indexed":
        _Y = {13, 27, 36}
        _G = {10, 17, 35, 42, 50}
        if fg.indexed in _Y:
            return "FFFFFF00"
        if fg.indexed in _G:
            return "FF92D050"
    return None


class GroupingBy(BaseStrategy):

    def process(self, loaded_files: dict, files: dict) -> bool:

        df_mapping_file, df_fip_original, df_fip_processed = self._process_fip(loaded_files)
        if df_fip_original is None:
            self.log_step(self.log, "Grouping By", "FIP processing failed — aborting.", 0)
            return False

        df_ebx_original, df_ebx_processed = self._process_ebx(loaded_files)
        if df_ebx_original is None:
            self.log_step(self.log, "Grouping By", "EBX processing failed — aborting.", 0)
            return False

        df_comparison = self._process_compare(df_fip_processed, df_ebx_processed)

        # When "Process only differences" is on, filter to rows whose Grouping By
        # cell is yellow (Changed) or green (New) in the original publication file.
        if files.get("process_only_differences", False):
            ebx_path  = files["files"].get(GROUPING_BY_UPLOAD_CONFIG.file_fields[1].label)
            ebx_sheet = files["sheet_names"].get(GROUPING_BY_UPLOAD_CONFIG.file_fields[1].label, "cross checks all")
            in_scope = self._diff_in_scope_xchecks(ebx_path, ebx_sheet)
            if in_scope is not None:
                # EBX Key format: "{xcheck}|{grouping_value}" — filter on X-Check part
                df_comparison = df_comparison[
                    df_comparison["EBX Key"].str.split("|").str[0].isin(in_scope)
                ].reset_index(drop=True)
                self.log_step(self.log, "Comparison",
                              f"Filtered to in-scope X-Checks (differences mode)",
                              len(df_comparison))

        # Apply known exceptions (annotation only — Result column unchanged)
        result = self._annotate_known_exceptions(
            df_comparison, files["files"].get("Known Exception List"),
            sheet_name="Grouping By", fingerprint_columns=["EBX Key"]
        )
        if result is False:
            return False
        df_comparison = result

        matched     = (df_comparison["Result"] == "Matched").sum()
        not_matched = (df_comparison["Result"] == "Not in FIP").sum()

        fip_path = files["files"][GROUPING_BY_UPLOAD_CONFIG.file_fields[0].label]
        ebx_path = files["files"][GROUPING_BY_UPLOAD_CONFIG.file_fields[1].label]

        output_path = self.build_output_path(
            files["output_directory"],
            "Grouping By Comparison",
            files["timestamp"],
        )

        sheets = OrderedDict([
            ("Mapping File",    df_mapping_file),
            ("FIP - Original",  df_fip_original),
            ("FIP - Processed", df_fip_processed),
            ("EBX - Original",  df_ebx_original),
            ("EBX - Processed", df_ebx_processed),
            ("Comparison",         df_comparison),
        ])

        summaries = {
            "FIP - Original":  OrderedDict([("Source filename:", fip_path), ("Number of rows:", len(df_fip_original))]),
            "FIP - Processed": OrderedDict([("Source filename:", fip_path), ("Number of rows:", len(df_fip_processed))]),
            "EBX - Original":  OrderedDict([("Source filename:", ebx_path), ("Number of rows:", len(df_ebx_original))]),
            "EBX - Processed": OrderedDict([("Source filename:", ebx_path), ("Number of rows:", len(df_ebx_processed))]),
            "Comparison":         OrderedDict([("Number of rows:", len(df_comparison)), ("Matched:", matched), ("Not in FIP:", not_matched)]),
        }

        self.write_excel_output(output_path, sheets, self.log, summaries=summaries)
        return True

    # ------------------------------------------------------------------
    # Differences-mode helper
    # ------------------------------------------------------------------

    def _diff_in_scope_xchecks(self, filepath: str | None, sheet_name: str) -> set | None:
        """
        Returns the set of X-Check Nos whose 'Grouping By' cell is yellow or green
        (i.e. Changed or New) in the publication file.  Returns None on any error
        so the caller can skip filtering gracefully.
        """
        if not filepath:
            return None
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return None
            ws = wb[sheet_name]

            # Locate header row and column indices
            xc_col = gb_col = None
            for row in ws.iter_rows(min_row=1, max_row=10):
                for cell in row:
                    v = str(cell.value).strip() if cell.value else ""
                    if v.casefold() == "x-check no.":
                        xc_col = cell.column
                    elif v.casefold() == "grouping by":
                        gb_col = cell.column
                if xc_col and gb_col:
                    header_row = row[0].row
                    break
            else:
                wb.close()
                return None

            in_scope: set = set()
            for row in ws.iter_rows(min_row=header_row + 1):
                row_dict = {cell.column: cell for cell in row}
                xc_cell = row_dict.get(xc_col)
                gb_cell = row_dict.get(gb_col)
                if not xc_cell or not xc_cell.value:
                    continue
                if not gb_cell:
                    continue
                rgb = _cell_rgb(gb_cell)
                if rgb and (rgb in _DIFF_YELLOW_RGBS or rgb in _DIFF_GREEN_RGBS):
                    in_scope.add(str(xc_cell.value).strip())

            wb.close()
            return in_scope
        except Exception:
            return None

    # ------------------------------------------------------------------
    # FIP processing
    # ------------------------------------------------------------------

    def apply_output_formatting(self, workbook):
        if "Comparison" not in workbook.sheetnames:
            return
        self.apply_conditional_formatting(
            worksheet=workbook["Comparison"],
            column_name="Result",
            rules={
                "Matched":    (self.FILL_GREEN,  self.FONT_GREEN),
                "Not in FIP": (self.FILL_ORANGE, self.FONT_ORANGE),
            },
        )

    def _process_fip(self, loaded_files) -> tuple:
        try:
            self.log_step(self.log, "Mapping File", "Started processing", 0)
            mapping_content = loaded_files[GROUPING_BY_UPLOAD_CONFIG.file_fields[2].label]
            self.log_step(self.log, "Mapping File", "Loaded", len(mapping_content.splitlines()), "Including header row")

            mapping_dict = {}
            for line in mapping_content.splitlines()[1:]:
                line = line.strip()
                if not line:
                    continue
                parts = line.split(",", maxsplit=1)
                if len(parts) == 2:
                    mapping_dict[parts[0].strip()] = parts[1].strip()
            self.log_step(self.log, "Mapping File", "Mapping dictionary created", len(mapping_dict))

            df_mapping_file = pd.DataFrame(
                [line.split(",", maxsplit=1) for line in mapping_content.splitlines()[1:] if line.strip()],
                columns=["FIP Data", "EBX item"]
            )
            self.log_step(self.log, "Mapping File", "Finished processing", len(mapping_dict))

            self.log_step(self.log, "FIP", "Started processing", 0)
            df_original = loaded_files[GROUPING_BY_UPLOAD_CONFIG.file_fields[0].label].copy()
            df_fip = df_original.copy()
            self.log_step(self.log, "FIP", "Original file loaded", len(df_original))

            df_fip["EBX Item"] = df_fip["Field name"].map(mapping_dict)
            self.log_step(self.log, "FIP", "Mapped 'Field name' to 'EBX Item'", len(df_fip))

            df_fip = df_fip[
                df_fip["EBX Item"].notna() &
                (df_fip["EBX Item"].str.strip() != "") &
                (df_fip["EBX Item"].str.strip().str.lower() != "ignore")
            ]
            df_fip = df_fip[
                df_fip["ValidRule"].notna() &
                (df_fip["ValidRule"].str.strip() != "")
            ]
            self.log_step(self.log, "FIP", "Removed unmapped and blank rows", len(df_fip))

            df_fip["Key"] = df_fip.apply(
                lambda row: f"{row['ValidRule']}|{row['EBX Item']}"
                if pd.notna(row["ValidRule"]) and str(row["ValidRule"]).strip() != ""
                else "",
                axis=1
            )
            self.log_step(self.log, "FIP", "Finished processing", len(df_fip))
            return df_mapping_file, df_original, df_fip

        except Exception as exc:
            import traceback
            self.log_step(self.log, "FIP", f"Exception: {exc}", 0)
            self.log_step(self.log, "FIP", traceback.format_exc(), 0)
            return None, None, None

    # ------------------------------------------------------------------
    # EBX processing
    # ------------------------------------------------------------------

    def _process_ebx(self, loaded_files) -> tuple:
        try:
            self.log_step(self.log, "EBX", "Started processing", 0)

            df_ebx_original = loaded_files[GROUPING_BY_UPLOAD_CONFIG.file_fields[1].label].copy()
            df_ebx = df_ebx_original.copy()
            self.log_step(self.log, "EBX", "Original file loaded", len(df_ebx_original))

            df_ebx = df_ebx[df_ebx["Grouping By"].notna() & (df_ebx["Grouping By"].str.strip() != "")]
            self.log_step(self.log, "EBX", "Filtered to rows with 'Grouping By'", len(df_ebx))

            df_ebx = df_ebx.drop_duplicates(subset=["X-Check No."], keep="first").reset_index(drop=True)
            self.log_step(self.log, "EBX", "Removed duplicate X-Check No. rows", len(df_ebx))

            split_cols = df_ebx["Grouping By"].str.split(",", expand=True)
            split_cols.columns = [f"Grouping By {i + 1}" for i in range(split_cols.shape[1])]
            split_cols = split_cols.apply(lambda col: col.str.strip())
            self.log_step(self.log, "EBX", f"Split 'Grouping By' into {split_cols.shape[1]} columns", split_cols.notna().sum().sum())

            col_position = df_ebx.columns.get_loc("Grouping By")
            df_ebx = df_ebx.drop(columns=["Grouping By"])
            for col in reversed(split_cols.columns.tolist()):
                df_ebx.insert(col_position, col, split_cols[col])
            self.log_step(self.log, "EBX", "Replaced 'Grouping By' with split columns", len(df_ebx))

            ref_col = df_ebx["Reference  X-Check (Condition)"].astype(str).str.strip()
            df_ebx["_base_key"] = ref_col.where(
                df_ebx["Reference  X-Check (Condition)"].notna() &
                (ref_col != "") & (ref_col.str.lower() != "nan"),
                other=df_ebx["X-Check No."].astype(str).str.strip()
            )
            self.log_step(self.log, "EBX", "Constructed base key column", df_ebx["_base_key"].notna().sum())

            for col in [c for c in df_ebx.columns if c.startswith("Grouping By ")]:
                key_col = col.replace("Grouping By ", "Grouping By Key ")
                df_ebx[key_col] = df_ebx.apply(
                    lambda row, c=col: (
                        f"{row['_base_key']}|{str(row[c]).strip()}"
                        if pd.notna(row[c]) and str(row[c]).strip() != ""
                        else ""
                    ),
                    axis=1
                )
            self.log_step(self.log, "EBX", "Constructed 'Grouping By Key n' columns", len(df_ebx))

            df_ebx = df_ebx.drop(columns=["_base_key"])

            key_cols = [c for c in df_ebx.columns if c.startswith("Grouping By Key ")]
            index_cols = [c for c in df_ebx.columns if c not in key_cols]
            stacked = (
                df_ebx.set_index(index_cols)
                .stack()
                .reset_index()
                .rename(columns={0: "Key"})
            )
            level_col = f"level_{len(index_cols)}"
            if level_col in stacked.columns:
                stacked = stacked.drop(columns=[level_col])
            self.log_step(self.log, "EBX", "Stacked key columns into single 'Key' column", len(stacked))

            df_ebx = stacked[stacked["Key"].str.strip() != ""].reset_index(drop=True)
            self.log_step(self.log, "EBX", "Finished processing", len(df_ebx))
            return df_ebx_original, df_ebx

        except Exception as exc:
            import traceback
            self.log_step(self.log, "EBX", f"Exception: {exc}", 0)
            self.log_step(self.log, "EBX", traceback.format_exc(), 0)
            return None, None

    # ------------------------------------------------------------------
    # Compare
    # ------------------------------------------------------------------

    def _process_compare(self, df_fip: pd.DataFrame, df_ebx: pd.DataFrame) -> pd.DataFrame:
        self.log_step(self.log, "Comparison", "Started comparison", 0)

        fip_keys = df_fip[["Key"]].drop_duplicates().copy()
        fip_keys["In FIP"] = True
        self.log_step(self.log, "Comparison", "FIP key lookup built", len(fip_keys))

        ebx_keys = df_ebx[["Key"]].drop_duplicates().copy()
        self.log_step(self.log, "Comparison", "EBX keys extracted", len(ebx_keys))

        df_compare = ebx_keys.merge(fip_keys, on="Key", how="left")
        df_compare["In FIP"] = df_compare["In FIP"].fillna(False)
        df_compare["Result"] = df_compare["In FIP"].map({True: "Matched", False: "Not in FIP"})

        matched     = df_compare["Result"].eq("Matched").sum()
        not_matched = df_compare["Result"].eq("Not in FIP").sum()
        self.log_step(self.log, "Comparison", f"Matched: {matched} | Not in FIP: {not_matched}", len(df_compare))

        df_compare = (
            df_compare
            .drop(columns=["In FIP"])
            .rename(columns={"Key": "EBX Key"})
            .sort_values("EBX Key")
            .reset_index(drop=True)
        )
        self.log_step(self.log, "Comparison", "Finished comparison", len(df_compare))
        return df_compare
