import pandas as pd
import openpyxl
import os
import re
import shutil
import threading
from abc import ABC, abstractmethod
from tkinter import messagebox
from typing import Optional
from file_upload_config import UploadTaskConfig
from datetime import datetime
from config import OUTPUT_TEMPLATE
from openpyxl.styles import PatternFill, Font

class BaseStrategy(ABC):
    """
    Handles everything that is ALWAYS the same:
    - Loading files into memory
    - Calling the use-case-specific processing
    - Writing Excel output
    - Logging processing steps
    """

    # Default sensitivity label applied to every workbook this base writes.
    DEFAULT_SENSITIVITY_LEVEL = "Internal_Use_Only"

    # Shared Zurich-brand colour palette used for comparison output formatting.
    # All strategies reference these constants — define once, inherit everywhere.
    FILL_GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    FONT_GREEN  = Font(color="276221")
    FILL_RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    FONT_RED    = Font(color="9C0006")
    FILL_ORANGE = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    FONT_ORANGE = Font(color="9C6500")
    FILL_BLUE   = PatternFill(start_color="91BFE3", end_color="91BFE3", fill_type="solid")
    FONT_BLUE   = Font(color="23366F")

    def __init__(self, config: UploadTaskConfig):
        self.config = config
        self._progress_dialog = None          # Set via set_progress_dialog()
        self._stop_event: threading.Event | None = None
        self._sensitivity_labeler = None      # lazy ExcelLabeler, created on first write

    def execute(self, files: dict):
        """
        Entry point called by main.py.
        Loads all files then hands off to the subclass.
        """
        from exceptions import FileLoadError, SheetNotFoundError, MissingColumnsError, UnsupportedFileTypeError
        from version import __version__

        self.log = []  # Initialised here — available to all strategies via self.log
        self.process_only_differences = files.get("process_only_differences", False)
        self.log_step(self.log, "System", f"X-Check Application v{__version__}", 0)

        # Store process flag on instance so _load_files can access it
        self.process_only_differences = files.get("process_only_differences", False)

        try:
            try:
                self.log_step(self.log, "System", "Loading files into memory...", 0)
                try:
                    loaded_files = self._load_files(files["files"], files["sheet_names"], self.config.file_fields)
                except (FileLoadError, SheetNotFoundError, MissingColumnsError, UnsupportedFileTypeError) as e:
                    self.log_step(self.log, "System", f"Error loading files: {e}", 0)
                    return

                if loaded_files is None:
                    return

                self.log_step(self.log, "System", "Files loaded successfully:", len(loaded_files))
                for label, data in loaded_files.items():
                    self.log_step(self.log, "    " + label, f"Loaded {type(data).__name__}", len(data))

                # Strategies that complete a full run return True. Strategies that
                # bail early (e.g. missing column, no rows to compare) return None
                # or False — propagate that so run_processing surfaces the failure
                # via 'Return to Form' instead of claiming success.
                return bool(self.process(loaded_files, files))
            finally:
                # Always shut down the cached Excel COM session, success or not.
                if self._sensitivity_labeler is not None:
                    try:
                        self._sensitivity_labeler.close()
                    except Exception:
                        pass
                    self._sensitivity_labeler = None

        except StopIteration:
            # User pressed Stop — log it, then return cleanly
            timestamp = __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            print(f"  [{timestamp}] System — Processing halted by user.")
            self.log.append({
                "Timestamp": timestamp,
                "File": "System",
                "Step": "Processing halted by user.",
                "Count": 0,
                "Notes": ""
            })
            if self._progress_dialog is not None:
                self._progress_dialog.append_entry("System", "Processing halted by user. You may now close this window.")

#        self.log_step(self.log, "System", "Loading files into memory...", 0)
#        try:
#            loaded_files = self._load_files(files["files"], files["sheet_names"], self.config.file_fields)
#
#        except (FileLoadError, SheetNotFoundError, MissingColumnsError, UnsupportedFileTypeError) as e:
#            self.log_step(self.log, "System", f"Error loading files: {e}", 0)
#            messagebox.showerror("File Loading Error", str(e))
#            return
#
#        if loaded_files is None:
#            return
#
#        self.log_step(self.log, "System", "Files loaded successfully:", len(loaded_files))
#        for label, data in loaded_files.items():
#            self.log_step(self.log, "    " + label, f"Loaded {type(data).__name__}", len(data))
#
#        self.process(loaded_files, files)

    # -------------------------------------------------------------------------
    # Excel output utilities — available to all strategies
    # -------------------------------------------------------------------------

    def build_output_path(self, output_directory: str, label: str, timestamp: str,
                          extension: str = ".xlsx") -> str:
        """
        Builds a safe, timestamped output file path from a label.
        Replaces characters that are invalid in Windows filenames.
        """
        safe_label = re.sub(r'[<>:"/\\|?*()]', '_', label)
        filename = f"{timestamp}_{safe_label}{extension}"
        return os.path.join(output_directory, filename)

    def set_progress_dialog(self, dialog):
        """
        Called by main.py (debug mode only) to attach the ProgressDialog.
        Also stores a reference to its stop event for checkpoint polling.
        """
        self._progress_dialog = dialog
        self._stop_event = dialog.stop_event

    def log_step(self, log: list, file: str, step: str, count: int, notes: str = ""):
        """
        Appends a timestamped entry to the processing log and prints to console.
        If a ProgressDialog is attached, pushes the entry to it.
        If the stop event has been set, raises StopIteration to unwind processing.
        The stop check happens AFTER the current step is recorded, so the step
        that was already running completes before the halt takes effect.
        """
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ui_timestamp = datetime.now().strftime("%Y%m%d %H%M%S")
        print(f"  [{timestamp}] {file} — {step} ({count})")
        log.append({
            "Timestamp": timestamp,
            "File": file,
            "Step": step,
            "Count": count,
            "Notes": notes
        })

        # Push to UI dialog if attached
        if self._progress_dialog is not None:
            self._progress_dialog.append_entry(file, step, count, notes, timestamp=ui_timestamp)

        # Check stop event AFTER completing this step
        if self._stop_event is not None and self._stop_event.is_set():
            raise StopIteration("Processing stopped by user.")

    def autofit_columns(self, worksheet, max_width: int = 90,  skip_rows: int = 0):
        """Auto-fits all columns in a worksheet to their content width, capped at max_width."""
        for column in worksheet.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                if cell.row <= skip_rows:
                    continue
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, max_width)

    def write_excel_output(self, output_path: str, sheets: dict, log: list, summaries: dict | None = None):
        """
        Writes a dictionary of DataFrames to a single timestamped Excel workbook.
        Copies the pre-labelled template, writes all sheets, auto-fits columns.

        Args:
            output_path:  Full path to the output file
            sheets:       Ordered dict of {sheet_name: DataFrame}
            log:          List of log entry dicts
            summaries:    Optional dict of {label: value} for summary blocks
        """
        # Copy pre-labelled template to output path
        shutil.copy(OUTPUT_TEMPLATE, output_path)
        df_log = pd.DataFrame(log)

        with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:

            # Write all provided sheets
            for sheet_name, df in sheets.items():
                if summaries and sheet_name in summaries:
                    self.write_sheet_with_summary(writer, sheet_name, df, summaries[sheet_name])
                else:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Write log sheet
            df_log.to_excel(writer, sheet_name="Processing Log", index=False)

            # Remove default Sheet1 from template
            if "Sheet1" in writer.book.sheetnames:
                del writer.book["Sheet1"]

            # Auto-fit all sheets
            for sheet_name in writer.book.sheetnames:
                if summaries and sheet_name in summaries:
                    skip = len(summaries[sheet_name]) + 2  # skip summary rows + blank row
                    self.autofit_columns(writer.book[sheet_name], skip_rows=skip)
                else:
                    self.autofit_columns(writer.book[sheet_name])

            self.log_step(self.log, "Output", f"  Sheets in workbook: {writer.book.sheetnames}", 0)

            # Allow subclasses to apply strategy-specific formatting
            self.apply_output_formatting(writer.book)

        self.log_step(self.log, "Output", f"Output written to: {output_path}", 0)

        # Apply MIP sensitivity label via Excel COM. Best-effort — failures
        # are logged but do not abort the strategy.
        self._apply_sensitivity_label(output_path)

    def _apply_sensitivity_label(self, path: str) -> None:
        """Lazy-creates the cached ExcelLabeler and labels `path`. No-op on
        non-Windows or when Excel COM is unavailable."""
        try:
            from .sensitivity import ExcelLabeler
        except Exception as e:
            self.log_step(self.log, "Sensitivity",
                          f"Module import failed: {e}", 0)
            return
        if self._sensitivity_labeler is None:
            self._sensitivity_labeler = ExcelLabeler()
        ok, msg = self._sensitivity_labeler.label_file(
            path, self.DEFAULT_SENSITIVITY_LEVEL
        )
        if ok:
            self.log_step(self.log, "Sensitivity",
                          f"Applied label: {msg}", 0)
        else:
            self.log_step(self.log, "Sensitivity",
                          f"Could not apply label: {msg}", 0)

    def apply_output_formatting(self, workbook):
        """
        Hook for subclasses to apply strategy-specific formatting.
        Override in subclass — default does nothing.
        """
        pass

    # -------------------------------------------------------------------------
    # File loading utilities — available to all strategies
    # -------------------------------------------------------------------------

    def _load_files(self, files: dict, sheet_names: dict, file_fields: list) -> Optional[dict]:
        """
        Reads each file into memory based on its extension.
        """
        from exceptions import FileLoadError, SheetNotFoundError, MissingColumnsError, UnsupportedFileTypeError

        loaded = {}

        column_map = {
            f.label: f.required_columns
            for f in file_fields
        }
        signals_map = {
            f.label: f.header_signals
            for f in file_fields
        }

        for label, path in files.items():
            if path is None:
                continue
            try:
                ext = os.path.splitext(path)[1].lower()

                if ext in (".xlsx", ".xls"):
                    sheet = sheet_names.get(label, "Sheet1")
                    try:
                        excel_file = pd.ExcelFile(path)
                    except Exception as e:
                        raise FileLoadError(
                            f"Could not open '{os.path.basename(path)}' as an Excel file.\n\n"
                            f"{str(e)}"
                        )

                    if sheet not in excel_file.sheet_names:
                        raise SheetNotFoundError(
                            f"Could not find sheet '{sheet}' in '{os.path.basename(path)}'.\n\n"
                            f"Please check the file and sheet name then try again."
                        )

                    header_row = self._detect_header_row(path, sheet, signals_map.get(label))
                    df = pd.read_excel(path, sheet_name=sheet, header=header_row)
                    df = self._select_columns(df, label, column_map.get(label), path)
                    if df is None:
                        return None

                    # NOTE: process_only_differences no longer filters the EBX DataFrame.
                    # Downstream comparison runs on the full dataset; the X-Check No
                    # selection (.txt) is computed in XChecks.process from a copy of df.

                    loaded[label] = df

                elif ext == ".csv":
                    df = pd.read_csv(path)
                    df = self._select_columns(df, label, column_map.get(label), path)
                    if df is None:
                        return None
                    loaded[label] = df

                elif ext == ".txt":
                    with open(path, "r") as f:
                        loaded[label] = f.read()
                else:
                    raise UnsupportedFileTypeError(
                        f"'{os.path.basename(path)}' is not a supported file type.\n\n"
                        f"Please select a csv or txt file."
                    )

            except PermissionError:
                raise FileLoadError(
                    f"'{os.path.basename(path)}' is currently open in another application.\n\n"
                    f"Please close it and try again."
                )

            except ValueError as e:
                # Catch unexpected pandas errors (e.g. malformed CSV, encoding issues)
                raise FileLoadError(
                    f"Error reading '{os.path.basename(path)}':\n\n{str(e)}"
                )
            
        return loaded


    _HEADER_SCAN_ROWS = 6  # how many top rows to inspect when looking for the header

    def _detect_header_row(self, filepath: str, sheet_name: str,
                           header_signals: Optional[list[str]]) -> int:
        """
        Returns the 0-indexed row to pass as `header=` to pd.read_excel.

        If `header_signals` is None or empty, returns 0 (treat row 1 as header).
        Otherwise scans the first _HEADER_SCAN_ROWS rows of the sheet for the
        first row whose cells contain ALL signal names (case-insensitive,
        stripped). If none match, falls back to 0 so the caller still gets a
        DataFrame and the strategy can surface its own 'column not found' error.
        """
        if not header_signals:
            return 0

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            ws = wb[sheet_name]
            wanted = {s.strip().casefold() for s in header_signals}
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=1, max_row=self._HEADER_SCAN_ROWS, values_only=True),
                start=0,
            ):
                cells = {str(v).strip().casefold() for v in row if v is not None}
                if wanted.issubset(cells):
                    wb.close()
                    return row_idx
            wb.close()
        except Exception:
            pass
        return 0


    def _select_columns(self, df: pd.DataFrame, label: str, required_columns: Optional[list[str]], filepath: str) -> Optional[pd.DataFrame]:
        """
        Reduces the DataFrame to only the required columns.
        Returns None if any required column is missing.
        """
        from exceptions import MissingColumnsError

        if required_columns is None:
            return df

        missing = [col for col in required_columns if col not in df.columns]
        if missing:
            raise MissingColumnsError(
                f"'{os.path.basename(filepath)}' is missing the following required column(s):\n\n"
                + "\n".join(f"  • {col}" for col in missing)
                + f"\n\nPlease check the file and try again."
            )
            
        return df[required_columns]
    
    def apply_conditional_formatting(self, worksheet, column_name: str, rules: dict):
        """
        Applies conditional formatting to a named column in a worksheet.

        Args:
            worksheet:    The openpyxl worksheet object
            column_name:  The header name of the column to format
            rules:        Dict of {cell_value: (PatternFill, Font)} or {cell_value: PatternFill}
        """
        from openpyxl.formatting.rule import CellIsRule

        self.log_step(self.log, "Formatting",
                    f"Applying to '{worksheet.title}', column '{column_name}'", 0)

        # Find the column letter by scanning all rows for the header
        target_col = None
        header_row = None
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value == column_name:
                    target_col = cell.column_letter
                    header_row = cell.row
                    break
            if target_col:
                break

        if target_col is None:
            self.log_step(self.log, "Formatting",
                        f"Column '{column_name}' not found — skipping", 0)
            return

        last_row = worksheet.max_row
        data_start = header_row + 1
        target_range = f"{target_col}{data_start}:{target_col}{last_row}"

        for value, formatting in rules.items():
            # Accept either a plain fill or a (fill, font) tuple
            if isinstance(formatting, tuple):
                fill, font = formatting
            else:
                fill, font = formatting, None

            worksheet.conditional_formatting.add(
                target_range,
                CellIsRule(
                    operator="equal",
                    formula=[f'"{value}"'],
                    fill=fill,
                    font=font
                )
            )

        self.log_step(self.log, "Formatting",
                    f"Applied {len(rules)} rule(s) to {target_range}", 0)

    def write_sheet_with_summary(self, writer, sheet_name: str, df: pd.DataFrame, summary: dict):
        """
        Writes a summary block followed by a blank row then the main DataFrame.

        Args:
            writer:       The pd.ExcelWriter instance
            sheet_name:   Name of the sheet to write to
            df:           The main DataFrame
            summary:      Ordered dict of {label: value} for the summary block
        """
        # Write main DataFrame first so the sheet is created, leaving room for summary
        start_row = len(summary) + 2  # +1 for blank row, +1 for header
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)

        # Write summary rows directly into cells
        worksheet = writer.sheets[sheet_name]
        for i, (label, value) in enumerate(summary.items()):
            worksheet.cell(row=i + 1, column=1, value=label)
            worksheet.cell(row=i + 1, column=2, value=value)

    def _load_known_exceptions(
        self,
        path: Optional[str],
        sheet_name: str,
        fingerprint_columns: list,
    ) -> dict:
        """
        Load a Known Exception List sheet and return a lookup dict.

        Keys are either plain strings (single fingerprint column) or tuples
        (multiple columns). Values are the Reason string.

        Row 2 of the sheet is a guidance/example row and is always skipped.
        Falls back to "Known Exceptions" sheet name if `sheet_name` is not found.

        Returns {} if path is None, sheet is empty, or sheet cannot be read.
        Raises ValueError if required columns are missing or rows are incomplete.
        """
        if not path:
            return {}

        required_cols = fingerprint_columns + ["Reason"]

        # Try the strategy-specific sheet name first, fall back for backwards compat
        df = None
        tried = []
        for sn in (sheet_name, "Known Exceptions"):
            if sn in tried:
                continue
            tried.append(sn)
            try:
                df = pd.read_excel(path, sheet_name=sn, skiprows=[1])
                break
            except Exception:
                df = None

        if df is None:
            self.log_step(self.log, "Exceptions",
                          f"Could not read sheet '{sheet_name}' from Known Exception List", 0)
            return {}

        df = df.dropna(how="all")
        if df.empty:
            self.log_step(self.log, "Exceptions", "Known Exception List sheet is empty", 0)
            return {}

        missing_cols = [c for c in required_cols if c not in df.columns]
        if missing_cols:
            raise ValueError(
                f"Known Exception List sheet '{sheet_name}' is missing columns: {missing_cols}"
            )

        exceptions = {}
        invalid_rows = []
        for i, row in df.iterrows():
            reason = str(row["Reason"]).strip()
            if reason in ("", "nan", "NaN", "None"):
                invalid_rows.append(f"row {i + 2} (missing Reason)")
                continue

            key_parts = []
            row_invalid = False
            for col in fingerprint_columns:
                val = str(row[col]).strip() if pd.notna(row[col]) else ""
                if val in ("", "nan", "NaN", "None"):
                    invalid_rows.append(f"row {i + 2} (missing {col})")
                    row_invalid = True
                    break
                key_parts.append(val)
            if row_invalid:
                continue

            key = key_parts[0] if len(key_parts) == 1 else tuple(key_parts)
            exceptions[key] = reason

        if invalid_rows:
            raise ValueError(
                f"Known Exception List has {len(invalid_rows)} incomplete row(s): "
                f"{'; '.join(invalid_rows)}"
            )

        return exceptions

    def _annotate_known_exceptions(
        self,
        df: "pd.DataFrame",
        exc_path: Optional[str],
        sheet_name: str,
        fingerprint_columns: list,
    ) -> "pd.DataFrame | bool":
        """
        Load the Known Exception List and annotate `df` with a 'Known Exception'
        column containing the Reason text where a row's fingerprint matches.

        Returns the annotated DataFrame, or False if the file is invalid (caller
        should abort and return False). Returns df unchanged if exc_path is None
        or the sheet is empty.
        """
        try:
            known_exceptions = self._load_known_exceptions(
                exc_path, sheet_name=sheet_name, fingerprint_columns=fingerprint_columns
            )
        except (ValueError, KeyError) as e:
            self.log_step(self.log, "Exceptions",
                          f"Known Exception List is invalid — aborting: {e}", 0)
            return False

        if known_exceptions:
            self.log_step(self.log, "Exceptions", "Known exceptions loaded",
                          len(known_exceptions))

            def _key(row):
                parts = [str(row[c]).strip() if pd.notna(row.get(c)) else ""
                         for c in fingerprint_columns]
                return parts[0] if len(parts) == 1 else tuple(parts)

            df["Known Exception"] = df.apply(
                lambda row: known_exceptions.get(_key(row), ""), axis=1
            )
        elif exc_path:
            self.log_step(self.log, "Exceptions",
                          "Known Exception List provided but empty — skipping", 0)
        else:
            self.log_step(self.log, "Exceptions",
                          "No Known Exception List provided — skipping", 0)

        return df

    @abstractmethod
    def process(self, loaded_files: dict, files: dict):
        """
        Subclasses implement THIS — not execute().
        By the time this is called, all files are already in memory.
        Access output_directory via files["output_directory"].
        Access timestamp via files["timestamp"].        
        """
        pass
