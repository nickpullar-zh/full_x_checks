"""
Tests for the v0.4 X-Check No Selection pipeline (select_x_check_nos)
and the .txt writer in XChecks._write_x_check_no_list.
"""
import os
import tempfile

import pandas as pd
import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from file_upload_config import UploadTaskConfig
from strategies.x_checks.x_check_no_selection import (
    _resolve_col,
    _ordered_unique_str,
    select_x_check_nos,
)
from strategies.x_checks.x_checks import XChecks


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_xlsx(tmp: str, headers: list[str], rows: list[list], yellow_cells=None) -> str:
    """
    Writes a single-sheet workbook ('cross checks all') and returns its path.
    yellow_cells: list of (excel_row, excel_col) tuples to fill #FFFF00.
    """
    path = os.path.join(tmp, "ebx.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "cross checks all"
    ws.append(headers)
    for r in rows:
        ws.append(r)
    if yellow_cells:
        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for excel_row, excel_col in yellow_cells:
            ws.cell(row=excel_row, column=excel_col).fill = yellow
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# _resolve_col — case-insensitive column lookup
# ---------------------------------------------------------------------------

def test_resolve_col_exact_match():
    df = pd.DataFrame(columns=["Status", "X-Check No."])
    assert _resolve_col(df, "Status") == "Status"


def test_resolve_col_different_case():
    df = pd.DataFrame(columns=["status", "x-check no."])
    assert _resolve_col(df, "Status") == "status"
    assert _resolve_col(df, "X-Check No.") == "x-check no."


def test_resolve_col_missing_returns_none():
    df = pd.DataFrame(columns=["A", "B"])
    assert _resolve_col(df, "Status") is None


# ---------------------------------------------------------------------------
# _ordered_unique_str
# ---------------------------------------------------------------------------

def test_ordered_unique_preserves_first_occurrence():
    s = pd.Series(["B2", "A1", "B2", "C3", "A1"])
    assert _ordered_unique_str(s) == ["B2", "A1", "C3"]


def test_ordered_unique_drops_blanks_and_nan():
    s = pd.Series(["", "A1", "nan", "B2", "None", " "])
    assert _ordered_unique_str(s) == ["A1", "B2"]


# ---------------------------------------------------------------------------
# select_x_check_nos — the full pipeline
# ---------------------------------------------------------------------------

def test_pipeline_drops_inactive_rows(tmp_path):
    headers = ["X-Check No.", "Status", "Type of Change", "Exclude Z-Core", "Category"]
    rows = [
        ["A1", "ACTIVE",   "Modified", "", ""],
        ["B2", "INACTIVE", "Modified", "", ""],   # dropped (Status)
        ["C3", "ACTIVE",   "New",      "", ""],
    ]
    path = _write_xlsx(str(tmp_path), headers, rows)
    df = pd.read_excel(path, sheet_name="cross checks all")
    assert select_x_check_nos(df, path, "cross checks all") == ["A1", "C3"]


def test_pipeline_drops_blank_type_of_change(tmp_path):
    headers = ["X-Check No.", "Status", "Type of Change", "Exclude Z-Core", "Category"]
    rows = [
        ["A1", "ACTIVE", "",         "", ""],   # blank Type of Change → not in scope
        ["B2", "ACTIVE", "Modified", "", ""],
        ["C3", "ACTIVE", "New",      "", ""],
    ]
    path = _write_xlsx(str(tmp_path), headers, rows)
    df = pd.read_excel(path, sheet_name="cross checks all")
    assert select_x_check_nos(df, path, "cross checks all") == ["B2", "C3"]


def test_pipeline_drops_exclude_z_core_x(tmp_path):
    headers = ["X-Check No.", "Status", "Type of Change", "Exclude Z-Core", "Category"]
    rows = [
        ["A1", "ACTIVE", "Modified", "X", ""],   # excluded
        ["B2", "ACTIVE", "Modified", "",  ""],
        ["C3", "ACTIVE", "New",      "x", ""],   # case-insensitive — also excluded
    ]
    path = _write_xlsx(str(tmp_path), headers, rows)
    df = pd.read_excel(path, sheet_name="cross checks all")
    assert select_x_check_nos(df, path, "cross checks all") == ["B2"]


def test_pipeline_drops_yellow_category(tmp_path):
    headers = ["X-Check No.", "Status", "Type of Change", "Exclude Z-Core", "Category"]
    rows = [
        ["A1", "ACTIVE", "Modified", "", "Cat 1"],
        ["B2", "ACTIVE", "Modified", "", "Cat 2"],   # this Category cell gets yellow
        ["C3", "ACTIVE", "New",      "", "Cat 3"],
    ]
    # Excel row 3 = the B2 row (header is row 1). Category is column 5.
    path = _write_xlsx(str(tmp_path), headers, rows, yellow_cells=[(3, 5)])
    df = pd.read_excel(path, sheet_name="cross checks all")
    assert select_x_check_nos(df, path, "cross checks all") == ["A1", "C3"]


def test_pipeline_case_insensitive_column_names(tmp_path):
    # Headers given in mixed case — pipeline must still find them
    headers = ["x-check no.", "STATUS", "type of change", "exclude z-core", "category"]
    rows = [
        ["A1", "active",   "modified", "",  ""],
        ["B2", "INACTIVE", "modified", "",  ""],
        ["C3", "ACTIVE",   "new",      "X", ""],
    ]
    path = _write_xlsx(str(tmp_path), headers, rows)
    df = pd.read_excel(path, sheet_name="cross checks all")
    assert select_x_check_nos(df, path, "cross checks all") == ["A1"]


def test_pipeline_does_not_mutate_input_df(tmp_path):
    headers = ["X-Check No.", "Status", "Type of Change", "Exclude Z-Core", "Category"]
    rows = [["A1", "INACTIVE", "Modified", "", ""], ["B2", "ACTIVE", "Modified", "", ""]]
    path = _write_xlsx(str(tmp_path), headers, rows)
    df = pd.read_excel(path, sheet_name="cross checks all")
    before = df.copy()
    select_x_check_nos(df, path, "cross checks all")
    pd.testing.assert_frame_equal(df, before)


def test_pipeline_returns_unique_in_order(tmp_path):
    headers = ["X-Check No.", "Status", "Type of Change", "Exclude Z-Core", "Category"]
    rows = [
        ["B2", "ACTIVE", "Modified", "", ""],
        ["A1", "ACTIVE", "Modified", "", ""],
        ["B2", "ACTIVE", "Modified", "", ""],   # duplicate of first
        ["C3", "ACTIVE", "New",      "", ""],
    ]
    path = _write_xlsx(str(tmp_path), headers, rows)
    df = pd.read_excel(path, sheet_name="cross checks all")
    assert select_x_check_nos(df, path, "cross checks all") == ["B2", "A1", "C3"]


# ---------------------------------------------------------------------------
# XChecks._write_x_check_no_list — wraps select_x_check_nos and writes .txt
# ---------------------------------------------------------------------------

def _make_xchecks_strategy():
    cfg = UploadTaskConfig(task_name="X-Checks", file_fields=[])
    s = XChecks(cfg)
    s.log = []
    s.process_only_differences = True
    return s


def test_writer_writes_one_per_line(tmp_path):
    headers = ["X-Check No.", "Status", "Type of Change", "Exclude Z-Core", "Category"]
    rows = [
        ["A1", "ACTIVE", "Modified", "", ""],
        ["B2", "ACTIVE", "New",      "", ""],
    ]
    path = _write_xlsx(str(tmp_path), headers, rows)
    df = pd.read_excel(path, sheet_name="cross checks all")

    s = _make_xchecks_strategy()
    out_dir = str(tmp_path)
    files = {
        "files":           {"X-Checks Publication File": path},
        "sheet_names":     {"X-Checks Publication File": "cross checks all"},
        "output_directory": out_dir,
        "timestamp":       "20260615_120000",
    }
    s._write_x_check_no_list({"X-Checks Publication File": df}, files)

    out = os.path.join(out_dir, "20260615_120000_X-Check_Nos.txt")
    assert os.path.exists(out)
    assert open(out).read() == "A1\nB2"


def test_writer_skips_when_no_x_check_nos(tmp_path):
    # All rows INACTIVE → nothing in scope
    headers = ["X-Check No.", "Status", "Type of Change", "Exclude Z-Core", "Category"]
    rows = [["A1", "INACTIVE", "Modified", "", ""]]
    path = _write_xlsx(str(tmp_path), headers, rows)
    df = pd.read_excel(path, sheet_name="cross checks all")

    s = _make_xchecks_strategy()
    out_dir = str(tmp_path)
    files = {
        "files":           {"X-Checks Publication File": path},
        "sheet_names":     {"X-Checks Publication File": "cross checks all"},
        "output_directory": out_dir,
        "timestamp":       "20260615_120000",
    }
    s._write_x_check_no_list({"X-Checks Publication File": df}, files)
    out = os.path.join(out_dir, "20260615_120000_X-Check_Nos.txt")
    assert not os.path.exists(out)
