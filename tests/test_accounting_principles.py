"""
Unit tests for the Accounting Principles strategy:
  - validation_methods.parse_validation_methods (Warning/Error/Both detection)
  - validation_methods._extract_method_codes (cell parser)
  - compare.compare (FIP-gated join + match rules)
"""
import os
import pandas as pd
import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from strategies.accounting_principles.validation_methods import (
    EventDefinition,
    _extract_method_codes,
    parse_validation_methods,
)
from strategies.accounting_principles.compare import compare


# ---------------------------------------------------------------------------
# _extract_method_codes
# ---------------------------------------------------------------------------

def test_extract_codes_single_line():
    assert _extract_method_codes("V900A - Part A (2023 onwards)") == ["V900A"]


def test_extract_codes_multiline():
    assert _extract_method_codes("V900A - Part A\nV900B - Part B") == ["V900A", "V900B"]


def test_extract_codes_empty():
    assert _extract_method_codes(None) == []
    assert _extract_method_codes("") == []
    assert _extract_method_codes("   ") == []


def test_extract_codes_dash():
    assert _extract_method_codes("-") == []
    assert _extract_method_codes("V900A - x\n-\nV900B - y") == ["V900A", "V900B"]


# ---------------------------------------------------------------------------
# parse_validation_methods (build a tiny xlsx in a tempfile)
# ---------------------------------------------------------------------------

def _make_vm_workbook(tmp_path, *, columns):
    """
    columns = list of (name, row4_value, row5_value, row6_value, merge_w_e)
    where merge_w_e=True means cells across rows 4-5 are merged (severity Both).
    """
    p = os.path.join(str(tmp_path), "vm.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Methods"
    # Header row
    ws.cell(row=1, column=1, value="Validation Event")
    for col_idx, (name, *_) in enumerate(columns, start=3):
        ws.cell(row=1, column=col_idx, value=name)
    for col_idx, (_, w, e1, e2, merge) in enumerate(columns, start=3):
        if w is not None:
            ws.cell(row=4, column=col_idx, value=w)
        if e1 is not None:
            ws.cell(row=5, column=col_idx, value=e1)
        if e2 is not None:
            ws.cell(row=6, column=col_idx, value=e2)
        if merge:
            from openpyxl.utils import get_column_letter
            letter = get_column_letter(col_idx)
            ws.merge_cells(f"{letter}4:{letter}5")
    wb.save(p)
    return p


def test_parse_warning_only(tmp_path):
    p = _make_vm_workbook(tmp_path, columns=[
        ("Ev1", "V900W - description", None, None, False),
    ])
    defs = parse_validation_methods(p, ["Ev1"])
    assert len(defs) == 1
    assert defs[0].severity == "Warning"
    assert defs[0].methods_w == ["V900W"]
    assert defs[0].methods_e == []


def test_parse_error_only_two_rows(tmp_path):
    p = _make_vm_workbook(tmp_path, columns=[
        ("Ev1", None, "V900A - x", "V900B - y", False),
    ])
    defs = parse_validation_methods(p, ["Ev1"])
    assert len(defs) == 1
    assert defs[0].severity == "Error"
    assert defs[0].methods_e == ["V900A", "V900B"]


def test_parse_independent_warning_and_error(tmp_path):
    p = _make_vm_workbook(tmp_path, columns=[
        ("Ev1", "V900W - w", "V900A - a", "V900B - b", False),
    ])
    defs = parse_validation_methods(p, ["Ev1"])
    sevs = sorted(d.severity for d in defs)
    assert sevs == ["Error", "Warning"]


def test_parse_both_via_merged_cell(tmp_path):
    p = _make_vm_workbook(tmp_path, columns=[
        ("Ev1", "V600A - both", None, None, True),
    ])
    defs = parse_validation_methods(p, ["Ev1"])
    assert len(defs) == 1
    assert defs[0].severity == "Both"
    assert defs[0].methods_w == ["V600A"]
    assert defs[0].methods_e == ["V600A"]


def test_parse_dash_treated_as_blank(tmp_path):
    p = _make_vm_workbook(tmp_path, columns=[
        ("Ev1", "-", None, None, False),
    ])
    defs = parse_validation_methods(p, ["Ev1"])
    assert defs == []   # nothing recorded → dropped


def test_parse_subset_filter(tmp_path):
    p = _make_vm_workbook(tmp_path, columns=[
        ("Ev1", "V001 - x", None, None, False),
        ("Ev2", "V002 - y", None, None, False),
        ("Ev3", "V003 - z", None, None, False),
    ])
    defs = parse_validation_methods(p, ["Ev1", "Ev3"])
    assert [d.event for d in defs] == ["Ev1", "Ev3"]


# ---------------------------------------------------------------------------
# compare — FIP-gated, match rules
# ---------------------------------------------------------------------------

def _make_cc_df(rows):
    """rows = list of dicts with X-Check No. + event letters."""
    return pd.DataFrame(rows)


def _make_fip_df(rows):
    """rows = list of (key, mt). Builds DataFrame with the columns used by compare()."""
    return pd.DataFrame([{"Key": k, "MT": v} for (k, v) in rows])


def test_compare_match_warning_w():
    defs = [EventDefinition(event="Ev1", severity="Warning", methods_w=["V001"])]
    cc  = _make_cc_df([{"X-Check No.": "X1", "Ev1": "w"}])
    fip = _make_fip_df([("V001|X1", "W")])
    out = compare(defs, cc, ["X1"], fip)
    assert out == [{"X-Check No.": "X1", "Event": "Ev1", "Expected": "Warning",
                    "FIP": "w", "Actual": "w", "Method": "V001", "Match": "Match"}]


def test_compare_mismatch_warning_actual_e():
    defs = [EventDefinition(event="Ev1", severity="Warning", methods_w=["V001"])]
    cc  = _make_cc_df([{"X-Check No.": "X1", "Ev1": "e"}])
    fip = _make_fip_df([("V001|X1", "W")])
    out = compare(defs, cc, ["X1"], fip)
    assert len(out) == 1
    assert out[0]["Match"] == "MisMatch"
    assert out[0]["Actual"] == "e"
    assert out[0]["FIP"] == "w"


def test_compare_fip_missing_no_row():
    """No FIP entry for the (method, X-Check) → no output row."""
    defs = [EventDefinition(event="Ev1", severity="Warning", methods_w=["V001"])]
    cc  = _make_cc_df([{"X-Check No.": "X1", "Ev1": "w"}])
    fip = _make_fip_df([])
    assert compare(defs, cc, ["X1"], fip) == []


def test_compare_actual_empty_no_row():
    """Even if FIP has the row, an empty cross-checks-all letter is skipped."""
    defs = [EventDefinition(event="Ev1", severity="Warning", methods_w=["V001"])]
    cc  = _make_cc_df([{"X-Check No.": "X1", "Ev1": ""}])
    fip = _make_fip_df([("V001|X1", "W")])
    assert compare(defs, cc, ["X1"], fip) == []


def test_compare_both_w():
    """Both event with actual=w → Match using methods_w."""
    defs = [EventDefinition(event="Ev1", severity="Both",
                            methods_w=["V600A"], methods_e=["V600A"])]
    cc  = _make_cc_df([{"X-Check No.": "X1", "Ev1": "w"}])
    fip = _make_fip_df([("V600A|X1", "W")])
    out = compare(defs, cc, ["X1"], fip)
    assert len(out) == 1
    assert out[0]["Match"] == "Match"


def test_compare_both_e():
    """Both event with actual=e → Match using methods_e."""
    defs = [EventDefinition(event="Ev1", severity="Both",
                            methods_w=["V600A"], methods_e=["V600A"])]
    cc  = _make_cc_df([{"X-Check No.": "X1", "Ev1": "e"}])
    fip = _make_fip_df([("V600A|X1", "E")])
    out = compare(defs, cc, ["X1"], fip)
    assert len(out) == 1
    assert out[0]["Match"] == "Match"


def test_compare_both_method_not_double_counted():
    """A Both event with the SAME method in both methods_w and methods_e
    must not produce two output rows for one FIP entry."""
    defs = [EventDefinition(event="Ev1", severity="Both",
                            methods_w=["V600A"], methods_e=["V600A"])]
    cc  = _make_cc_df([{"X-Check No.": "X1", "Ev1": "w"}])
    fip = _make_fip_df([("V600A|X1", "W")])
    out = compare(defs, cc, ["X1"], fip)
    assert len(out) == 1


def test_compare_skips_out_of_scope_xcheck():
    defs = [EventDefinition(event="Ev1", severity="Warning", methods_w=["V001"])]
    cc  = _make_cc_df([
        {"X-Check No.": "X1", "Ev1": "w"},
        {"X-Check No.": "X2", "Ev1": "w"},
    ])
    fip = _make_fip_df([("V001|X1", "W"), ("V001|X2", "W")])
    out = compare(defs, cc, ["X1"], fip)
    assert [r["X-Check No."] for r in out] == ["X1"]


def test_compare_event_name_punctuation_insensitive():
    """Validation methods uses 'DE-GAAP RFD'; cross-checks-all uses 'DE GAAP RFD'.
    The strategy must match them despite the hyphen-vs-space difference."""
    defs = [EventDefinition(event="DE-GAAP RFD", severity="Error", methods_e=["V791A"])]
    cc  = _make_cc_df([{"X-Check No.": "X1", "DE GAAP RFD": "e"}])  # space, not hyphen
    fip = _make_fip_df([("V791A|X1", "E")])
    out = compare(defs, cc, ["X1"], fip)
    assert len(out) == 1
    assert out[0]["Match"] == "Match"
    assert out[0]["Event"] == "DE-GAAP RFD"   # output keeps the validation-methods spelling


def test_compare_pandas_dot_n_dedup_column_ignored():
    """When the same column name appears twice in cross-checks-all, pandas
    appends '.1', '.2', etc. The matcher should pick the FIRST occurrence
    only (the un-suffixed one)."""
    defs = [EventDefinition(event="Ev1", severity="Warning", methods_w=["V001"])]
    # Build a df with a duplicate-named column; pandas adds '.1' to the second
    cc = pd.DataFrame([{"X-Check No.": "X1", "Ev1": "w", "Ev1.1": "BOGUS"}])
    fip = _make_fip_df([("V001|X1", "W")])
    out = compare(defs, cc, ["X1"], fip)
    assert len(out) == 1
    assert out[0]["Actual"] == "w"
