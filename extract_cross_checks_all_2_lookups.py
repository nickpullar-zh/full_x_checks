"""
One-off extraction from the 'cross checks all (2)' sheet.

Each method has a 5-column block starting at BS (col 71) and repeating every
5 columns (BS, BX, CC, CH, ...). Within a block:
  +0  Method block-start: row1 = method code (e.g. 'V900W'),
                          row2 = concat label (e.g. 'IFRS New RFD Warning'),
                          data rows = '<method>|<X-Check No.>'
  +1  row2 = 'Lookup',     data rows = '<method>|<X-Check No.>' (if found in VALMSG)
  +2  row2 = 'This Sheet', data rows = letter from cross-checks-all (e.g. 'w')
  +3  row2 = 'VALMSG',     data rows = letter from VALMSG (e.g. 'W')
  +4  row2 = 'Match',      data rows = True / False

This script keeps every data row whose Lookup sub-column is non-empty and
writes them to <input_basename>_lookups.xlsx in the same folder.

Run:
  python extract_cross_checks_all_2_lookups.py "<path to .xlsx>"

If no path is supplied it defaults to the Q2 X-Checks reference workbook.
"""
import os
import sys

import openpyxl


SHEET_NAME = "cross checks all (2)"
DATA_HEADER_ROW = 2          # row 2 of the EBX sheet holds 'Type of change' etc.
FIRST_DATA_ROW  = 3          # data starts at row 3
X_CHECK_COL     = 5          # column E
FIRST_BLOCK_COL = 71         # column BS
BLOCK_STRIDE    = 5

OUTPUT_HEADERS = [
    "X-Check No.", "Method", "Concat",
    "Key", "Lookup", "This Sheet", "VALMSG", "Match",
]


def _norm(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none") else s


def extract(input_path: str, output_path: str | None = None) -> str:
    wb = openpyxl.load_workbook(input_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in {input_path}")
    ws = wb[SHEET_NAME]

    # Discover blocks: every column from FIRST_BLOCK_COL whose row 1 is non-empty
    # AND falls on the BLOCK_STRIDE grid.
    blocks: list[tuple[int, str, str]] = []
    for c in range(FIRST_BLOCK_COL, ws.max_column + 1, BLOCK_STRIDE):
        method = _norm(ws.cell(row=1, column=c).value)
        concat = _norm(ws.cell(row=2, column=c).value)
        if method:
            blocks.append((c, method, concat))

    out_rows: list[list] = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        x_check = _norm(ws.cell(row=r, column=X_CHECK_COL).value)
        for start_col, method, concat in blocks:
            key        = _norm(ws.cell(row=r, column=start_col).value)
            lookup     = _norm(ws.cell(row=r, column=start_col + 1).value)
            this_sheet = _norm(ws.cell(row=r, column=start_col + 2).value)
            valmsg     = _norm(ws.cell(row=r, column=start_col + 3).value)
            match      = ws.cell(row=r, column=start_col + 4).value
            if not lookup:
                continue
            out_rows.append([
                x_check, method, concat, key, lookup, this_sheet, valmsg, match,
            ])

    wb.close()

    if output_path is None:
        base, _ = os.path.splitext(input_path)
        output_path = base + "_lookups.xlsx"

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Lookups"
    out_ws.append(OUTPUT_HEADERS)
    for row in out_rows:
        out_ws.append(row)
    out_wb.save(output_path)

    print(f"  Read {ws.max_row - FIRST_DATA_ROW + 1} data rows x {len(blocks)} method blocks")
    print(f"  Wrote {len(out_rows)} rows to: {output_path}")
    return output_path


if __name__ == "__main__":
    if len(sys.argv) > 1:
        path = sys.argv[1]
    else:
        path = (
            r"C:\Users\NICK.PULLAR\OneDrive - Zurich Insurance"
            r"\Projects\2026 Testing\06 2026\Q2 X-Checks"
            r"\20260602 VALMSG (Accounting Principle).XLSX"
        )
    extract(path)
