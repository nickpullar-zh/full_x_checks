"""
Known Exception List Builder

Opens a small dialog that creates a correctly structured Known Exception List
.xlsx file. Each strategy gets its own sheet with headers derived from its
fingerprint columns. An optional comparison output workbook can be imported to
pre-fill mismatch rows.
"""

import tkinter as tk
import sys as _sys
from tkinter import ttk, filedialog, messagebox


def _exe_dir() -> str:
    if getattr(_sys, 'frozen', False):
        return os.path.dirname(_sys.executable)
    return os.path.dirname(os.path.abspath(__file__))
import os

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Strategy definitions — single source of truth.
# fingerprint_cols must exactly match what each strategy passes to
# _annotate_known_exceptions(). If a strategy's fingerprint changes, update here.
# ---------------------------------------------------------------------------

STRATEGY_DEFINITIONS = {
    "X-Checks": {
        "kel_sheet":         "X-Checks",
        "fingerprint_cols":  [
            "X-Check No.", "EBX Formula", "FIP Formula",
            "EBX Formula (Excl)", "FIP Formula (Excl)",
            "EBX Variables", "FIP Variables", "FIP Variable (Builder)",
        ],
        "comparison_sheets": ["Comparison", "XC — Comparison"],
        "mismatch_col":      "Formula Match",
        "mismatch_val":      "Match",
    },
    "Grouping By": {
        "kel_sheet":         "Grouping By",
        "fingerprint_cols":  ["EBX Key"],
        "comparison_sheets": ["Comparison", "GB — Comparison"],
        "mismatch_col":      "Result",
        "mismatch_val":      "Matched",
    },
    "Accounting Principles": {
        "kel_sheet":         "Accounting Principles",
        "fingerprint_cols":  ["X-Check No.", "Event", "Expected", "FIP", "Actual", "Method"],
        "comparison_sheets": ["Comparison", "AP — Comparison"],
        "mismatch_col":      "Match",
        "mismatch_val":      "Match",
    },
    "Conditions": {
        "kel_sheet":         "Conditions",
        "fingerprint_cols":  ["EBX Data", "FIP Data"],
        "comparison_sheets": ["Comparison", "Cond — Comparison"],
        "mismatch_col":      "Comparison",
        "mismatch_val":      "Matched",
    },
}

_METADATA_COLS = ["Reason", "Added By", "Date Added", "Resolution Status", "Resolution Notes"]

_HEADER_FILL  = PatternFill("solid", fgColor="91BFE3")
_HEADER_FONT  = Font(name="Zurich Sans", bold=True, color="23366F")
_GUIDANCE_FILL = PatternFill("solid", fgColor="ECEEEF")
_GUIDANCE_FONT = Font(name="Zurich Sans", italic=True, color="23366F")
_BODY_FONT    = Font(name="Zurich Sans", color="23366F")
_WRAP         = Alignment(wrap_text=True, vertical="top")


class KnownExceptionBuilderDialog:
    """Modal dialog for building a Known Exception List workbook."""

    def __init__(self, parent: tk.Tk):
        self._parent = parent
        self._out_path = tk.StringVar()
        self._import_path = tk.StringVar()
        # {strategy_name: (BooleanVar, detected_sheet_name)}
        self._strategy_vars: dict[str, tuple[tk.BooleanVar, str]] = {}

        self._win = tk.Toplevel(parent)
        self._win.title("Known Exception List Builder")
        self._win.resizable(False, False)
        self._win.grab_set()
        self._build_ui()
        self._centre()

    # ------------------------------------------------------------------ UI

    def _build_ui(self):
        pad = {"padx": 10, "pady": 5}
        f = ttk.Frame(self._win, padding=15)
        f.grid(sticky="nsew")

        row = 0

        # Title
        ttk.Label(f, text="Known Exception List Builder",
                  font=("Zurich Sans Semibold", 13)).grid(
            row=row, column=0, columnspan=3, sticky="w", pady=(0, 12))
        row += 1

        ttk.Separator(f, orient="horizontal").grid(
            row=row, column=0, columnspan=3, sticky="ew", pady=(0, 8))
        row += 1

        # Output file
        ttk.Label(f, text="Save as:").grid(row=row, column=0, sticky="w", **pad)
        self._out_lbl = ttk.Label(f, text="Click Browse and select a folder, then type the filename",
                                   foreground="#888888", width=38, anchor="w", wraplength=220)
        self._out_lbl.grid(row=row, column=1, sticky="w", **pad)
        ttk.Button(f, text="Browse…", command=self._browse_output).grid(
            row=row, column=2, **pad)
        row += 1

        ttk.Separator(f, orient="horizontal").grid(
            row=row, column=0, columnspan=3, sticky="ew", pady=(4, 8))
        row += 1

        # Import section
        ttk.Label(f, text="Import mismatches from\n(optional):",
                  justify="left").grid(row=row, column=0, sticky="nw", **pad)
        self._import_lbl = ttk.Label(f, text="No file selected", foreground="#888888",
                                      width=38, anchor="w")
        self._import_lbl.grid(row=row, column=1, sticky="w", **pad)
        ttk.Button(f, text="Browse…", command=self._browse_import).grid(
            row=row, column=2, **pad)
        row += 1

        # Strategy checkboxes (hidden until import file selected)
        self._cb_frame = ttk.Frame(f)
        self._cb_frame.grid(row=row, column=0, columnspan=3, sticky="w",
                             padx=10, pady=(0, 4))
        self._detected_label = ttk.Label(self._cb_frame,
                                          text="Strategies detected:")
        # Initialise checkbox vars (disabled until import file chosen)
        self._checkboxes: dict[str, ttk.Checkbutton] = {}
        for i, name in enumerate(STRATEGY_DEFINITIONS):
            var = tk.BooleanVar(value=False)
            self._strategy_vars[name] = (var, "")
            cb = ttk.Checkbutton(self._cb_frame, text=name, variable=var, state="disabled")
            self._checkboxes[name] = cb
        row += 1

        ttk.Separator(f, orient="horizontal").grid(
            row=row, column=0, columnspan=3, sticky="ew", pady=(4, 8))
        row += 1

        # Open after build checkbox
        self._open_after_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(f, text="Open file after building",
                        variable=self._open_after_var).grid(
            row=row, column=0, columnspan=3, sticky="w", padx=10, pady=(0, 6))
        row += 1

        # Buttons
        btn_frame = ttk.Frame(f)
        btn_frame.grid(row=row, column=0, columnspan=3, pady=(0, 4))
        self._build_btn = ttk.Button(btn_frame, text="Build",
                                      command=self._on_build, state="disabled", width=14)
        self._build_btn.pack(side="left", padx=(0, 8))
        ttk.Button(btn_frame, text="Cancel", command=self._win.destroy,
                   width=14).pack(side="left")

    def _centre(self):
        self._win.update_idletasks()
        w = self._win.winfo_reqwidth()
        h = self._win.winfo_reqheight()

        # Ideal position: centred over the parent window
        pw = self._parent.winfo_x() + self._parent.winfo_width() // 2
        ph = self._parent.winfo_y() + self._parent.winfo_height() // 2
        x = pw - w // 2
        y = ph - h // 2

        # Clamp to screen with a 40px margin from every edge
        margin = 40
        screen_w = self._win.winfo_screenwidth()
        screen_h = self._win.winfo_screenheight()
        x = max(margin, min(x, screen_w - w - margin))
        y = max(margin, min(y, screen_h - h - margin))

        self._win.geometry(f"+{x}+{y}")

    # ------------------------------------------------------------------ Browse

    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            parent=self._win,
            title="Save Known Exception List as…",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialdir=_exe_dir(),
        )
        if path:
            self._out_path.set(path)
            self._out_lbl.config(text=os.path.basename(path), foreground="#23366F")
            self._check_ready()

    def _browse_import(self):
        path = filedialog.askopenfilename(
            parent=self._win,
            title="Select comparison output workbook…",
            filetypes=[("Excel Files", "*.xlsx")],
            initialdir=_exe_dir(),
        )
        if not path:
            return
        self._import_path.set(path)
        self._import_lbl.config(text=os.path.basename(path), foreground="#23366F")
        self._detect_strategies(path)

    def _detect_strategies(self, path: str):
        """Open the import workbook, check which strategy comparison sheets are present."""
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet_names = set(wb.sheetnames)
            wb.close()
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file:\n{e}", parent=self._win)
            return

        detected = {}
        for strategy_name, defn in STRATEGY_DEFINITIONS.items():
            for sn in defn["comparison_sheets"]:
                if sn in sheet_names:
                    detected[strategy_name] = sn
                    break

        # Rebuild checkbox layout
        for widget in self._cb_frame.winfo_children():
            widget.grid_forget()

        if detected:
            self._detected_label.grid(row=0, column=0, columnspan=4, sticky="w",
                                       padx=0, pady=(4, 2))
            for i, (name, sheet) in enumerate(detected.items()):
                var, _ = self._strategy_vars[name]
                var.set(True)
                self._strategy_vars[name] = (var, sheet)
                cb = self._checkboxes[name]
                cb.config(state="normal")
                cb.grid(row=1 + i // 2, column=i % 2, sticky="w", padx=(0, 16), pady=1)
        else:
            ttk.Label(self._cb_frame,
                      text="No comparison sheets detected in this file.",
                      foreground="#888888").grid(row=0, column=0, sticky="w", pady=4)

    def _check_ready(self):
        if self._out_path.get():
            self._build_btn.config(state="normal")
        else:
            self._build_btn.config(state="disabled")

    # ------------------------------------------------------------------ Build

    def _on_build(self):
        out = self._out_path.get()
        if not out:
            return

        import_path = self._import_path.get()
        import_wb = None
        if import_path:
            try:
                import_wb = openpyxl.load_workbook(import_path, read_only=True, data_only=True)
            except Exception as e:
                messagebox.showerror("Error", f"Could not open import file:\n{e}",
                                     parent=self._win)
                return

        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            for strategy_name, defn in STRATEGY_DEFINITIONS.items():
                var, detected_sheet = self._strategy_vars[strategy_name]
                import_rows = []
                if import_wb and var.get() and detected_sheet:
                    import_rows = _extract_mismatch_rows(import_wb, detected_sheet, defn)

                _build_strategy_sheet(wb, defn, import_rows)

            _build_instructions_sheet(wb)

            if import_wb:
                import_wb.close()

            wb.save(out)

            # Verify the file was actually written before closing
            if not os.path.isfile(out):
                raise RuntimeError(
                    f"File was not found at the expected path after saving:\n{out}"
                )

            # Apply MIP sensitivity label — best-effort, does not block on failure
            try:
                from strategies.sensitivity import ExcelLabeler
                labeler = ExcelLabeler()
                labeler.label_file(out, "Internal_Use_Only")
                labeler.close()
            except Exception:
                pass

            open_after = self._open_after_var.get()
            messagebox.showinfo("Done",
                                f"Known Exception List saved to:\n{out}",
                                parent=self._win)
            self._win.destroy()

            if open_after:
                import subprocess as _sp
                _sp.Popen(["start", "", out], shell=True)

        except Exception as e:
            if import_wb:
                try:
                    import_wb.close()
                except Exception:
                    pass
            messagebox.showerror("Build failed", str(e), parent=self._win)


# ---------------------------------------------------------------------------
# Excel building helpers
# ---------------------------------------------------------------------------

def _build_strategy_sheet(wb: openpyxl.Workbook, defn: dict, import_rows: list[dict]):
    ws = wb.create_sheet(defn["kel_sheet"])

    all_cols = defn["fingerprint_cols"] + _METADATA_COLS

    # Row 1 — headers
    for col_idx, col_name in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP

    # Row 2 — guidance
    guidance = {col: "Enter value exactly as shown in the Comparison output"
                for col in defn["fingerprint_cols"]}
    guidance["Reason"]            = "Required — why these values are expected to differ"
    guidance["Added By"]          = "Your name"
    guidance["Date Added"]        = "YYYY-MM-DD"
    guidance["Resolution Status"] = "Open / In Progress / Resolved / Permanent"
    guidance["Resolution Notes"]  = ""

    for col_idx, col_name in enumerate(all_cols, start=1):
        cell = ws.cell(row=2, column=col_idx, value=guidance.get(col_name, ""))
        cell.fill = _GUIDANCE_FILL
        cell.font = _GUIDANCE_FONT
        cell.alignment = _WRAP

    # Rows 3+ — imported data rows
    for row_offset, row_data in enumerate(import_rows):
        row_idx = 3 + row_offset
        for col_idx, col_name in enumerate(all_cols, start=1):
            val = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _BODY_FONT
            cell.alignment = _WRAP

    # Column widths
    _autofit(ws, all_cols, min_w=15, max_w=60)

    # Freeze header rows
    ws.freeze_panes = "A3"


def _extract_mismatch_rows(import_wb: openpyxl.Workbook, sheet_name: str,
                            defn: dict) -> list[dict]:
    """Read the comparison sheet and return mismatch rows as dicts of fingerprint cols."""
    import pandas as pd

    ws = import_wb[sheet_name]

    # Read header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if not headers or headers[0] is None:
        return []

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        rows.append(row_dict)

    if not rows:
        return []

    mismatch_col = defn["mismatch_col"]
    mismatch_val = defn["mismatch_val"]
    fingerprint_cols = defn["fingerprint_cols"]

    # Filter to mismatches only
    mismatches = [r for r in rows if str(r.get(mismatch_col, "")).strip() != mismatch_val]

    # Deduplicate by fingerprint key
    seen = set()
    deduped = []
    for r in mismatches:
        key_parts = tuple(str(r.get(c, "") or "").strip() for c in fingerprint_cols)
        if key_parts not in seen:
            seen.add(key_parts)
            deduped.append({c: r.get(c, "") for c in fingerprint_cols})

    return deduped


def _autofit(ws, col_names: list[str], min_w: int = 15, max_w: int = 60):
    for col_idx, col_name in enumerate(col_names, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(col_name)), min_w)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    # Cap individual cell contribution — long formula strings shouldn't widen too much
                    max_len = min(max(max_len, len(str(cell.value).split("\n")[0])), max_w)
        ws.column_dimensions[col_letter].width = max_len + 2


def _build_instructions_sheet(wb: openpyxl.Workbook):
    ws = wb.create_sheet("Instructions")

    title_font  = Font(name="Zurich Sans Semibold", bold=True, size=13, color="23366F")
    body_font   = Font(name="Zurich Sans", size=10, color="23366F")
    head2_font  = Font(name="Zurich Sans Semibold", bold=True, size=10, color="23366F")

    lines = [
        ("Known Exception List — Usage Guide", title_font),
        ("", body_font),
        ("Purpose", head2_font),
        ("This workbook records X-Check mismatches that are known, understood, and accepted. "
         "When loaded into the X-Checks application the app annotates matching rows in the "
         "comparison output with the Reason text rather than treating them as new issues.", body_font),
        ("", body_font),
        ("Sheet structure", head2_font),
        ("One sheet per strategy: X-Checks, Grouping By, Accounting Principles, Conditions.", body_font),
        ("", body_font),
        ("Row layout", head2_font),
        ("Row 1  — Column headers (do not edit).", body_font),
        ("Row 2  — Guidance row (always skipped by the app — do not delete).", body_font),
        ("Row 3+ — One exception per row.", body_font),
        ("", body_font),
        ("Required columns", head2_font),
        ("The fingerprint columns (all columns before 'Reason') must exactly match the values "
         "shown in the Comparison output sheet of a run. The 'Reason' column is required — "
         "rows with a blank Reason will be rejected.", body_font),
        ("", body_font),
        ("Optional columns", head2_font),
        ("Added By, Date Added, Resolution Status, Resolution Notes are for your records only. "
         "The app ignores them.", body_font),
        ("", body_font),
        ("How to use the builder", head2_font),
        ("Run any strategy to produce a comparison output. Open the builder (⚙ button), "
         "set a save path, optionally browse to the comparison output to import mismatch rows, "
         "and click Build. Then open the file and fill in the Reason column for each row "
         "you want to accept as a known exception.", body_font),
    ]

    ws.column_dimensions["A"].width = 90
    for i, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 30 if font == body_font else 20
