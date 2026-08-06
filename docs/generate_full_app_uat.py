"""
Generate the Full Application v1.0 UAT Test Plan workbook.

Output: docs/<YYYYMMDD> Full_Application_v<version> Test Plan.xlsx

Run from the repo root:
    python docs/generate_full_app_uat.py
"""

import os
import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from version import __version__

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

VERSION  = __version__
TODAY    = date.today().strftime("%Y%m%d")
OUT_DIR  = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "docs")


def _next_suffix(out_dir, stem_pattern):
    """Return the next letter suffix (_a, _b, ..., _z, _aa, _ab, ...) for a versioned file.
    Scans out_dir for existing files whose name contains stem_pattern and extracts used suffixes."""
    import re
    import string
    used = set()
    if os.path.isdir(out_dir):
        for f in os.listdir(out_dir):
            if stem_pattern in f:
                m = re.search(r'_([a-z]+)\.xlsx$', f)
                if m:
                    used.add(m.group(1))

    def _seq():
        for c in string.ascii_lowercase:
            yield c
        for c1 in string.ascii_lowercase:
            for c2 in string.ascii_lowercase:
                yield c1 + c2

    for s in _seq():
        if s not in used:
            return s
    return "a"


SUFFIX   = _next_suffix(OUT_DIR, f"Full_Application_v{VERSION} Test Plan_")
FILENAME = f"{TODAY} Full_Application_v{VERSION} Test Plan_{SUFFIX}.xlsx"
OUT_PATH = os.path.join(OUT_DIR, FILENAME)

# Test data files (all in test_data\ relative to repo root)
TD = "test_data"
FILES = {
    "xc_pub":     "20251205 EPM X-Checks - Original.xlsx",
    "xc_fip":     "20251205 FIP X-Checks - Original.txt",
    "gb_fip":     "VALFLDGR file with 12348 Data rows on sheet Sheet1.XLSX",
    "gb_pub":     "20260313 Cross Checks All.xlsx",
    "mapping":    "Mapping Table with 20 rows.txt",
    "ap_val_mth": "validation methods.xlsx",
    "ap_pub":     "EPM X-Checks file with 3345 Data rows on sheet cross checks all.xlsx",
    "ap_fip":     "20260728_VALMSG_File_direct_from_FIP.XLSX",
    "cond_pub":   "20260313 Cross Checks All.xlsx",
    "cond_fip":   "20260602 VALMETH (Conditions).xlsx",
    "known_exc":  "Known_Exception_List.xlsx",
}
SHEETS = {
    "xc_pub":     "cross checks all",
    "gb_fip":     "Sheet1",
    "gb_pub":     "cross checks all",
    "ap_val_mth": "Validation Methods",
    "ap_pub":     "cross checks all",
    "ap_fip":     "Sheet1",   # raw ZQ9_VALMSG: MK + ValidRule → Key built at load time
    "cond_pub":   "cross checks all",
    "cond_fip":   "FIP Conditions",
    "known_exc":  "Known Exceptions",
}

# Actual counts derived by running the app logic against these files
XC_EBX_ROWS          = 3345   # source rows in pub file
XC_EBX_EXTRACTED     = 664    # X-Check entries extracted by EBX extraction
XC_FIP_ROWS          = 27737  # source rows in FIP file
XC_FIP_MATCHED       = 650    # X-Checks found in FIP
XC_RESULTS           = 653    # comparison output rows
XC_FORMULA_MATCH     = 525
XC_FORMULA_MISMATCH  = 125
XC_FORMULA_NOTFOUND  = 3

GB_PUB_ROWS          = 1499   # source rows in pub file (excl header)
GB_FIP_ROWS          = 12348  # source rows in VALFLDGR file
GB_FIP_PROCESSED     = 6776   # after mapping + filtering
GB_MAPPING_ROWS      = 19     # mapping dict entries

AP_PUB_ROWS          = 3345   # source rows in AP pub file
AP_FIP_ROWS          = 17184  # source rows in VALMSG file (20260728 raw ZQ9_VALMSG)

COND_FIP_ROWS        = 4816   # processed rows in VALMETH FIP file
COND_DIFF_PAIRS      = 92     # comparison rows in differences mode
COND_FULL_PAIRS      = 130    # comparison rows in full-file mode

KNOWN_EXC_ROWS       = 2      # data rows in Known_Exception_List.xlsx

# Dropdown order in the app
TASKS = [
    "Collect Live X-Checks",
    "X-Checks",
    "Grouping By",
    "Accounting Principles",
    "Conditions",
    "Full Run",
]

# Zurich brand colours
DARK_BLUE  = "FF23366F"
LIGHT_BLUE = "FF91BFE3"
LOGIC_BLUE = "FFD6E4F7"
WHITE      = "FFFFFFFF"
ALT_GREY   = "FFECEEEF"

def _font(name="Zurich Sans", size=10, bold=False, color=DARK_BLUE):
    return Font(name=name, size=size, bold=bold, color=color)

def _semibold(size=10, color=DARK_BLUE):
    return Font(name="Zurich Sans Semibold", size=size, bold=True, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

_thin = Side(style="thin")
_med  = Side(style="medium")
ALL_THIN   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
TOP_WRAP   = Alignment(vertical="top", wrap_text=True)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# TEST CASES
# ---------------------------------------------------------------------------
# Format: (ID, Area, Test Type, Precondition, Steps, Expected Result)
# Test Type: "Logic" | "Whole App"

TEST_CASES = [

    # ── Launch & version ────────────────────────────────────────────────────
    (
        "FA-01", "Launch & version", "Whole App",
        f"Production EXE present in dist\\.",
        f"Double-click dist\\X-Checks_v{VERSION}.exe.",
        f"Splash screen appears showing 'X-Check Application v{VERSION} Loading...'. "
        f"After ~2–5 s the task selector form opens with 'X-Check Application v{VERSION}' "
        f"in the title bar and UI label.",
    ),
    (
        "FA-02", "Launch & version", "Whole App",
        "App is open at the task selector.",
        "Confirm the version shown in the title bar and UI label.",
        f"Both display v{VERSION}.",
    ),

    # ── Task selector ────────────────────────────────────────────────────────
    (
        "FA-03", "Task selector — dropdown", "Whole App",
        "App is open at the task selector.",
        "Open the task dropdown and count the entries.",
        f"{len(TASKS)} tasks are listed in this order:\n" +
        "\n".join(f"{i+1}. {t}" for i, t in enumerate(TASKS)),
    ),
    (
        "FA-04", "Task selector — navigation", "Whole App",
        "App is open at the task selector.",
        "Select each task in turn and confirm the file-selection form title.",
        "Each task opens a correctly titled form:\n"
        "• Collect Live X-Checks → 'Collect Live X-Checks'\n"
        "• X-Checks → 'X-Check Files'\n"
        "• Grouping By → 'Grouping By Files'\n"
        "• Accounting Principles → 'Accounting Principles Files'\n"
        "• Conditions → 'Conditions Files'\n"
        "• Full Run → 'Full Run — All Strategies'",
    ),

    # ── X-Checks ─────────────────────────────────────────────────────────────
    (
        "FA-05", "X-Checks — file fields", "Whole App",
        "X-Checks task selected, file-selection form open.",
        "Inspect the form fields.",
        "Four fields present:\n"
        f"1. 'FIP File' (.txt)\n"
        f"2. 'X-Checks Publication File' (.xlsx, default sheet: '{SHEETS['xc_pub']}')\n"
        f"3. 'GCoA Publication File' (.xlsx, optional)\n"
        f"4. 'Known Exception List' (.xlsx, optional)\n"
        "Two experimental checkboxes: 'Apply Version Spanning Validation' and "
        "'Apply Prior Year Balance Formula', both unchecked by default.\n"
        "'Process only differences' checkbox is present and checked by default.",
    ),
    (
        "FA-06", "X-Checks — run (differences only)", "Logic",
        "X-Checks form open.",
        f"Set FIP File to '{FILES['xc_fip']}', "
        f"X-Checks Publication File to '{FILES['xc_pub']}' (sheet: '{SHEETS['xc_pub']}'), "
        f"no GCoA or Known Exception List. "
        f"'Process only differences' checked. click Proceed.",
        f"Run completes. Progress log shows {XC_EBX_EXTRACTED} X-Check entries extracted "
        f"from the publication file and {XC_FIP_MATCHED} matched in the FIP file.",
    ),
    (
        "FA-07", "X-Checks — output structure", "Whole App",
        "FA-06 complete. Output workbook open.",
        "Check the sheet tabs.",
        "Workbook contains 7 sheets:\n"
        "1. EBX Data        — raw 'cross checks all' publication data\n"
        "2. FIP Data        — parsed FIP results (one row per X-Check found in FIP text)\n"
        "3. Comparison      — full combined output\n"
        "4. Matched Data    — only rows where all 4 match columns = Match\n"
        "5. MisMatched Data — only rows where any match column = MisMatch\n"
        "6. Not Found Data  — only rows where any match column = Not Found\n"
        "7. Processing Log",
    ),
    (
        "FA-08", "X-Checks — comparison counts", "Logic",
        "FA-06 output, Comparison sheet open.",
        f"Count the rows and check the Formula Match column breakdown. "
        f"Also verify the filtered sheet counts.",
        f"Comparison: {XC_RESULTS} rows total.\n"
        f"Formula Match: {XC_FORMULA_MATCH} Match, {XC_FORMULA_MISMATCH} MisMatch, "
        f"{XC_FORMULA_NOTFOUND} Not Found.\n\n"
        f"Matched Data:    {XC_FORMULA_MATCH} rows.\n"
        f"MisMatched Data: {XC_FORMULA_MISMATCH} rows.\n"
        f"Not Found Data:  {XC_FORMULA_NOTFOUND} rows.",
    ),
    (
        "FA-09", "X-Checks — Known Exception List", "Logic",
        "X-Checks form open.",
        f"Add '{FILES['known_exc']}' (sheet: '{SHEETS['known_exc']}') as the "
        f"Known Exception List. Run with the same files as FA-06.",
        f"Run completes. Progress log references the Known Exception List. "
        f"Rows matching the {KNOWN_EXC_ROWS} exception entries are annotated in the output.",
    ),

    # ── Grouping By ──────────────────────────────────────────────────────────
    (
        "FA-10", "Grouping By — file fields", "Whole App",
        "Grouping By task selected, file-selection form open.",
        "Inspect the form fields.",
        "Four fields present:\n"
        f"1. 'FIP File (ZQ9_VALFLDGR)' (.xlsx, default sheet: '{SHEETS['gb_fip']}')\n"
        f"2. 'X-Checks Publication File' (.xlsx, default sheet: '{SHEETS['gb_pub']}')\n"
        f"3. 'Mapping File' (.csv / .txt)\n"
        f"4. 'Known Exception List' (.xlsx, optional)",
    ),
    (
        "FA-11", "Grouping By — run", "Logic",
        "Grouping By form open.",
        f"Set FIP File (ZQ9_VALFLDGR) to '{FILES['gb_fip']}' (sheet: '{SHEETS['gb_fip']}'), "
        f"X-Checks Publication File to '{FILES['gb_pub']}' (sheet: '{SHEETS['gb_pub']}'), "
        f"Mapping File to '{FILES['mapping']}'. click Proceed.",
        f"Run completes without error. Progress log shows:\n"
        f"• FIP original loaded: {GB_FIP_ROWS} rows\n"
        f"• FIP processed (after mapping + filtering): {GB_FIP_PROCESSED} rows\n"
        f"• Mapping dictionary: {GB_MAPPING_ROWS} entries\n"
        f"• EBX rows: {GB_PUB_ROWS}",
    ),
    (
        "FA-12", "Grouping By — output structure", "Whole App",
        "FA-11 complete. Output workbook open.",
        "Check the sheet tabs.",
        "Workbook contains:\n"
        "1. FIP - Original\n"
        "2. FIP - Processed\n"
        "3. EBX Processed\n"
        "4. Comparison\n"
        "5. Processing Log",
    ),
    (
        "FA-13", "Grouping By — colour coding", "Whole App",
        "FA-11 output, Comparison sheet open.",
        "Review the Result column.",
        "'Matched' rows have a green fill. 'Not in FIP' rows have an orange fill.",
    ),

    # ── Accounting Principles ────────────────────────────────────────────────
    (
        "FA-14", "Accounting Principles — file fields", "Whole App",
        "Accounting Principles task selected, file-selection form open.",
        "Inspect the form fields.",
        "Four fields present:\n"
        f"1. 'Validation Methods File' (.xlsx, default sheet: '{SHEETS['ap_val_mth']}')\n"
        f"2. 'X-Checks Publication File' (.xlsx, default sheet: '{SHEETS['ap_pub']}')\n"
        "3. 'FIP File (VALMSG)' (.xlsx, sheet: 'FIP Methods Rules and Condition')\n"
        "4. 'Known Exception List' (.xlsx, optional)",
    ),
    (
        "FA-15", "Accounting Principles — run", "Logic",
        "Accounting Principles form open.",
        f"Set Validation Methods File to '{FILES['ap_val_mth']}' (sheet: '{SHEETS['ap_val_mth']}'), "
        f"X-Checks Publication File to '{FILES['ap_pub']}' (sheet: '{SHEETS['ap_pub']}'), "
        f"FIP File (VALMSG) to '{FILES['ap_fip']}' (sheet: '{SHEETS['ap_fip']}'). "
        f"This is the raw ZQ9_VALMSG extract — the app builds the Key column from MK + ValidRule. click Proceed.",
        f"Run completes. Progress log confirms {AP_PUB_ROWS} source rows in the "
        f"publication file and {AP_FIP_ROWS} rows in the VALMSG file. "
        f"Log shows 'Built Key column from MK + ValidRule'.",
    ),
    (
        "FA-16", "Accounting Principles — output structure", "Whole App",
        "FA-15 complete. Output workbook open.",
        "Check the sheet tabs.",
        "Workbook contains:\n"
        "1. EBX\n"
        "2. FIP\n"
        "3. Comparison\n"
        "4. Processing Log",
    ),

    # ── Conditions ───────────────────────────────────────────────────────────
    (
        "FA-17", "Conditions — file fields", "Whole App",
        "Conditions task selected, file-selection form open.",
        "Inspect the form fields.",
        "Three fields present:\n"
        f"1. 'X-Checks Publication File' (.xlsx, default sheet: '{SHEETS['cond_pub']}')\n"
        f"2. 'FIP File (ZQ9_VALMETH)' (.xlsx, default sheet: '{SHEETS['cond_fip']}')\n"
        "3. 'Known Exception List' (.xlsx, optional)\n"
        "'Process only differences' checkbox present and checked by default.",
    ),
    (
        "FA-18", "Conditions — run (differences only)", "Logic",
        "Conditions form open. 'Process only differences' checked (default).",
        f"Set X-Checks Publication File to '{FILES['cond_pub']}' (sheet: '{SHEETS['cond_pub']}'), "
        f"FIP File (ZQ9_VALMETH) to '{FILES['cond_fip']}' (sheet: '{SHEETS['cond_fip']}'). click Proceed.",
        f"Run completes. Progress log shows {COND_FIP_ROWS} FIP rows processed "
        f"and {COND_DIFF_PAIRS} pairs compared.",
    ),
    (
        "FA-19", "Conditions — run (full file)", "Logic",
        "Conditions form open. 'Process only differences' unchecked.",
        "Run with same files as FA-18.",
        f"Progress log reports full-file extraction mode. "
        f"{COND_FULL_PAIRS} pairs compared (more than the {COND_DIFF_PAIRS} in differences mode).",
    ),
    (
        "FA-20", "Conditions — colour coding", "Whole App",
        "FA-18 or FA-19 complete. Output workbook open, Comparison sheet.",
        "Review the Comparison column.",
        "'Matched' cells have a green fill. 'Not Matched' cells have a red fill.",
    ),
    (
        "FA-21", "Conditions — FIP Data column header", "Whole App",
        "FA-18 or FA-19 complete. Output workbook open, FIP Data sheet.",
        "Check the first column header.",
        "First column is labelled 'Key (Concatenated)'.",
    ),

    # ── Full Run ─────────────────────────────────────────────────────────────
    (
        "FA-22", "Full Run — file fields", "Whole App",
        "Full Run task selected, file-selection form open.",
        "Count the file fields.",
        "All unique fields from every strategy are merged into one form with no "
        "duplicates. Fields include FIP File, X-Checks Publication File, "
        "GCoA Publication File, FIP File (ZQ9_VALFLDGR), Mapping File, "
        "Validation Methods File, FIP File (VALMSG), FIP File (ZQ9_VALMETH), "
        "Known Exception List.",
    ),
    (
        "FA-23", "Full Run — run all strategies", "Logic",
        "Full Run form open. All test files loaded.",
        f"Set all required files using the test_data\\ files:\n"
        f"• FIP File: '{FILES['xc_fip']}'\n"
        f"• X-Checks Publication File: '{FILES['xc_pub']}'\n"
        f"• FIP File (ZQ9_VALFLDGR): '{FILES['gb_fip']}' (sheet: '{SHEETS['gb_fip']}')\n"
        f"• Mapping File: '{FILES['mapping']}'\n"
        f"• Validation Methods File: '{FILES['ap_val_mth']}'\n"
        f"• FIP File (VALMSG): '{FILES['ap_fip']}' (sheet: '{SHEETS['ap_fip']}')\n"
        f"• FIP File (ZQ9_VALMETH): '{FILES['cond_fip']}' (sheet: '{SHEETS['cond_fip']}')\n"
        f"• Known Exception List: '{FILES['known_exc']}'\n"
        "click Proceed.",
        "All strategies run in sequence without error. Progress dialog shows "
        "steps for each strategy. Run completes with 'Processing complete'.",
    ),
    (
        "FA-24", "Full Run — combined output structure", "Whole App",
        "FA-23 complete. Combined output workbook open.",
        "Check all sheet tabs and their tab colours.",
        "One workbook containing all strategy output sheets, each prefixed with "
        "the strategy name (e.g. 'XC — Comparison', 'GB — Comparison', etc.). "
        "Tabs are colour-coded by strategy group. A single 'Processing Log' sheet "
        "at the end contains the combined log.",
    ),

    # ── Settings menu / Known Exception Builder ───────────────────────────────
    (
        "FA-25", "Settings — gear menu", "Whole App",
        "App is open at the task selector.",
        "Click the ⚙ gear button at the bottom-right of the task selector.",
        "A popup menu appears below the button containing at least one entry: "
        "'Build Known Exception List…'. No dialog opens directly.",
    ),
    (
        "FA-26", "Settings — open Known Exception Builder", "Whole App",
        "Settings popup menu open (FA-25).",
        "Click 'Build Known Exception List…'.",
        "The Known Exception Builder dialog opens as a modal window. "
        "It contains: a 'Save as' path field showing hint text "
        "'Click Browse and select a folder, then type the filename', "
        "a Browse button, an optional 'Import from comparison output' section, "
        "an 'Open file after building' checkbox (checked by default), "
        "and a Build button.",
    ),
    (
        "FA-27", "Known Exception Builder — build and open", "Whole App",
        "Known Exception Builder dialog open. An output folder is available.",
        "Click Browse, select an output folder, type a filename (e.g. 'test_kel'). "
        "Leave 'Open file after building' checked. Click Build.",
        "A .xlsx file is created at the chosen path. The dialog closes. "
        "The file opens automatically in Excel. "
        "It contains one sheet per strategy (X-Checks, Grouping By, "
        "Accounting Principles, Conditions) plus an Instructions sheet. "
        "Row 2 of each strategy sheet is a guidance/example row (skipped by the app). "
        "The file carries the 'Internal Use Only' sensitivity label.",
    ),

    # ── Processing Log ────────────────────────────────────────────────────────
    (
        "FA-28", "Processing Log — content", "Whole App",
        "Any completed run. Output workbook open, Processing Log sheet.",
        "Review the log entries.",
        f"First entry shows v{VERSION}. "
        "Log includes: files loaded, strategy-specific steps, output path, "
        "expected sensitivity label (Internal_Use_Only). "
        "All entries have a Timestamp, File, Step, and Count column.",
    ),
    (
        "FA-29", "Processing Log — output path entry", "Whole App",
        "Any completed run. Processing Log sheet open.",
        "Find the 'Output written to' entry.",
        "An entry with File='Output' and Step starting 'Output written to:' "
        "is present in the log. This confirms the path was captured before the "
        "file was closed.",
    ),
    (
        "FA-30", "Processing Log — sensitivity label entry", "Whole App",
        "Any completed run. Processing Log sheet open.",
        "Find the sensitivity label entry.",
        "An entry with File='Sensitivity' and Step='Expected label: Internal_Use_Only' "
        "is present. This records intent; the actual COM result is shown in the "
        "progress dialog only.",
    ),

    # ── Sensitivity label ──────────────────────────────────────────────────────
    (
        "FA-31", "Sensitivity label — applied", "Whole App",
        "Any completed run. Output file saved to disk.",
        "Right-click the output .xlsx in Explorer → Properties → Details, "
        "or open in Excel and check the sensitivity bar.",
        "File carries the 'Internal Use Only' Microsoft Information Protection label. "
        "Progress dialog shows 'Applied label: Internal_Use_Only'.",
    ),

    # ── Stop / error handling ──────────────────────────────────────────────────
    (
        "FA-32", "Stop / Return to Form", "Whole App",
        "Any task started (click Proceed).",
        "Click Stop during processing.",
        "Processing halts cleanly. Progress dialog shows 'Processing halted by user'. "
        "'Return to Form' button available. Clicking it reopens the file-selection "
        "form with previously chosen files pre-filled.",
    ),
    (
        "FA-33", "UI — sheet name constrained to file contents", "Whole App",
        "Any task's file-selection form open. Browse to any Excel file.",
        "After selecting an Excel file, inspect the sheet name dropdown.",
        "The sheet name dropdown is populated with only the sheets present in the selected file. "
        "It is not possible to type or select a sheet name that does not exist in the file — "
        "the combobox is read-only and restricted to the file's own sheet names.",
    ),
    (
        "FA-34", "UI — Proceed button disabled until required files selected", "Whole App",
        "Any task's file-selection form open. No files selected.",
        "Observe the Proceed button before selecting any files. "
        "Then select only optional files. Then select all required files.",
        "Proceed button is disabled (greyed out) while any required file field is empty. "
        "Selecting only optional files does not enable it. "
        "Proceed button becomes enabled only once all required file fields are populated.",
    ),
]

LOGIC_COUNT     = sum(1 for t in TEST_CASES if t[2] == "Logic")
WHOLE_APP_COUNT = sum(1 for t in TEST_CASES if t[2] == "Whole App")


# ---------------------------------------------------------------------------
# OVERVIEW content
# ---------------------------------------------------------------------------

OVERVIEW_ROWS = [
    (
        "Purpose",
        f"End-to-end validation of X-Checks Full Application v{VERSION} against the "
        f"test_data\\ files. Covers all five strategies (Collect Live X-Checks, X-Checks, "
        f"Grouping By, Accounting Principles, Conditions) and Full Run, plus shared "
        f"infrastructure: Processing Log, sensitivity labelling, stop/error handling.\n\n"
        f"Test cases are categorised:\n"
        f"  Logic ({LOGIC_COUNT} cases)     — verifies row counts and processing output. "
        f"Run after every code change to confirm strategy logic is correct.\n"
        f"  Whole App ({WHOLE_APP_COUNT} cases) — verifies UI, structure, labelling, error handling. "
        f"Run for full release sign-off.",
    ),
    (
        "Scope",
        "All strategies in the production EXE, Full Run combined output, Processing Log "
        "content (output path + expected sensitivity label entries), colour coding on "
        "Comparison sheets, and error/stop handling. "
        "Collect Live X-Checks output (clipboard .txt) is noted but its data accuracy "
        "is out of scope for this plan.",
    ),
    (
        "Test executable",
        f"dist\\X-Checks_v{VERSION}.exe (production build).",
    ),
    (
        "Test data",
        f"All files are in test_data\\ (repo root):\n"
        f"• X-Checks pub:            {FILES['xc_pub']}  (sheet: {SHEETS['xc_pub']}, {XC_EBX_ROWS} rows)\n"
        f"• X-Checks FIP:            {FILES['xc_fip']}  ({XC_FIP_ROWS} rows)\n"
        f"• Grouping By FIP:         {FILES['gb_fip']}  (sheet: {SHEETS['gb_fip']}, {GB_FIP_ROWS} rows)\n"
        f"• Grouping By pub:         {FILES['gb_pub']}  (sheet: {SHEETS['gb_pub']})\n"
        f"• Mapping File:            {FILES['mapping']}  ({GB_MAPPING_ROWS} mapping entries)\n"
        f"• AP Validation Methods:   {FILES['ap_val_mth']}  (sheet: {SHEETS['ap_val_mth']})\n"
        f"• AP pub:                  {FILES['ap_pub']}  (sheet: {SHEETS['ap_pub']}, {AP_PUB_ROWS} rows)\n"
        f"• AP FIP (VALMSG):         {FILES['ap_fip']}  (sheet: {SHEETS['ap_fip']}, {AP_FIP_ROWS} rows — raw ZQ9_VALMSG; app builds Key column)\n"
        f"• Conditions pub:          {FILES['cond_pub']}  (sheet: {SHEETS['cond_pub']})\n"
        f"• Conditions FIP (VALMETH):{FILES['cond_fip']}  (sheet: {SHEETS['cond_fip']}, {COND_FIP_ROWS} rows)\n"
        f"• Known Exceptions:        {FILES['known_exc']}  (sheet: {SHEETS['known_exc']}, {KNOWN_EXC_ROWS} rows)",
    ),
    (
        "Pre-conditions",
        "1) Tester is on Windows 10/11 with permission to run unsigned EXEs.\n"
        "2) No test_data files are open in Excel before launching.\n"
        "3) An output folder is available and writable.\n"
        "4) Microsoft Information Protection client is installed (for sensitivity-label checks).",
    ),
    (
        "How to run a test",
        "Read 'Steps', perform them on the running EXE, compare to 'Expected Result'. "
        "Enter the actual outcome in 'Actual Result' and set 'Pass / Fail'. "
        "If failed, add detail (screenshot path, error message, count observed).",
    ),
    (
        "Sign-off",
        "Once all in-scope cases are Pass (or any Fail has been logged and triaged), "
        "complete the Sign-off sheet.",
    ),
]


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

def build():
    wb = Workbook()
    wb.remove(wb.active)
    _build_overview(wb)
    _build_test_cases(wb)
    _build_signoff(wb)
    os.makedirs(OUT_DIR, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"  Written: {OUT_PATH}")


def _build_overview(wb):
    ws = wb.create_sheet("Overview")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 22, "B": 26, "C": 26, "D": 26, "E": 26})

    tc = ws.cell(row=1, column=1,
                 value=f"Full Application v{VERSION} — User Acceptance Testing")
    tc.font = _semibold(size=16)
    tc.alignment = TOP_WRAP
    ws.merge_cells("A1:E1")

    vc = ws.cell(row=2, column=1, value=f"Version: v{VERSION}")
    vc.font = _font(size=11)
    vc.alignment = TOP_WRAP
    ws.merge_cells("A2:E2")

    for i, (label, content) in enumerate(OVERVIEW_ROWS, start=4):
        ws.row_dimensions[i].height = 80
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = _semibold()
        lc.fill = _fill(LIGHT_BLUE)
        lc.border = ALL_THIN
        lc.alignment = TOP_WRAP
        cc = ws.cell(row=i, column=2, value=content)
        cc.font = _font()
        cc.border = ALL_THIN
        cc.alignment = TOP_WRAP
        ws.merge_cells(f"B{i}:E{i}")


def _build_test_cases(wb):
    ws = wb.create_sheet("Test Cases")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {
        "A": 10, "B": 26, "C": 14, "D": 36, "E": 50,
        "F": 50, "G": 36, "H": 12, "I": 18, "J": 14,
    })

    headers = ["ID", "Area", "Test Type", "Precondition", "Steps", "Expected Result",
               "Actual Result", "Pass / Fail", "Tester", "Date"]

    tc = ws.cell(row=1, column=1, value=f"Full Application v{VERSION} — UAT Test Cases")
    tc.font = _semibold(size=16)
    tc.alignment = TOP_WRAP
    ws.merge_cells("A1:J1")

    # Legend row
    ws.row_dimensions[2].height = 18
    leg = ws.cell(row=2, column=1,
                  value="Test Type:   Logic = verify processing output counts     Whole App = verify UI, structure, labelling, error handling")
    leg.font = _font(size=9, color="FF444444")
    leg.alignment = TOP_WRAP
    ws.merge_cells("A2:J2")

    ws.row_dimensions[3].height = 28
    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.font = _semibold(size=11, color=WHITE)
        cell.fill = _fill(DARK_BLUE)
        cell.border = ALL_THIN
        cell.alignment = CENTER_WRAP

    last_data_row = 3 + len(TEST_CASES)
    for row_offset, (tc_id, area, test_type, precond, steps, expected) in enumerate(TEST_CASES):
        row = 4 + row_offset
        ws.row_dimensions[row].height = 80
        for col_idx, val in enumerate(
            [tc_id, area, test_type, precond, steps, expected, "", "", "", ""], start=1
        ):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = _font()
            cell.border = ALL_THIN
            cell.alignment = TOP_WRAP
        ws.cell(row=row, column=3).font = _semibold()

    # Light-blue CF for Logic rows
    cf_range = f"A4:J{last_data_row}"
    cf_fill = PatternFill(patternType="solid", fgColor=LOGIC_BLUE, bgColor=LOGIC_BLUE)
    ws.conditional_formatting.add(
        cf_range,
        FormulaRule(formula=['$C4="Logic"'], fill=cf_fill),
    )


def _build_signoff(wb):
    ws = wb.create_sheet("Sign-off")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 22, "B": 26, "C": 26, "D": 26})

    tc = ws.cell(row=1, column=1, value=f"Full Application v{VERSION} — UAT Sign-off")
    tc.font = _semibold(size=16)
    tc.alignment = TOP_WRAP
    ws.merge_cells("A1:D1")

    ws.row_dimensions[3].height = 30
    ol = ws.cell(row=3, column=1, value="Outcome")
    ol.font = _semibold()
    ol.fill = _fill(LIGHT_BLUE)
    ol.border = ALL_THIN
    ol.alignment = TOP_WRAP
    ov = ws.cell(row=3, column=2,
                 value="All in-scope cases passed (or any failure has been logged and triaged)?")
    ov.font = _font()
    ov.border = ALL_THIN
    ov.alignment = TOP_WRAP
    ws.merge_cells("B3:D3")

    for i, field in enumerate([
        "Tester (name)", "Tester (role)", "Test start date", "Test end date",
        "Pass / Fail", "Failures (count)", "Notes", "Approver (name)",
        "Approver (role)", "Approval date",
    ], start=5):
        ws.row_dimensions[i].height = 26
        lc = ws.cell(row=i, column=1, value=field)
        lc.font = _semibold()
        lc.fill = _fill(LIGHT_BLUE)
        lc.border = ALL_THIN
        lc.alignment = TOP_WRAP
        vc = ws.cell(row=i, column=2, value="")
        vc.font = _font()
        vc.border = ALL_THIN
        vc.alignment = TOP_WRAP
        ws.merge_cells(f"B{i}:D{i}")


if __name__ == "__main__":
    build()
    try:
        import sys as _sys
        _sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from strategies.sensitivity import ExcelLabeler
        labeler = ExcelLabeler()
        ok, msg = labeler.label_file(OUT_PATH, "Internal_Use_Only")
        labeler.close()
        print(f"  Sensitivity label: {msg}")
    except Exception as e:
        print(f"  Sensitivity label skipped: {e}")
