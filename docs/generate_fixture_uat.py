"""
Generate the Fixture-Based UAT Test Plan workbook.

Standalone plan covering the full app using test_data/fixtures/ files.
Each test case is marked as either:
  Logic       — verifies comparison output correctness (row-level results)
  Whole App   — verifies UI, output structure, colour coding, labelling, error handling

Output: docs/<YYYYMMDD> Fixture_UAT_v<version> Test Plan.xlsx

Run from the repo root:
    python docs/generate_fixture_uat.py
"""

import os
import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from version import __version__

VERSION  = __version__
TODAY    = date.today().strftime("%Y%m%d")
OUT_DIR  = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "docs")
FILENAME = f"{TODAY} Fixture_UAT_v{VERSION} Test Plan.xlsx"
OUT_PATH = os.path.join(OUT_DIR, FILENAME)

F = "test_data\\fixtures"

TASKS = [
    "Collect Live X-Checks",
    "X-Checks",
    "Grouping By",
    "Accounting Principles",
    "Conditions",
    "Full Run",
]

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
DARK_BLUE  = "FF23366F"
LIGHT_BLUE = "FF91BFE3"
LOGIC_BLUE = "FFD6E4F7"   # light tint for Logic rows
WHITE      = "FFFFFFFF"
ALT_GREY   = "FFECEEEF"
GOLD       = "FFFFC000"   # whole-app row tint

def _font(size=10, bold=False, color=DARK_BLUE):
    return Font(name="Zurich Sans", size=size, bold=bold, color=color)

def _semibold(size=10, color=DARK_BLUE):
    return Font(name="Zurich Sans Semibold", size=size, bold=True, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

_thin = Side(style="thin")
ALL_THIN  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
TOP_WRAP  = Alignment(vertical="top", wrap_text=True)
CTR_WRAP  = Alignment(horizontal="center", vertical="center", wrap_text=True)

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Test cases
# Format: (ID, Area, Test Type, Files/Setup, Steps, Expected Result)
# Test Type: "Logic" | "Whole App"
# ---------------------------------------------------------------------------

TEST_CASES = [

    # ── Launch & version ──────────────────────────────────────────────────────
    (
        "FX-01", "Launch & version", "Whole App",
        f"Production EXE: dist\\X-Checks_FullRun_{VERSION}.exe",
        f"Double-click dist\\X-Checks_FullRun_{VERSION}.exe.",
        f"Splash screen shows 'X-Check Application v{VERSION} Loading...'. "
        f"Task selector opens with 'X-Check Application v{VERSION}' in title bar and UI label.",
    ),

    # ── Task selector ─────────────────────────────────────────────────────────
    (
        "FX-02", "Task selector", "Whole App",
        "App open at task selector.",
        "Open the task dropdown and count entries.",
        f"{len(TASKS)} tasks listed in order:\n" +
        "\n".join(f"{i+1}. {t}" for i, t in enumerate(TASKS)),
    ),
    (
        "FX-03", "Task selector", "Whole App",
        "App open at task selector.",
        "Select each task in turn and confirm the form title.",
        "• Collect Live X-Checks → 'Collect Live X-Checks'\n"
        "• X-Checks → 'X-Check Files'\n"
        "• Grouping By → 'Grouping By Files'\n"
        "• Accounting Principles → 'Accounting Principles Files'\n"
        "• Conditions → 'Conditions Files'\n"
        "• Full Run → 'Full Run — All Strategies'",
    ),

    # ── X-Checks ─────────────────────────────────────────────────────────────
    (
        "FX-04", "X-Checks — file fields", "Whole App",
        "X-Checks task selected, form open.",
        "Inspect the form fields.",
        "Four fields present:\n"
        "1. 'FIP File' (.txt)\n"
        "2. 'X-Checks Publication File' (.xlsx, default sheet: cross checks all)\n"
        "3. 'GCoA Publication File' (.xlsx, optional)\n"
        "4. 'Known Exception List' (.xlsx, optional)\n"
        "Two experimental checkboxes unchecked by default. "
        "'Process only differences' present and checked by default.",
    ),
    (
        "FX-05", "X-Checks — comparison output", "Logic",
        f"FIP File: {F}\\fip_xc.txt\n"
        f"X-Checks Publication File: {F}\\xc_pub.xlsx  (sheet: cross checks all)\n"
        "No GCoA. No Known Exception List. 'Process only differences' unchecked.",
        "Load the files above into X-Checks and click Start.",
        "Run completes. Comparison sheet contains exactly 30 rows. "
        "Formula Match column values (sorted alphabetically):\n\n"
        "  XC_ABS_FORMULA        - Match   (Operator 2 set - ABS() wrapping)\n"
        "  XC_ALL_MATCH          - Match\n"
        "  XC_ALL_MISMATCH       - MisMatch (formula AND variables wrong)\n"
        "  XC_DIFF_EXCLUDED      - Not Found (Exclude Z-Core=X, no FIP block)\n"
        "  XC_DIFF_GREEN         - Match   (Type of change=New, green fill - in scope)\n"
        "  XC_DIFF_INACTIVE      - Not Found (Status=INACTIVE, no FIP block)\n"
        "  XC_DIFF_IN_SCOPE      - Match   (Type of change=Changed, yellow fill - in scope)\n"
        "  XC_DIFF_NO_TOC        - Not Found (blank Type of change, no FIP block)\n"
        "  XC_DIFF_ORANGE        - Not Found (Type of change=Removed, excluded)\n"
        "  XC_DIFF_YELLOW        - Match   (Type of change=Changed, yellow fill - in scope)\n"
        "  XC_DIFF_YELLOW_CAT    - Not Found (Category yellow, no FIP block)\n"
        "  XC_EXCL_MATCH         - Match   (FIP @2A@, EBX excl - Excl cols Match)\n"
        "  XC_EXCL_MISMATCH      - Match   (Formula Match=Match; Excl=MisMatch - see FX-06c)\n"
        "  XC_FF_SUFFIX          - Match   (two accounts - ff suffix)\n"
        "  XC_FORMULA_MISMATCH   - MisMatch (operator differs <=0 vs >=0; vars same)\n"
        "  XC_GTE_OPERATOR       - Match   (Operator 1 = >=)\n"
        "  XC_KEL_MISMATCH       - MisMatch (without KEL; annotated when KEL supplied)\n"
        "  XC_KEL_NO_MATCH       - MisMatch (KEL entry exists but wrong fingerprint)\n"
        "  XC_LC_CONST           - Match   (Shareholders' Equity + non-zero limit - CONST_LC)\n"
        "  XC_LC_YTD             - Match   (Shareholders' Equity - LC_YTD)\n"
        "  XC_NONZERO_LIMIT      - Match   (Limit 1=100 - CONST(100,...))\n"
        "  XC_NOT_IN_EBX         - Not Found\n"
        "  XC_NOT_IN_FIP         - Not Found\n"
        "  XC_PCT_FORMAT         - Match   (% column=X - percentage right-hand side)\n"
        "  XC_REORDER_MATCH      - MisMatch (known edge case in reorder logic)\n"
        "  XC_REX_CORRECTION     - Match   (FIP uses REX; parser - ToM)\n"
        "  XC_SUBTRACT           - Match   (+ and - operators - subtraction formula)\n"
        "  XC_THOUSANDS_CORR     - Match   (FIP '1.000' stripped to '1000')\n"
        "  XC_TOM_CORRECTION     - Match   (FIP 'TOM' normalised to 'ToM')\n"
        "  XC_VARIABLE_MISMATCH  - Match   (formula matches; FS Account differs)\n\n"
        "X-Check No. column: green for all-Match rows, red for any MisMatch, "
        "orange for Not Found rows.",
    ),
    (
        "FX-06", "X-Checks — column-level checks", "Logic",
        "FX-05 output open, Comparison sheet.",
        "Check the following specific columns on the rows listed.",
        "a) XC_VARIABLE_MISMATCH: Variables Match = MisMatch; Formula Match = Match.\n"
        "b) XC_FORMULA_MISMATCH: Variables Match = Match (operator differs, account same).\n"
        "c) XC_EXCL_MISMATCH: Formula Match = Match; Formula Match (Excl) = MisMatch "
        "(EBX has excl.acc.type=2, FIP has no @2A@ row).\n"
        "d) XC_EXCL_MATCH: Formula Match (Excl) = Match "
        "(both EBX and FIP carry excl.acc.type=2).",
    ),
    (
        "FX-07", "X-Checks — output structure", "Whole App",
        "FX-05 complete. Output workbook open.",
        "Check the sheet tabs.",
        "Workbook contains:\n"
        "1. Comparison\n"
        "2. Processing Log",
    ),
    (
        "FX-08", "X-Checks — colour coding", "Whole App",
        "FX-05 output open, Comparison sheet.",
        "Review fill colours.",
        "Comparison columns: Match → green, MisMatch → red, Not Found → orange.\n"
        "X-Check No. column: green when all 4 columns Match; red when any MisMatch; "
        "orange when any Not Found (and no MisMatch).",
    ),
    (
        "FX-09", "X-Checks — Known Exception annotation", "Logic",
        f"Same files as FX-05. Known Exception List: {F}\\known_exception_list.xlsx  (sheet: X-Checks).\n"
        "The KEL contains 2 entries: one correct fingerprint for XC_KEL_MISMATCH, "
        "one with wrong fingerprint for XC_KEL_NO_MATCH.",
        "Add the Known Exception List file and run.",
        "Run completes. Progress log shows 'Known exceptions loaded  (2)'.\n\n"
        "XC_KEL_MISMATCH: MisMatch cells → 'MisMatch (Excepted)' with blue fill; "
        "X-Check No. cell blue; Known Exception column = 'Test fixture — expected mismatch'.\n\n"
        "XC_KEL_NO_MATCH: Formula Match remains MisMatch (red); Known Exception blank "
        "(fingerprint in KEL does not match the actual row values).\n\n"
        "All other rows: Known Exception column blank; colours unchanged.",
    ),

    (
        "FX-10", "X-Checks — differences mode (X-Check No Selection)", "Logic",
        f"FIP File: {F}\\fip_xc.txt\n"
        f"X-Checks Publication File: {F}\\xc_pub.xlsx  (sheet: cross checks all)\n"
        "'Process only differences' checked.",
        "Check 'Process only differences'. Load files and click Start.",
        "Run completes with two outputs:\n\n"
        "1. .txt file contains exactly 3 X-Check Nos (Yellow=Changed and Green=New are in scope;\n"
        "   Removed=Orange and blank are excluded):\n"
        "     XC_DIFF_IN_SCOPE   (Type of change=Changed, yellow fill)\n"
        "     XC_DIFF_YELLOW     (Type of change=Changed, yellow fill)\n"
        "     XC_DIFF_GREEN      (Type of change=New x-check, green fill)\n\n"
        "2. Comparison sheet shows ONLY those 3 rows:\n"
        "     XC_DIFF_IN_SCOPE - Match\n"
        "     XC_DIFF_YELLOW   - Match\n"
        "     XC_DIFF_GREEN    - Match\n\n"
        "Absent from Comparison:\n"
        "  XC_DIFF_ORANGE    (Type of change=Removed - excluded from processing)\n"
        "  XC_DIFF_NO_TOC    (blank Type of change - not in scope)\n"
        "  XC_DIFF_EXCLUDED  (Exclude Z-Core=X)\n"
        "  XC_DIFF_INACTIVE  (Status=INACTIVE)\n"
        "  XC_DIFF_YELLOW_CAT (Category cell yellow)\n"
        "  All other rows    (no Type of change value)",
    ),

    # ── Grouping By ──────────────────────────────────────────────────────────
    (
        "FX-11", "Grouping By — file fields", "Whole App",
        "Grouping By task selected, form open.",
        "Inspect the form fields.",
        "Four fields present:\n"
        "1. 'FIP File (ZQ9_VALFLDGR)' (.xlsx, default sheet: Sheet1)\n"
        "2. 'X-Checks Publication File' (.xlsx, default sheet: cross checks all)\n"
        "3. 'Mapping File' (.csv / .txt)\n"
        "4. 'Known Exception List' (.xlsx, optional)",
    ),
    (
        "FX-12", "Grouping By — full comparison output (all 14 rows)", "Logic",
        f"FIP File (ZQ9_VALFLDGR): fixtures\\gb\\fip_ZQ9_VALFLDGR.xlsx  (sheet: Sheet1)\n"
        f"X-Checks Publication File: fixtures\\gb\\gb_pub.xlsx  (sheet: cross checks all)\n"
        f"Mapping File: fixtures\\gb\\gb_mapping.txt\n"
        "'Process only differences' unchecked.",
        "Load the files above into Grouping By and click Start.",
        "Comparison sheet contains exactly 14 rows. Verify each row by EBX Key and Result:\n\n"
        "  GB_BLANK_VR|ITEM_A    Not in FIP  (FIP row has blank ValidRule - dropped during FIP processing)\n"
        "  GB_DEDUP|ITEM_A       Not in FIP  (first of two identical X-Check rows; no FIP entry)\n"
        "  GB_DIFF_GREEN|ITEM_A  Matched     (Grouping By cell is green - in scope for diff mode)\n"
        "  GB_DIFF_WHITE|ITEM_A  Not in FIP  (Grouping By cell is white - out of scope for diff mode)\n"
        "  GB_DIFF_YELLOW|ITEM_A Matched     (Grouping By cell is yellow - in scope for diff mode)\n"
        "  GB_IGNORE_FIELD|ITEM_A Not in FIP (FIP field 'GB_IGNORE_FIELD_FIP' maps to 'ignore' - dropped)\n"
        "  GB_KEL_MATCH|ITEM_A   Not in FIP  (KEL entry with correct fingerprint - annotated)\n"
        "  GB_KEL_NO_MATCH|ITEM_A Not in FIP (KEL entry exists but wrong EBX Key - not annotated)\n"
        "  GB_MATCHED|ITEM_A     Matched     (standard match: ValidRule=GB_MATCHED maps via GB_FIP_FIELD to ITEM_A)\n"
        "  GB_MULTI|ITEM_A       Matched     (first of two comma-separated Grouping By values)\n"
        "  GB_MULTI|ITEM_B       Not in FIP  (second comma-separated value; no FIP entry for ITEM_B)\n"
        "  GB_NOT_IN_FIP|ITEM_A  Not in FIP  (EBX key present; no matching FIP row)\n"
        "  GB_UNMAPPED|ITEM_A    Not in FIP  (FIP field 'UNMAPPED_FIELD' not in mapping file - dropped)\n"
        "  REF_BASE|ITEM_A       Matched     (GB_REF_XC_KEY row: 'Reference X-Check (Condition)'=REF_BASE\n"
        "                                     overrides X-Check No. as base key; FIP has REF_BASE|ITEM_A)\n\n"
        "Key logic to verify:\n"
        "- GB_DEDUP second row (Grouping By=ITEM_B) never appears: deduplication keeps only the FIRST row per X-Check No.\n"
        "- GB_IGNORE_FIELD and GB_UNMAPPED test FIP-side filtering: if a field is mapped to 'ignore' or is absent\n"
        "  from the mapping file, the FIP row is dropped and the EBX key becomes Not in FIP.\n"
        "- GB_BLANK_VR tests that FIP rows with a blank ValidRule are dropped during processing.\n"
        "- REF_BASE row confirms that the Reference X-Check (Condition) column replaces X-Check No. as the key prefix.",
    ),
    (
        "FX-13", "Grouping By — Known Exception annotation", "Logic",
        f"Same files as FX-12. Known Exception List: fixtures\\gb\\gb_kel.xlsx  (sheet: Grouping By).\n"
        "KEL contains 2 entries: correct fingerprint for GB_KEL_MATCH|ITEM_A; wrong key for GB_KEL_NO_MATCH.",
        "Add the Known Exception List and run.",
        "GB_KEL_MATCH|ITEM_A: Result = Not in FIP (unchanged); Known Exception column populated with reason text.\n"
        "GB_KEL_NO_MATCH|ITEM_A: Result = Not in FIP; Known Exception column BLANK (wrong fingerprint - no match).\n"
        "All other rows: Known Exception column blank.",
    ),
    (
        "FX-14", "Grouping By — differences mode (Grouping By cell colour)", "Logic",
        f"Same files as FX-12. 'Process only differences' checked.",
        "Check 'Process only differences'. Run.",
        "Comparison sheet shows ONLY rows whose Grouping By cell is yellow or green:\n\n"
        "  GB_DIFF_GREEN|ITEM_A  Matched  (Grouping By cell green - New x-check)\n"
        "  GB_DIFF_YELLOW|ITEM_A Matched  (Grouping By cell yellow - Changed)\n\n"
        "All other rows absent: their Grouping By cells are plain white (unchanged).\n"
        "Confirm: GB_MATCHED, GB_NOT_IN_FIP, GB_MULTI, REF_BASE etc. do NOT appear.",
    ),
    (
        "FX-15", "Grouping By — output structure and colour coding", "Whole App",
        "FX-12 complete. Output workbook open.",
        "Check the sheet tabs and Comparison fill colours.",
        "Workbook contains 7 sheets: Mapping File, FIP - Original, FIP - Processed,\n"
        "EBX - Original, EBX - Processed, Comparison, Processing Log.\n\n"
        "Comparison sheet: 'Matched' rows have green fill; 'Not in FIP' rows have orange fill.",
    ),

    # ── Accounting Principles ─────────────────────────────────────────────────
    (
        "FX-16", "Accounting Principles — file fields", "Whole App",
        "Accounting Principles task selected, form open.",
        "Inspect the form fields.",
        "Four fields present:\n"
        "1. 'Validation Methods File' (.xlsx, default sheet: Validation Methods)\n"
        "2. 'X-Checks Publication File' (.xlsx, default sheet: cross checks all)\n"
        "3. 'FIP File (VALMSG)' (.xlsx, sheet: FIP Methods Rules and Condition)\n"
        "4. 'Known Exception List' (.xlsx, optional)",
    ),
    (
        "FX-17", "Accounting Principles — full comparison output (all rows)", "Logic",
        f"Validation Methods File: fixtures\\validation_methods.xlsx  (sheet: Validation Methods)\n"
        f"X-Checks Publication File: fixtures\\ap\\ap_pub.xlsx  (sheet: cross checks all)\n"
        f"FIP File (VALMSG): fixtures\\ap\\ap_fip_ZQ9_VALMSG.xlsx  (sheet: FIP Methods Rules and Condition)\n"
        "Note: fip_ZQ9_VALMSG.xlsx is a raw ZQ9_VALMSG export with no pre-built Key column.\n"
        "'Process only differences' unchecked.",
        "Load the files above into Accounting Principles and click Start.",
        "Progress log shows 'Built Key column from MK + ValidRule'.\n\n"
        "Comparison sheet contains 11 rows. Verify each (X-Check No., Event, Expected, FIP, Actual, Match):\n\n"
        "  AP_BOTH_E      Stammhaus SLST SFD  Both     e  e  Match    (Both severity: FIP=e matches actual=e)\n"
        "  AP_BOTH_W      Stammhaus SLST RFD  Both     w  w  Match    (Both severity: FIP=w matches actual=w)\n"
        "  AP_DIFF_GREEN  IFRS New RFD        Warning  w  w  Match    (event col green - included in full run)\n"
        "  AP_DIFF_YELLOW IFRS New RFD        Warning  w  w  Match    (event col yellow - included in full run)\n"
        "  AP_EXCL_ZCORE  IFRS New RFD        Warning  w  w  Match    (Exclude Z-Core=X - excluded only in diff mode)\n"
        "  AP_GREY_WINS   IFRS New SFD        Warning  w  w  Match    (black binding column empty; grey binding fires)\n"
        "  AP_MATCH_W     IFRS New RFD        Warning  w  w  Match    (FIP letter w = actual w)\n"
        "  AP_MISMATCH    IFRS New RFD        Warning  e  w  MisMatch (FIP letter e != actual w)\n"
        "  AP_NOT_SCOPE_INA IFRS New RFD      Warning  w  w  Match    (INACTIVE - excluded only in diff mode)\n"
        "  AP_NOT_SCOPE_TOC IFRS New RFD      Warning  w  w  Match    (blank Type of change - excluded only in diff mode)\n"
        "  AP_YELLOW_CAT  IFRS New RFD        Warning  w  w  Match    (yellow Category - excluded only in diff mode)\n\n"
        "Rows NOT appearing (skipped entirely):\n"
        "  AP_NO_BINDING  - FIP sends method V_UNKNOWN_METHOD which has no binding in Validation Methods\n"
        "  AP_NO_ACTUAL   - IFRS New RFD column is blank; no binding produces a non-empty actual letter\n"
        "  AP_DIFF_WHITE  - no event column has any fill colour; excluded in diff mode (not relevant here)\n\n"
        "Key logic to verify:\n"
        "- Both severity rows (AP_BOTH_W, AP_BOTH_E) demonstrate that a single binding accepts either w or e.\n"
        "- AP_GREY_WINS: the black-font binding's event column is blank in ap_pub, so the grey binding fires.\n"
        "  Verify AP_GREY_WINS appears with Event=IFRS New SFD (the grey event), not IFRS New RFD.\n"
        "- AP_NO_BINDING row is absent: the method is not in the Validation Methods file at all.\n"
        "- The 'Process only differences' filter rows (EXCL_ZCORE, NOT_SCOPE_INA, etc.) appear here\n"
        "  because diff mode is OFF - confirming they are correctly included when the checkbox is unchecked.",
    ),
    (
        "FX-18", "Accounting Principles — differences mode (event column colour)", "Logic",
        f"Same files as FX-17. 'Process only differences' checked.",
        "Check 'Process only differences'. Run.",
        "Comparison sheet shows ONLY rows where at least one validation event column\n"
        "is yellow or green in ap_pub.xlsx:\n\n"
        "  AP_DIFF_GREEN  IFRS New RFD  Warning  w  w  Match  (event col green)\n"
        "  AP_DIFF_YELLOW IFRS New RFD  Warning  w  w  Match  (event col yellow)\n\n"
        "All other rows absent: AP_MATCH_W, AP_MISMATCH, AP_BOTH_W, AP_BOTH_E etc.\n"
        "have no coloured event columns and are therefore out of scope.\n\n"
        "Also confirm: AP_NOT_SCOPE_TOC absent (blank Type of change),\n"
        "AP_NOT_SCOPE_INA absent (INACTIVE), AP_EXCL_ZCORE absent (Z-Core=X),\n"
        "AP_YELLOW_CAT absent (yellow Category cell).",
    ),
    (
        "FX-19", "Accounting Principles — Known Exception annotation", "Logic",
        f"Same files as FX-17. Known Exception List: fixtures\\ap\\ap_kel.xlsx  (sheet: Accounting Principles).\n"
        "KEL contains 2 entries: correct 6-column fingerprint for AP_MISMATCH; wrong FIP value for AP_KEL_NO_MATCH.",
        "Add the Known Exception List and run.",
        "AP_MISMATCH: Match = MisMatch (unchanged); Known Exception column populated with reason text.\n"
        "AP_MATCH_W: Known Exception column BLANK (no mismatch - KEL never fires on Match rows).\n"
        "All other rows: Known Exception blank.",
    ),
    (
        "FX-20", "Accounting Principles — output structure and colour coding", "Whole App",
        "FX-17 complete. Output workbook open.",
        "Check the sheet tabs and Comparison fill colours.",
        "Workbook contains 4 sheets: EBX, FIP, Comparison, Processing Log.\n\n"
        "Comparison sheet: 'Match' rows have green fill; 'MisMatch' rows have red fill.",
    ),

    # ── Conditions ────────────────────────────────────────────────────────────
    (
        "FX-21", "Conditions — file fields", "Whole App",
        "Conditions task selected, form open.",
        "Inspect the form fields.",
        "Three fields present:\n"
        "1. 'X-Checks Publication File' (.xlsx, default sheet: cross checks all)\n"
        "2. 'FIP File (ZQ9_VALMETH)' (.xlsx, default sheet: FIP Conditions)\n"
        "3. 'Known Exception List' (.xlsx, optional)\n"
        "'Process only differences' checkbox present and checked by default.",
    ),
    (
        "FX-22", "Conditions — full file run (all 15 rows)", "Logic",
        f"X-Checks Publication File: fixtures\\cond\\cond_pub.xlsx  (sheet: cross checks all)\n"
        f"FIP File (ZQ9_VALMETH): fixtures\\cond\\cond_fip_ZQ9_VALMETH.xlsx  (sheet: FIP Conditions)\n"
        "'Process only differences' unchecked.",
        "Uncheck 'Process only differences'. Load the files and click Start.",
        "Comparison sheet contains exactly 15 rows (one per non-blank condition cell value).\n"
        "Verify each row by EBX Data, FIP Data, and Comparison:\n\n"
        "  COND_APPL_QTRS|Q1        FIP=COND_APPL_QTRS|Q1        Matched   (Applicable Quarters column)\n"
        "  COND_DIFF_GREEN|RU_NORTH FIP=COND_DIFF_GREEN|RU_NORTH  Matched   (Included RUs col, green cell)\n"
        "  COND_DIFF_WHITE|Q1       FIP=(blank)                   Not Matched (white cell - no colour, not in FIP)\n"
        "  COND_DIFF_YELLOW|Q1      FIP=COND_DIFF_YELLOW|Q1       Matched   (Applicable Quarters col, yellow cell)\n"
        "  COND_EXCL_RUS|RU_SOUTH   FIP=COND_EXCL_RUS|RU_SOUTH   Matched   (Excluded RUs column)\n"
        "  COND_INCL_RUS|RU_NORTH   FIP=COND_INCL_RUS|RU_NORTH   Matched   (Included RUs column)\n"
        "  COND_KEL_MISMATCH|Q3     FIP=(blank)                   Not Matched (KEL annotated - see FX-24)\n"
        "  COND_KEL_NO_MATCH|Q4     FIP=(blank)                   Not Matched (wrong KEL fingerprint)\n"
        "  COND_LIMIT_PCT|10.5      FIP=COND_LIMIT_PCT|10.5       Matched   (Reference X-Check (Limit, %) column)\n"
        "  COND_MULTI_COL|Q1        FIP=COND_MULTI_COL|Q1         Matched   (multi-col row: Applicable Quarters)\n"
        "  COND_MULTI_COL|RU_IN     FIP=COND_MULTI_COL|RU_IN      Matched   (multi-col row: Included RUs)\n"
        "  COND_MULTI_COL|RU_OUT    FIP=(blank)                   Not Matched (multi-col row: Excluded RUs - no FIP entry)\n"
        "  COND_NOT_MATCHED|Q2      FIP=(blank)                   Not Matched (Applicable Quarters value not in FIP)\n"
        "  COND_REF_XC|COND_REF_XC FIP=(blank)                   Not Matched (Reference X-Check col value itself)\n"
        "  COND_REF_XC|Q1          FIP=COND_REF_XC|Q1            Matched   (Applicable Quarters using ref override)\n\n"
        "Key logic to verify:\n"
        "- COND_APPL_QTRS, COND_INCL_RUS, COND_EXCL_RUS, COND_LIMIT_PCT each confirm a different one of the\n"
        "  5 condition columns produces output rows.\n"
        "- COND_MULTI_COL produces 3 rows from a single X-Check: one per non-blank condition column.\n"
        "  Q1 and RU_IN are in FIP (Matched); RU_OUT is not (Not Matched).\n"
        "- COND_REF_XC row demonstrates the Reference X-Check (Condition) override:\n"
        "  The pub row has X-Check No.=COND_REF_XC and Reference X-Check (Condition)=COND_REF_XC.\n"
        "  The effective key prefix becomes COND_REF_XC (not the X-Check No. itself).\n"
        "  This produces two rows: COND_REF_XC|COND_REF_XC (col value - Not Matched) and\n"
        "  COND_REF_XC|Q1 (Applicable Quarters with override prefix - Matched).\n"
        "- COND_DIFF_WHITE is Not Matched even in full-file mode: the cell is white but the\n"
        "  value Q1 is not in the FIP file for that X-Check No.",
    ),
    (
        "FX-23", "Conditions — differences mode (condition cell colour)", "Logic",
        f"Same files as FX-22. 'Process only differences' checked (default).",
        "Run with 'Process only differences' checked (the default setting).",
        "Comparison sheet shows ONLY rows where the condition cell itself is yellow or green:\n\n"
        "  COND_DIFF_GREEN|RU_NORTH  FIP=COND_DIFF_GREEN|RU_NORTH  Matched  (Included RUs cell is green)\n"
        "  COND_DIFF_YELLOW|Q1       FIP=COND_DIFF_YELLOW|Q1       Matched  (Applicable Quarters cell is yellow)\n\n"
        "All other rows absent - their condition cells are plain white (unchanged).\n"
        "Confirm: COND_MATCHED, COND_NOT_MATCHED, COND_MULTI_COL rows all absent.\n\n"
        "Note: COND_DIFF_WHITE is also absent because its Applicable Quarters cell has no fill.",
    ),
    (
        "FX-24", "Conditions — Known Exception annotation", "Logic",
        f"Same files as FX-22. Known Exception List: fixtures\\cond\\cond_kel.xlsx  (sheet: Conditions).\n"
        "KEL contains: correct 2-column fingerprint for COND_APPL_QTRS|Q1 (Matched row);\n"
        "wrong FIP Data fingerprint for COND_INCL_RUS|RU_NORTH.\n"
        "Note: Not Matched rows have blank FIP Data so cannot be fingerprint-matched by the KEL.",
        "Add the Known Exception List and run with 'Process only differences' unchecked.",
        "COND_APPL_QTRS|Q1: Comparison = Matched (unchanged); Known Exception column populated with reason text.\n"
        "COND_INCL_RUS|RU_NORTH: Comparison = Matched; Known Exception BLANK (wrong FIP Data fingerprint).\n"
        "All other rows: Known Exception blank.",
    ),
    (
        "FX-25", "Conditions — output structure and colour coding", "Whole App",
        "FX-22 complete. Output workbook open.",
        "Check the sheet tabs, FIP Data first column header, and Comparison fill colours.",
        "Workbook contains 4 sheets: Working Sheet, FIP Data, Comparison, Processing Log.\n"
        "FIP Data sheet: first column header is 'Key (Concatenated)'.\n\n"
        "Comparison sheet: 'Matched' rows have green fill; 'Not Matched' rows have red fill.",
    ),

    # ── Full Run ──────────────────────────────────────────────────────────────
    (
        "FX-26", "Full Run — file fields", "Whole App",
        "Full Run task selected, form open.",
        "Count the file fields and verify no duplicates.",
        "All unique fields from every strategy are merged into one form. "
        "Fields include: FIP File, X-Checks Publication File, GCoA Publication File, "
        "FIP File (ZQ9_VALFLDGR), Mapping File, Validation Methods File, "
        "FIP File (VALMSG), FIP File (ZQ9_VALMETH), Known Exception List. "
        "No field label appears twice.",
    ),
    (
        "FX-27", "Full Run — all strategies", "Logic",
        f"All fixture files. 'Process only differences' unchecked.\n"
        f"• FIP File: {F}\\fip_xc.txt\n"
        f"• X-Checks Publication File: {F}\\xc_pub.xlsx\n"
        f"• FIP File (ZQ9_VALFLDGR): fixtures\\gb\\fip_ZQ9_VALFLDGR.xlsx\n"
        f"• Mapping File: fixtures\\gb\\gb_mapping.txt\n"
        f"• Validation Methods File: {F}\\validation_methods.xlsx\n"
        f"• FIP File (VALMSG): fixtures\\ap\\ap_fip_ZQ9_VALMSG.xlsx\n"
        f"• FIP File (ZQ9_VALMETH): fixtures\\cond\\cond_fip_ZQ9_VALMETH.xlsx",
        "Load all fixture files into Full Run. Uncheck 'Process only differences'. Click Start.",
        "All four strategies run without error. Combined output contains:\n\n"
        "  XC — Comparison    30 rows  (matches FX-05)\n"
        "  GB — Comparison    14 rows  (matches FX-12)\n"
        "  AP — Comparison    11 rows  (matches FX-17)\n"
        "  Cond — Comparison  15 rows  (matches FX-22)\n\n"
        "Single 'Processing Log' sheet at the end.",
    ),
    (
        "FX-28", "Full Run — combined output structure", "Whole App",
        "FX-27 complete. Combined output workbook open.",
        "Check all sheet tabs and tab colours.",
        "One workbook with all strategy sheets prefixed by strategy name "
        "(e.g. 'XC — Comparison', 'GB — Comparison', 'AP — Comparison', 'Cond — Comparison'). "
        "Tabs are colour-coded by strategy. Single 'Processing Log' sheet at the end.",
    ),
    (
        "FX-29", "Full Run — abort on strategy failure", "Whole App",
        "Full Run form open.",
        "Set an incorrect sheet name for one file, then click Start.",
        "Failing strategy logs a clear error. Full Run aborts immediately — "
        "does not continue to the next strategy. 'Return to Form' is available.",
    ),

    # ── Settings / Known Exception Builder ────────────────────────────────────
    (
        "FX-30", "Settings — gear menu", "Whole App",
        "App open at task selector.",
        "Click the gear button at the bottom-right.",
        "A popup menu appears with at least 'Build Known Exception List...'. "
        "No dialog opens directly.",
    ),
    (
        "FX-31", "Settings — Known Exception Builder", "Whole App",
        "Settings popup open (FX-30).",
        "Click 'Build Known Exception List...'.",
        "Modal dialog opens with: 'Save as' hint text "
        "'Click Browse and select a folder, then type the filename', "
        "Browse button, optional comparison import section, "
        "'Open file after building' checkbox (checked by default), Build button.",
    ),
    (
        "FX-32", "Settings — build and open KEL", "Whole App",
        "Known Exception Builder dialog open. Output folder available.",
        "Click Browse, select a folder, type a filename. Leave 'Open file after building' checked. Click Build.",
        "File created at the chosen path. Dialog closes. File opens in Excel. "
        "Contains sheets: X-Checks, Grouping By, Accounting Principles, Conditions, Instructions. "
        "Row 2 of each strategy sheet is a guidance row. "
        "File carries the 'Internal Use Only' sensitivity label.",
    ),

    # ── Processing Log ────────────────────────────────────────────────────────
    (
        "FX-33", "Processing Log — content", "Whole App",
        "Any completed run. Output workbook open, Processing Log sheet.",
        "Review the log entries.",
        f"First entry shows v{VERSION}. "
        "Log includes: files loaded, strategy steps, output path, expected sensitivity label. "
        "All entries have Timestamp, File, Step, Count columns.",
    ),
    (
        "FX-34", "Processing Log — output path entry", "Whole App",
        "Any completed run. Processing Log sheet open.",
        "Find the 'Output written to' entry.",
        "Entry with File='Output' and Step starting 'Output written to:' is present.",
    ),
    (
        "FX-35", "Processing Log — sensitivity label entry", "Whole App",
        "Any completed run. Processing Log sheet open.",
        "Find the sensitivity label entry.",
        "Entry with File='Sensitivity' and Step='Expected label: Internal_Use_Only' is present.",
    ),

    # ── Sensitivity label ──────────────────────────────────────────────────────
    (
        "FX-36", "Sensitivity label", "Whole App",
        "Any completed run. Output .xlsx saved to disk.",
        "Right-click the output file in Explorer, Properties, Details, "
        "or open in Excel and check the sensitivity bar.",
        "File carries the 'Internal Use Only' MIP label. "
        "Progress dialog shows 'Applied label: Internal_Use_Only'.",
    ),

    # ── Stop / error handling ──────────────────────────────────────────────────
    (
        "FX-37", "Stop / Return to Form", "Whole App",
        "Any task started.",
        "Click Stop during processing.",
        "Processing halts. Dialog shows 'Processing halted by user'. "
        "'Return to Form' reopens the form with previously chosen files pre-filled.",
    ),
    (
        "FX-38", "Error — wrong sheet name", "Whole App",
        "Any task's file-selection form open.",
        "Set a sheet name to 'does_not_exist', then click Start.",
        "Run aborts with a clear error identifying the missing sheet. "
        "App returns to form — does not crash.",
    ),
    (
        "FX-39", "Error — missing required file", "Whole App",
        "Any task's file-selection form open.",
        "Click Start without selecting any required files.",
        "Start does not begin processing. Form indicates missing required fields.",
    ),
]

LOGIC_COUNT    = sum(1 for t in TEST_CASES if t[2] == "Logic")
WHOLE_APP_COUNT = sum(1 for t in TEST_CASES if t[2] == "Whole App")

# ---------------------------------------------------------------------------
# Overview
# ---------------------------------------------------------------------------

OVERVIEW_ROWS = [
    (
        "Purpose",
        f"Standalone UAT for X-Checks Full Application v{VERSION} using the minimal "
        f"fixture files in test_data\\fixtures\\. "
        f"Covers all strategies, UI, output structure, colour coding, sensitivity labelling, "
        f"and error handling.\n\n"
        f"Test cases are categorised:\n"
        f"  Logic ({LOGIC_COUNT} cases)     — verifies row-level comparison output. "
        f"Run after every code change to confirm strategy logic is correct.\n"
        f"  Whole App ({WHOLE_APP_COUNT} cases) — verifies UI, structure, labelling, error handling. "
        f"Run for full release sign-off.",
    ),
    (
        "Fixture files",
        f"All in test_data\\fixtures\\:\n"
        "  xc_pub.xlsx               — shared EBX publication (all strategies)\n"
        "  fip_xc.txt                — FIP X-Checks text\n"
        "  fip_ZQ9_VALFLDGR.xlsx     — FIP Grouping By\n"
        "  mapping.txt               — Grouping By mapping\n"
        "  validation_methods.xlsx   — AP Validation Methods (real reference file)\n"
        "  fip_ZQ9_VALMSG.xlsx       — FIP AP (raw ZQ9_VALMSG)\n"
        "  fip_ZQ9_VALMETH.xlsx      — FIP Conditions (raw ZQ9_VALMETH)\n\n"
        "Regenerate (except validation_methods.xlsx) with:\n"
        "  python test_data/generate_test_fixtures.py",
    ),
    (
        "Test executable",
        f"dist\\X-Checks_FullRun_{VERSION}.exe",
    ),
    (
        "Pre-conditions",
        "1) EXE present in dist\\.\n"
        "2) No fixture files open in Excel.\n"
        "3) An output folder is available and writable.\n"
        "4) MIP client installed (for sensitivity-label checks).\n"
        "5) 'Process only differences' unchecked unless the test case specifies otherwise.",
    ),
]


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

def build():
    wb = Workbook()
    wb.remove(wb.active)
    _build_overview(wb)
    _build_test_cases(wb)
    _build_signoff(wb)
    os.makedirs(OUT_DIR, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"  Written: {OUT_PATH}")


def _build_overview(wb):
    ws = wb.create_sheet("Overview")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 22, "B": 100})

    tc = ws.cell(row=1, column=1,
                 value=f"Fixture-Based UAT v{VERSION} — Test Plan")
    tc.font = _semibold(size=16)
    tc.alignment = TOP_WRAP
    ws.merge_cells("A1:B1")

    vc = ws.cell(row=2, column=1,
                 value=f"Version: v{VERSION}  |  {LOGIC_COUNT} Logic cases  |  {WHOLE_APP_COUNT} Whole App cases")
    vc.font = _font(size=11)
    vc.alignment = TOP_WRAP
    ws.merge_cells("A2:B2")

    for i, (label, content) in enumerate(OVERVIEW_ROWS, start=4):
        ws.row_dimensions[i].height = 90
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = _semibold()
        lc.fill = _fill(LIGHT_BLUE)
        lc.border = ALL_THIN
        lc.alignment = TOP_WRAP
        cc = ws.cell(row=i, column=2, value=content)
        cc.font = _font()
        cc.border = ALL_THIN
        cc.alignment = TOP_WRAP


def _build_test_cases(wb):
    ws = wb.create_sheet("Test Cases")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {
        "A": 10, "B": 26, "C": 14, "D": 55, "E": 38,
        "F": 55, "G": 38, "H": 12, "I": 16, "J": 14,
    })

    headers = ["ID", "Area", "Test Type", "Files / Setup", "Steps",
               "Expected Result", "Actual Result", "Pass / Fail", "Tester", "Date"]

    tc = ws.cell(row=1, column=1,
                 value=f"Fixture-Based UAT v{VERSION} — Test Cases")
    tc.font = _semibold(size=16)
    tc.alignment = TOP_WRAP
    ws.merge_cells("A1:J1")

    # Legend row
    ws.row_dimensions[2].height = 18
    leg = ws.cell(row=2, column=1,
                  value="Test Type:   Logic = verify comparison output logic     Whole App = verify UI, structure, labelling, error handling")
    leg.font = _font(size=9, color="FF444444")
    leg.alignment = TOP_WRAP
    ws.merge_cells("A2:J2")

    ws.row_dimensions[3].height = 28
    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.font = _semibold(size=11, color=WHITE)
        cell.fill = _fill(DARK_BLUE)
        cell.border = ALL_THIN
        cell.alignment = CTR_WRAP

    for row_offset, (tc_id, area, test_type, setup, steps, expected) in enumerate(TEST_CASES):
        row = 4 + row_offset
        ws.row_dimensions[row].height = 100
        # Logic rows: light blue tint; Whole App rows: plain white
        row_fill = _fill(LOGIC_BLUE) if test_type == "Logic" else None
        for col_idx, val in enumerate(
            [tc_id, area, test_type, setup, steps, expected, "", "", "", ""], start=1
        ):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = _font()
            cell.border = ALL_THIN
            cell.alignment = TOP_WRAP
            if row_fill:
                cell.fill = row_fill
        # Bold the test-type cell
        ws.cell(row=row, column=3).font = _semibold()


def _build_signoff(wb):
    ws = wb.create_sheet("Sign-off")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, {"A": 22, "B": 26, "C": 26, "D": 26})

    tc = ws.cell(row=1, column=1,
                 value=f"Fixture-Based UAT v{VERSION} — Sign-off")
    tc.font = _semibold(size=16)
    tc.alignment = TOP_WRAP
    ws.merge_cells("A1:D1")

    ws.row_dimensions[3].height = 30
    ol = ws.cell(row=3, column=1, value="Outcome")
    ol.font = _semibold()
    ol.fill = _fill(LIGHT_BLUE)
    ol.border = ALL_THIN
    ol.alignment = TOP_WRAP
    ov = ws.cell(row=3, column=2,
                 value="All in-scope cases passed (or any failure logged and triaged)?")
    ov.font = _font()
    ov.border = ALL_THIN
    ov.alignment = TOP_WRAP
    ws.merge_cells("B3:D3")

    for i, field in enumerate([
        "Tester (name)", "Tester (role)", "Test start date", "Test end date",
        "Pass / Fail", "Failures (count)", "Notes", "Approver (name)",
        "Approver (role)", "Approval date",
    ], start=5):
        ws.row_dimensions[i].height = 26
        lc = ws.cell(row=i, column=1, value=field)
        lc.font = _semibold()
        lc.fill = _fill(LIGHT_BLUE)
        lc.border = ALL_THIN
        lc.alignment = TOP_WRAP
        vc = ws.cell(row=i, column=2, value="")
        vc.font = _font()
        vc.border = ALL_THIN
        vc.alignment = TOP_WRAP
        ws.merge_cells(f"B{i}:D{i}")


if __name__ == "__main__":
    build()
