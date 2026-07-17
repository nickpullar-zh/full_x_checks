"""
Generate the Conditions strategy UAT Test Plan workbook.

Output: docs/<YYYYMMDD> Conditions_v<version> Test Plan.xlsx

Run from the repo root:
    python docs/generate_uat.py
"""

import os
import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from version import __version__

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

VERSION   = __version__
TODAY     = date.today().strftime("%Y%m%d")
OUT_DIR   = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "docs")
FILENAME  = f"{TODAY} Conditions_v{VERSION} Test Plan.xlsx"
OUT_PATH  = os.path.join(OUT_DIR, FILENAME)

# Test files (actual files used for UAT)
PUB_FILE  = "20260313 Cross Checks All.xlsx"
FIP_FILE  = "20260602 VALMETH (Conditions).xlsx"
PUB_SHEET = "cross checks all"
FIP_SHEET = "FIP Conditions"

# Actual counts derived from the test files
FIP_ROW_COUNT          = 6235
DIFF_XCHECKS           = 88
DIFF_PAIRS             = 92
DIFF_MATCHED           = 52
DIFF_NOT_MATCHED       = 40
FULL_XCHECKS           = 92
FULL_PAIRS             = 130
FULL_MATCHED           = 86
FULL_NOT_MATCHED       = 44

# Zurich brand colours
DARK_BLUE  = "FF23366F"
LIGHT_BLUE = "FF91BFE3"
WHITE      = "FFFFFFFF"
ALT_GREY   = "FFECEEEF"
BODY_TEXT  = "FF23366F"

# Fonts
def _font(name="Zurich Sans", size=10, bold=False, color=BODY_TEXT):
    return Font(name=name, size=size, bold=bold, color=color)

def _semibold(size=10, color=BODY_TEXT):
    return Font(name="Zurich Sans Semibold", size=size, bold=True, color=color)

# Fills
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

# Borders
_thin  = Side(style="thin")
_med   = Side(style="medium")
ALL_THIN   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
HDR_BORDER = Border(left=_thin, right=_thin, top=_med, bottom=_med)

# Alignments
TOP_WRAP    = Alignment(vertical="top", wrap_text=True)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP_LEFT    = Alignment(vertical="top", wrap_text=True)

# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def apply_row(ws, row_idx, values, font, fill=None, alignment=TOP_WRAP,
              border=ALL_THIN, height=None):
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = font
        cell.alignment = alignment
        cell.border = border
        if fill:
            cell.fill = fill
    if height:
        ws.row_dimensions[row_idx].height = height


# ---------------------------------------------------------------------------
# TEST CASE DATA
# ---------------------------------------------------------------------------
# Format: (ID, Area, Precondition, Steps, Expected Result)
# ---------------------------------------------------------------------------

TEST_CASES = [
    # --- Launch ---
    (
        "CON-01",
        "Launch & version",
        f"Production EXE present in dist\\.",
        f"Double-click dist\\X-Checks_Conditions_v{VERSION}.exe.",
        f"Splash screen appears with 'X-Check Application v{VERSION} Loading...'. "
        f"After ~2–5 s the task selector form opens with 'X-Check Application v{VERSION}' "
        f"in the title bar.",
    ),
    (
        "CON-02",
        "Launch & version",
        "App is open at the task selector.",
        "Confirm the version shown in the title bar and UI label.",
        f"Both show v{VERSION}. Confirms the correct build is under test.",
    ),
    # --- Task selector ---
    (
        "CON-03",
        "Task selector",
        "App is open at the task selector.",
        "Confirm 'Conditions' appears in the task dropdown.",
        "'Conditions' is listed and selectable. No other tasks appear (this is the Conditions-only build).",
    ),
    (
        "CON-04",
        "Task selector",
        "App is open at the task selector.",
        "Select 'Conditions' and click Proceed.",
        "The Conditions file-selection form opens with the title 'Conditions Files'.",
    ),
    # --- File selection form ---
    (
        "CON-05",
        "File selection — layout",
        "Conditions file-selection form is open.",
        "Inspect the form layout.",
        "Two file fields are shown:\n"
        f"1. 'X-Checks Publication File' with default sheet '{PUB_SHEET}'\n"
        f"2. 'FIP File' with default sheet '{FIP_SHEET}'\n"
        "An output directory picker is present.\n"
        "A 'Process only differences' checkbox is present and checked by default.",
    ),
    (
        "CON-06",
        "File selection — browse",
        "Conditions file-selection form is open.",
        f"Click the Browse button next to 'X-Checks Publication File' and select "
        f"'{PUB_FILE}'.",
        f"The file path populates in the field. The sheet name combobox pre-fills with "
        f"'{PUB_SHEET}' and lists all available sheets.",
    ),
    (
        "CON-07",
        "File selection — browse",
        "CON-06 complete.",
        f"Click the Browse button next to 'FIP File' and select '{FIP_FILE}'. "
        f"Confirm the sheet combobox shows '{FIP_SHEET}'.",
        "Path and sheet both populated correctly.",
    ),
    (
        "CON-08",
        "File selection — validation",
        "Conditions file-selection form is open, no files selected.",
        "Click Start without selecting any files.",
        "Start does not begin processing. The form indicates the missing required fields "
        "and does not crash or close.",
    ),
    (
        "CON-09",
        "File selection — output directory",
        "Both files selected as per CON-06/07.",
        "Set the output directory to a writable folder, then click Start.",
        "Processing begins and the progress dialog opens.",
    ),
    # --- Process only differences (checkbox checked) ---
    (
        "CON-10",
        "Processing — differences only (checkbox checked)",
        f"'Process only differences' checkbox is checked (default). "
        f"Publication file: '{PUB_FILE}' (sheet: '{PUB_SHEET}'). "
        f"FIP file: '{FIP_FILE}' (sheet: '{FIP_SHEET}').",
        "Click Start and wait for the run to complete.",
        "Progress dialog shows steps:\n"
        "1. Loading files\n"
        f"2. Extracting condition cells (changed/new rows only)\n"
        "3. FIP processed\n"
        "4. Comparison complete\n"
        f"Log reports {DIFF_XCHECKS} unique X-Check entries extracted and "
        f"{DIFF_PAIRS} pairs checked.",
    ),
    (
        "CON-11",
        "Processing — differences only (result)",
        "CON-10 complete. Output workbook is open.",
        "Open the 'Conditions' sheet of the output workbook.",
        "Sheet has 3 columns: EBX Data | FIP Data | Comparison.\n"
        f"{DIFF_PAIRS} rows (one per yellow/green condition pair).\n"
        f"Matched (green): {DIFF_MATCHED} rows. Not Matched (red): {DIFF_NOT_MATCHED} rows.\n"
        "Matched rows show the same value in EBX Data and FIP Data. "
        "Not Matched rows have a blank FIP Data cell.",
    ),
    # --- Full file (checkbox unchecked) ---
    (
        "CON-12",
        "Processing — full file (checkbox unchecked)",
        "Return to form. Uncheck 'Process only differences'.",
        "Click Start with the same files as CON-10.",
        "Progress dialog log reports extraction mode 'full file'.\n"
        f"{FULL_XCHECKS} unique X-Check entries extracted "
        f"(more than the {DIFF_XCHECKS} in differences mode).\n"
        "Run completes without error.",
    ),
    (
        "CON-13",
        "Processing — full file (result)",
        "CON-12 complete. Output workbook is open.",
        "Open the 'Conditions' sheet.",
        f"{FULL_PAIRS} rows (more than the {DIFF_PAIRS} rows in differences mode).\n"
        f"Matched (green): {FULL_MATCHED} rows. Not Matched (red): {FULL_NOT_MATCHED} rows.\n"
        "Not Matched rows have a blank FIP Data cell.",
    ),
    # --- Output workbook structure ---
    (
        "CON-14",
        "Output — sheet structure",
        "Either output workbook from CON-11 or CON-13 is open.",
        "Check all sheet tabs.",
        "Workbook contains exactly 4 sheets:\n"
        "1. Conditions\n"
        "2. Working Sheet\n"
        "3. FIP Data\n"
        "4. Processing Log",
    ),
    (
        "CON-15",
        "Output — Working Sheet",
        "Output workbook open.",
        "Open the 'Working Sheet' tab.",
        "Columns: X-Check No. + 5 condition value columns + 5 concat key columns.\n"
        "One row per unique X-Check No.\n"
        "Concat columns contain 'XCheck|ConditionValue' strings or are blank.\n"
        "When a row has a value in 'Reference  X-Check (Condition)', the concat keys "
        "use that reference number, not the row's own X-Check No.",
    ),
    (
        "CON-16",
        "Output — FIP Data",
        "Output workbook open.",
        "Open the 'FIP Data' tab.",
        "First column is labelled 'Key (Concatenated)'.\n"
        "Remaining columns: MethC, MK, Medium Text MK, Normal X-Check No, "
        "X-Check Medium Text, UCFV20G-TRUE_BRANCH, Condition No, Condition Medium Text.\n"
        f"{FIP_ROW_COUNT} data rows (matching source '{FIP_SHEET}' sheet row count).",
    ),
    (
        "CON-17",
        "Output — Processing Log",
        "Output workbook open.",
        "Open the 'Processing Log' tab.",
        f"First entry shows v{VERSION}.\n"
        "Steps logged include: loading files, extraction mode and X-Check count, "
        "FIP row count, comparison summary (pairs / matched / not matched), "
        f"output path, expected sensitivity label (Internal_Use_Only).",
    ),
    (
        "CON-18",
        "Output — colour coding",
        "CON-11 output workbook open, 'Conditions' sheet.",
        "Review the Comparison column.",
        "Cells showing 'Matched' have a green fill. "
        "Cells showing 'Not Matched' have a red fill.",
    ),
    (
        "CON-19",
        "Output — sensitivity label",
        "Output workbook saved to disk.",
        "Right-click the output .xlsx in Explorer → Properties → Details tab, "
        "or open in Excel and check the sensitivity bar.",
        "File carries the 'Internal Use Only' Microsoft Information Protection label. "
        "Progress dialog shows 'Applied label: Internal_Use_Only'.",
    ),
    # --- Data accuracy ---
    (
        "CON-20",
        "Data accuracy — concat key format",
        "CON-11 output workbook open, 'Conditions' sheet.",
        "Inspect any EBX Data value.",
        "Format is 'XCheckNo|ConditionValue' with a pipe separator and no spaces, "
        "e.g. 'S168_00|CON_Q2_Q4'.\n"
        "FIP Data cell contains the identical string when Comparison = Matched.\n"
        "FIP Data cell is blank when Comparison = Not Matched.",
    ),
    (
        "CON-21",
        "Data accuracy — Reference X-Check override",
        "CON-11 output open, 'Working Sheet' tab.",
        "Find any row where 'Reference  X-Check (Condition)' is non-blank.",
        "The concat key for that row uses the Reference X-Check value as the "
        "X-Check identifier, not the row's own X-Check No. value.",
    ),
    (
        "CON-22",
        "Data accuracy — differences count",
        "CON-11 output open, 'Conditions' sheet.",
        f"Count the rows.",
        f"{DIFF_PAIRS} rows total: {DIFF_MATCHED} Matched, {DIFF_NOT_MATCHED} Not Matched.",
    ),
    # --- Stop and return to form ---
    (
        "CON-23",
        "Stop / Return to Form",
        "New run started (click Start).",
        "Click Stop during processing.",
        "Processing halts cleanly. Progress dialog shows 'Processing halted by user'. "
        "A 'Return to Form' button is available.\n"
        "Clicking it reopens the file-selection form with the previously chosen "
        "files pre-filled.",
    ),
    (
        "CON-24",
        "Error handling — wrong sheet name",
        "Conditions file-selection form is open.",
        "Set the sheet name for the X-Checks Publication File to a sheet that does "
        "not exist (e.g. 'does_not_exist'), then click Start.",
        "Run aborts with a clear error message identifying the missing sheet. "
        "The app returns to the form (does not crash or exit).",
    ),
]

# ---------------------------------------------------------------------------
# OVERVIEW content
# ---------------------------------------------------------------------------

OVERVIEW_ROWS = [
    (
        "Purpose",
        f"Validate the Conditions strategy end-to-end against the actual UAT test files. "
        f"The tester confirms each acceptance criterion in the Test Cases sheet and logs "
        f"the actual result and pass/fail.",
    ),
    (
        "Scope",
        "Conditions strategy only: launching the app, selecting files, running in "
        "differences mode (checkbox checked) and full-file mode (checkbox unchecked), "
        "verifying output workbook structure, colour coding, data accuracy (counts and "
        "Reference X-Check override), and error/stop handling. "
        "Other strategies (X-Checks, Accounting Principles, Grouping By) are NOT in scope.",
    ),
    (
        "Test executable",
        f"dist\\X-Checks_Conditions_v{VERSION}.exe (production build, no test data bundled). "
        f"Debug build dist\\X-Checks_Debug_Conditions_v{VERSION}.exe is available "
        f"for developer triage but is not part of the UAT pass.",
    ),
    (
        "Test data",
        f"Publication file: '{PUB_FILE}' (sheet: '{PUB_SHEET}').\n"
        f"FIP file: '{FIP_FILE}' (sheet: '{FIP_SHEET}').\n"
        f"Both files located in the tester's Downloads\\X-Checks Conditions testing\\ folder.",
    ),
    (
        "Pre-conditions",
        "1) Tester is on Windows 10/11 with permission to run unsigned EXEs.\n"
        "2) Neither test file is open in Excel before launching.\n"
        "3) An output folder is available and writable.\n"
        "4) Microsoft Information Protection client is installed (for sensitivity-label check).",
    ),
    (
        "How to run a test",
        "Read 'Steps' in the Test Cases sheet, perform them on the running EXE, and "
        "compare what you observe to 'Expected Result'. Type the actual outcome in "
        "'Actual Result' and set 'Pass / Fail'. If failed, add detail (screenshot path, "
        "error message, row count observed).",
    ),
    (
        "Sign-off",
        "Once all in-scope cases are Pass (or any Fail has been logged and triaged), "
        "complete the Sign-off sheet with name, role, date, and outcome.",
    ),
]


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

def build():
    wb = Workbook()
    wb.remove(wb.active)

    _build_overview(wb)
    _build_test_cases(wb)
    _build_signoff(wb)

    os.makedirs(OUT_DIR, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"  Written: {OUT_PATH}")


# ---------------------------------------------------------------------------
# Sheet 1 — Overview
# ---------------------------------------------------------------------------

def _build_overview(wb):
    ws = wb.create_sheet("Overview")
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 22, "B": 26, "C": 26, "D": 26, "E": 26})

    title_cell = ws.cell(row=1, column=1,
                         value=f"Conditions Application v{VERSION} — User Acceptance Testing")
    title_cell.font      = _semibold(size=16, color=DARK_BLUE)
    title_cell.alignment = TOP_LEFT
    ws.merge_cells("A1:E1")

    ver_cell = ws.cell(row=2, column=1, value=f"Version: v{VERSION}")
    ver_cell.font      = _font(size=11, color=DARK_BLUE)
    ver_cell.alignment = TOP_LEFT
    ws.merge_cells("A2:E2")

    for i, (label, content) in enumerate(OVERVIEW_ROWS, start=4):
        row = i
        ws.row_dimensions[row].height = 70

        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font      = _semibold(size=10, color=DARK_BLUE)
        label_cell.fill      = _fill(LIGHT_BLUE)
        label_cell.border    = ALL_THIN
        label_cell.alignment = TOP_WRAP

        content_cell = ws.cell(row=row, column=2, value=content)
        content_cell.font      = _font(size=10, color=DARK_BLUE)
        content_cell.border    = ALL_THIN
        content_cell.alignment = TOP_WRAP
        ws.merge_cells(f"B{row}:E{row}")


# ---------------------------------------------------------------------------
# Sheet 2 — Test Cases
# ---------------------------------------------------------------------------

def _build_test_cases(wb):
    ws = wb.create_sheet("Test Cases")
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {
        "A": 12, "B": 22, "C": 38, "D": 50, "E": 50,
        "F": 36, "G": 12, "H": 18, "I": 14,
    })

    headers = ["ID", "Area", "Precondition", "Steps",
               "Expected Result", "Actual Result", "Pass / Fail", "Tester", "Date"]

    title_cell = ws.cell(row=1, column=1,
                         value=f"Conditions v{VERSION} — UAT Test Cases")
    title_cell.font      = _semibold(size=16, color=DARK_BLUE)
    title_cell.alignment = TOP_LEFT
    ws.merge_cells("A1:I1")

    ws.row_dimensions[3].height = 28
    hdr_fill = _fill(DARK_BLUE)
    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.font      = _semibold(size=11, color=WHITE)
        cell.fill      = hdr_fill
        cell.border    = ALL_THIN
        cell.alignment = CENTER_WRAP

    for row_offset, (tc_id, area, precond, steps, expected) in enumerate(TEST_CASES):
        row = 4 + row_offset
        ws.row_dimensions[row].height = 80
        fill = _fill(ALT_GREY) if row_offset % 2 == 1 else None

        for col_idx, val in enumerate([tc_id, area, precond, steps, expected, "", "", "", ""], start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font      = _font(size=10, color=DARK_BLUE)
            cell.border    = ALL_THIN
            cell.alignment = TOP_WRAP
            if fill:
                cell.fill = fill


# ---------------------------------------------------------------------------
# Sheet 3 — Sign-off
# ---------------------------------------------------------------------------

def _build_signoff(wb):
    ws = wb.create_sheet("Sign-off")
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 22, "B": 26, "C": 26, "D": 26})

    title_cell = ws.cell(row=1, column=1,
                         value=f"Conditions v{VERSION} — UAT Sign-off")
    title_cell.font      = _semibold(size=16, color=DARK_BLUE)
    title_cell.alignment = TOP_LEFT
    ws.merge_cells("A1:D1")

    ws.row_dimensions[3].height = 30
    out_label = ws.cell(row=3, column=1, value="Outcome")
    out_label.font      = _semibold(size=10, color=DARK_BLUE)
    out_label.fill      = _fill(LIGHT_BLUE)
    out_label.border    = ALL_THIN
    out_label.alignment = TOP_WRAP

    out_val = ws.cell(row=3, column=2,
                      value="All in-scope cases passed (or any failure has been logged and triaged)?")
    out_val.font      = _font(size=10, color=DARK_BLUE)
    out_val.border    = ALL_THIN
    out_val.alignment = TOP_WRAP
    ws.merge_cells("B3:D3")

    fields = [
        "Tester (name)", "Tester (role)", "Test start date", "Test end date",
        "Pass / Fail", "Failures (count)", "Notes", "Approver (name)",
        "Approver (role)", "Approval date",
    ]
    for i, field in enumerate(fields, start=5):
        ws.row_dimensions[i].height = 26

        label = ws.cell(row=i, column=1, value=field)
        label.font      = _semibold(size=10, color=DARK_BLUE)
        label.fill      = _fill(LIGHT_BLUE)
        label.border    = ALL_THIN
        label.alignment = TOP_WRAP

        val = ws.cell(row=i, column=2, value="")
        val.font      = _font(size=10, color=DARK_BLUE)
        val.border    = ALL_THIN
        val.alignment = TOP_WRAP
        ws.merge_cells(f"B{i}:D{i}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    build()
