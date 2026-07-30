"""
Generate minimal test fixture files for all four strategies.

Run:  python test_data/generate_test_fixtures.py

Produces test_data/fixtures/:
  xc_pub.xlsx                 shared EBX publication file (all strategies)
  fip_xc.txt                  FIP X-Checks text (SAP Validation Rule export format)
  fip_ZQ9_VALFLDGR.xlsx       FIP Grouping By (ZQ9_VALFLDGR extract)
  mapping.txt                 Grouping By mapping file
  validation_methods.xlsx     AP Validation Methods (not generated — copy real file)
  fip_ZQ9_VALMSG.xlsx         FIP Accounting Principles (ZQ9_VALMSG extract)
  fip_ZQ9_VALMETH.xlsx        FIP Conditions (ZQ9_VALMETH extract)
  known_exception_list.xlsx   Known Exception List with one XC_KEL_MISMATCH entry

X-Checks fixture design
=======================
Every code path in ebx_extraction.py, fip_extraction.py, and compare.py
that performs a transformation is covered by a dedicated row.

Row                   What it tests                                    Expected
---------             -----------------------------------------        --------
XC_ALL_MATCH          Standard VAL_YTD formula+vars — full match       all Match
XC_ALL_MISMATCH       Formula wrong AND variables wrong                 all MisMatch
XC_FORMULA_MISMATCH   Operator differs (<=0 vs >=0); vars same         Formula=MisMatch, Vars=Match
XC_VARIABLE_MISMATCH  Formula matches; FIP FS Account differs          Formula=Match, Vars=MisMatch
XC_NOT_IN_FIP         EBX row present; no FIP block                    all Not Found
XC_NOT_IN_EBX         FIP block present; no EBX Account No row         all Not Found
XC_TOM_CORRECTION     FIP var/formula uses TOM; parser rewrites→ToM    all Match
XC_REX_CORRECTION     FIP var/formula uses REX; parser rewrites→ToM    all Match
XC_THOUSANDS_CORR     FIP formula has 1.000; _clean_text→1000          all Match
XC_REORDER_MATCH      FIP addition terms reversed; reorder logic       Formula=MisMatch (edge case)
XC_ABS_FORMULA        Operator 2 set → ABS(...) wrapping               all Match
XC_LC_YTD             Category=Shareholders' Equity → LC_YTD           all Match
XC_LC_CONST           LC_YTD with non-zero limit → CONST_LC            all Match
XC_PCT_FORMAT         % column=X → percentage format right-hand side   all Match
XC_FF_SUFFIX          Two accounts, no SubA → ff suffix, one variable  all Match
XC_SUBTRACT           Two accounts, + and - operators                  all Match
XC_NONZERO_LIMIT      Non-zero Limit 1 → CONST(...) right-hand side    all Match
XC_GTE_OPERATOR       Operator 1 = >= (not <=)                         all Match
XC_EXCL_MATCH         Exclude Account Type=2, FIP has @2A@             all Match (incl Excl cols)
XC_EXCL_MISMATCH      Exclude Account Type=2, FIP no excl              Formula=Match, Excl=MisMatch
XC_KEL_MISMATCH       Formula mismatch; KEL entry annotates reason      all MisMatch + Known Exception
XC_DIFF_IN_SCOPE      Status=ACTIVE, Type of change set                In-scope for selection
XC_DIFF_EXCLUDED      Type of change set, Exclude Z-Core=X             Excluded from selection
XC_DIFF_INACTIVE      Status=INACTIVE, Type of change set              Excluded (INACTIVE)
XC_DIFF_YELLOW_CAT    Type of change set, Category cell yellow          Excluded (yellow category)
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font

OUT = Path(__file__).parent / "fixtures"
OUT.mkdir(exist_ok=True)


def _write_rows(ws, header, rows):
    ws.append(header)
    for r in rows:
        ws.append(r)


# ---------------------------------------------------------------------------
# 1. Shared EBX Publication file
# ---------------------------------------------------------------------------

def _make_xc_pub():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "cross checks all"

    headers = [
        "X-Check No.",                    # A  col 1
        "Account No.",                    # B  col 2
        "SubA No.",                       # C  col 3
        "Operator (X-Check Term)",        # D  col 4
        "Absolute (result)",              # E  col 5
        "Category",                       # F  col 6
        "Operator 1",                     # G  col 7
        "Operator 2",                     # H  col 8
        "Limit 1",                        # I  col 9
        "Limit 2",                        # J  col 10
        "%",                              # K  col 11
        "Exclude Account Type",           # L  col 12
        "Version Spanning Validation",    # M  col 13
        "Ending Balance Prior Year",      # N  col 14
        "Grouping By",                    # O  col 15
        "Reference  X-Check (Condition)", # P  col 16  (two spaces — real file header)
        "IFRS New RFD",                   # Q  col 17  (real AP validation event)
        "Applicable Quarters",            # R  col 18
        "Included RUs",                   # S  col 19
        "Excluded RUs",                   # T  col 20
        "Reference X-Check (Limit, %)",   # U  col 21
        "Status",                         # V  col 22
        "Type of change",                 # W  col 23
        "Exclude Z-Core",                 # X  col 24
        "Sense Check",                    # Y  col 25
        "In Scope",                       # Z  col 26
    ]
    ws.append(headers)

    # Track rows that need fills: list of (row_idx, col_idx, fill)
    pending_fills = []

    def row(xc, acct="", suba="", op_term="+", absolute="", category="",
            op1="<=", op2="", lim1="0", lim2="", pct="",
            excl_acc="", vsv="", ebpy="",
            grouping_by="", ref_xc_cond="",
            test_event="",
            app_qtrs="", incl_ru="", excl_ru="", ref_xc_lim="",
            status="ACTIVE", toc="", excl_zcore=""):
        return [
            xc, acct, suba, op_term, absolute, category,
            op1, op2, lim1, lim2, pct,
            excl_acc, vsv, ebpy,
            grouping_by, ref_xc_cond,
            test_event,
            app_qtrs, incl_ru, excl_ru, ref_xc_lim,
            status, toc, excl_zcore, "", "",
        ]

    def append_with_fill(data, col_idx, fill):
        ws.append(data)
        pending_fills.append((ws.max_row, col_idx, fill))

    yellow_fill = openpyxl.styles.PatternFill("solid", fgColor="FFFFFF00")
    green_fill  = openpyxl.styles.PatternFill("solid", fgColor="FF92D050")

    # ==========================================================================
    # X-Checks rows
    # ==========================================================================

    # Standard match/mismatch
    ws.append(row("XC_ALL_MATCH",
                  acct="ACC001", op1="<=", lim1="0"))

    ws.append(row("XC_ALL_MISMATCH",        # FIP has ACC999 → formula AND vars wrong
                  acct="ACC001", op1="<=", lim1="0"))

    ws.append(row("XC_FORMULA_MISMATCH",    # EBX <=0, FIP >=0 → formula wrong, vars same
                  acct="ACC001", op1="<=", lim1="0"))

    ws.append(row("XC_VARIABLE_MISMATCH",   # formula matches, FIP FS Account differs
                  acct="ACC001", op1="<=", lim1="0"))

    ws.append(row("XC_NOT_IN_FIP",          # EBX only → Not Found
                  acct="ACC001", op1="<=", lim1="0"))

    ws.append(row("XC_NOT_IN_EBX"))         # no Account No → skipped by extract_ebx

    # FIP text normalisation
    ws.append(row("XC_TOM_CORRECTION",      # FIP uses TOM; parser → ToM
                  acct="ACC001", suba="AA", op_term="+", op1="<=", lim1="0"))

    ws.append(row("XC_REX_CORRECTION",      # FIP uses REX; parser → ToM
                  acct="ACC001", suba="CC", op_term="+", op1="<=", lim1="0"))

    ws.append(row("XC_THOUSANDS_CORR",      # FIP has 1.000; _clean_text → 1000
                  acct="ACC001", op1="<=", lim1="1000"))

    # Addition-term reorder (two rows, same XC — SubA "AA"/"BB")
    ws.append(row("XC_REORDER_MATCH", acct="A_ACC", suba="AA", op_term="+", op1="<=", lim1="0"))
    ws.append(row("XC_REORDER_MATCH", acct="B_ACC", suba="BB", op_term="+", op1="<=", lim1="0"))

    # EBX formula-building paths
    ws.append(row("XC_ABS_FORMULA",         # Operator 2 = <= with Limit 2 = 5 → ABS(...)
                  acct="ACC001", op1="<=", op2="<=", lim1="0", lim2="5"))

    ws.append(row("XC_LC_YTD",              # Category = Shareholders' Equity → LC_YTD
                  acct="ACC001", category="Shareholders' Equity", op1="<=", lim1="0"))

    ws.append(row("XC_LC_CONST",            # LC_YTD with non-zero limit → CONST_LC
                  acct="ACC001", category="Shareholders' Equity", op1="<=", lim1="100"))

    ws.append(row("XC_PCT_FORMAT",          # % = X → percentage right-hand side
                  acct="ACC001", op1="<", lim1="1.5", pct="X"))

    # ff suffix: two accounts, no SubA → grouped, formula uses ACC001ff
    ws.append(row("XC_FF_SUFFIX", acct="ACC001", op_term="+", op1="<=", lim1="0"))
    ws.append(row("XC_FF_SUFFIX", acct="ACC002", op_term="+", op1="<=", lim1="0"))

    # Subtraction: ACC_POS (+) and ACC_NEG (-), no SubA
    ws.append(row("XC_SUBTRACT", acct="ACC_POS", op_term="+", op1="<=", lim1="0"))
    ws.append(row("XC_SUBTRACT", acct="ACC_NEG", op_term="-", op1="<=", lim1="0"))

    # Non-zero limit
    ws.append(row("XC_NONZERO_LIMIT",       # lim1=100 → CONST(100,'USD','E')
                  acct="ACC001", op1="<=", lim1="100"))

    # >= operator
    ws.append(row("XC_GTE_OPERATOR",        # Operator 1 = >= instead of <=
                  acct="ACC001", op1=">=", lim1="0"))

    # Excl account type: FIP has @2A@ → Formula Match (Excl) = Match
    ws.append(row("XC_EXCL_MATCH",
                  acct="ACC001", excl_acc="2", op1="<=", lim1="0"))

    # Excl account type: FIP has no @2A@ → Formula Match (Excl) = MisMatch
    ws.append(row("XC_EXCL_MISMATCH",
                  acct="ACC001", excl_acc="2", op1="<=", lim1="0"))

    # Known Exception List annotation
    ws.append(row("XC_KEL_MISMATCH",        # MisMatch without KEL; annotated with KEL
                  acct="ACC001", op1="<=", lim1="0"))

    ws.append(row("XC_KEL_NO_MATCH",       # MisMatch; KEL entry exists for this ID but
                  acct="ACC001", op1="<=", lim1="0"))  # with wrong fingerprint → no annotation

    # ==========================================================================
    # Grouping By rows
    # ==========================================================================
    ws.append(row("GB_MATCHED",    grouping_by="GB_GROUPING_ITEM"))
    ws.append(row("GB_NOT_IN_FIP", grouping_by="GB_GROUPING_ITEM"))

    # ==========================================================================
    # Accounting Principles rows
    # ==========================================================================
    ws.append(row("AP_MATCH",    test_event="w"))
    ws.append(row("AP_MISMATCH", test_event="w"))

    # ==========================================================================
    # Conditions rows
    # ==========================================================================
    ws.append(row("COND_MATCHED",     app_qtrs="Q1"))
    ws.append(row("COND_NOT_MATCHED", app_qtrs="Q2"))
    ws.append(row("COND_REF_XCHECK",  ref_xc_cond="COND_BASE", app_qtrs="Q1"))

    # ==========================================================================
    # "Process only differences" / "Test Changes Only" rows
    # ==========================================================================

    # XC: ACTIVE + Type of change "Changed" (yellow fill on ToC cell) → in scope
    ws.append(row("XC_DIFF_IN_SCOPE",   acct="ACC001", op1="<=", lim1="0", toc="Changed"))
    pending_fills.append((ws.max_row, 23, yellow_fill))  # col 23 = Type of change

    # XC: Type of change set BUT Exclude Z-Core=X → excluded
    ws.append(row("XC_DIFF_EXCLUDED",   acct="ACC001", op1="<=", lim1="0",
                  toc="Changed", excl_zcore="X"))
    pending_fills.append((ws.max_row, 23, yellow_fill))

    # XC: INACTIVE + Type of change → excluded (step 1: drop INACTIVE)
    ws.append(row("XC_DIFF_INACTIVE",   acct="ACC001", op1="<=", lim1="0",
                  status="INACTIVE", toc="Changed"))
    pending_fills.append((ws.max_row, 23, yellow_fill))

    # XC: Type of change set BUT Category cell is yellow → excluded
    ws.append(row("XC_DIFF_YELLOW_CAT", acct="ACC001", op1="<=", lim1="0",
                  category="Test Cat", toc="Changed"))
    pending_fills.append((ws.max_row, 6, yellow_fill))   # col 6 = Category
    pending_fills.append((ws.max_row, 23, yellow_fill))  # col 23 = Type of change

    # XC: Type of change "Changed" — yellow fill → in scope; FIP matches → Match
    ws.append(row("XC_DIFF_YELLOW",     acct="ACC001", op1="<=", lim1="0", toc="Changed"))
    pending_fills.append((ws.max_row, 23, yellow_fill))

    # XC: Type of change "New x-check or association" — green fill → in scope; FIP matches
    ws.append(row("XC_DIFF_GREEN",      acct="ACC001", op1="<=", lim1="0",
                  toc="New x-check or association"))
    pending_fills.append((ws.max_row, 23, green_fill))

    # XC: Type of change "Removed" — orange fill → EXCLUDED (removed rows not compared)
    orange_fill = openpyxl.styles.PatternFill("solid", fgColor="FFFF9900")
    ws.append(row("XC_DIFF_ORANGE",     acct="ACC001", op1="<=", lim1="0", toc="Removed"))
    pending_fills.append((ws.max_row, 23, orange_fill))

    # XC: No Type of change value (plain white) → EXCLUDED
    ws.append(row("XC_DIFF_NO_TOC",     acct="ACC001", op1="<=", lim1="0"))

    # Conditions: yellow condition cell → collected in differences mode
    ws.append(row("COND_DIFF_YELLOW", app_qtrs="Q1"))
    pending_fills.append((ws.max_row, 18, yellow_fill))  # col 18 = Applicable Quarters

    # Conditions: green condition cell → collected in differences mode
    ws.append(row("COND_DIFF_GREEN",  app_qtrs="Q1"))
    pending_fills.append((ws.max_row, 18, green_fill))

    # Conditions: plain white cell → NOT collected in differences mode
    ws.append(row("COND_DIFF_WHITE",  app_qtrs="Q1"))

    # Apply all fills now that row indices are stable
    for row_idx, col_idx, fill in pending_fills:
        ws.cell(row=row_idx, column=col_idx).fill = fill

    wb.save(OUT / "xc_pub.xlsx")
    print("  wrote xc_pub.xlsx")


# ---------------------------------------------------------------------------
# 2. FIP X-Checks text
# ---------------------------------------------------------------------------
# Block format notes (what survives _clean_text):
#
#   FS Account:   "| 1 | 2 | 3 | FS Account | <acct> |"
#                  → token[3]="FS", token[5]=<acct>
#
#   Excl type:    "| 0 | 1 | @2A@ | Account | Type | <n> |"
#                  → token[2]="@2A@", token[4]="Type", token[5]=<n>
#
#   Movement type (first):  "| |- Movement Type @20@ <mt> desc |"
#                            → 'Movement Type' in line, token[3]=<mt>
#
#   Variable name: bare name on its own line (no delimiters)
#
# State machine transitions:
#   VARIABLE → BLANK → FS_ACCOUNT
#   FS_ACCOUNT: BLANK saves var, → VARIABLE
#   FS_ACCOUNT: BLOCK_END saves var, breaks (last var in block)

_SEGMENT_END = "|-Segment @28@ * |"
_BLOCK_END   = "-|"
_BLANK       = "|"
_FORMULA_HDR = "|Formula String |"
_VAR_HDR     = "|-Characteristic Sel Opt Attributes Node Characteristic From To |"


def _fip_block_single(xc_id, formula, var_name, fs_accounts,
                      movement_types=None, excl_types=None):
    """One variable, optional movement types, optional @2A@ excl account types."""
    if movement_types is None:
        movement_types = []
    if excl_types is None:
        excl_types = []

    fs_lines   = "".join(f"| 1 | 2 | 3 | FS Account | {a} |\n" for a in fs_accounts)
    excl_lines = "".join(f"| 0 | 1 | @2A@ | Account | Type | {t} |\n" for t in excl_types)

    if movement_types:
        mt_lines = f"| |- Movement Type @20@ {movement_types[0]} desc |\n"
        for mt in movement_types[1:]:
            mt_lines += f"| 1 | 2 | 3 | @20@ | {mt} | desc |\n"
    else:
        mt_lines = ""

    return (
        f"{xc_id} {xc_id} {xc_id} test description\n"
        f"{_FORMULA_HDR}\n{formula}\n{_BLOCK_END}\n"
        f"{_VAR_HDR}\n{var_name}\n{_BLANK}\n"
        f"{fs_lines}{excl_lines}{mt_lines}"
        f"{_BLOCK_END}\n"
        f"{_SEGMENT_END}\n{_BLOCK_END}\n\n"
    )


def _fip_block_two_vars(xc_id, formula,
                        var1_name, var1_fs, var1_mt,
                        var2_name, var2_fs, var2_mt):
    """Two variables each with one movement type."""
    def fs(accts): return "".join(f"| 1 | 2 | 3 | FS Account | {a} |\n" for a in accts)
    def mt(m):     return f"| |- Movement Type @20@ {m} desc |\n"

    return (
        f"{xc_id} {xc_id} {xc_id} test description\n"
        f"{_FORMULA_HDR}\n{formula}\n{_BLOCK_END}\n"
        f"{_VAR_HDR}\n"
        f"{var1_name}\n{_BLANK}\n{fs(var1_fs)}{mt(var1_mt)}"
        f"{_BLANK}\n"           # BLANK in MOV_GENERAL → save var1, back to VARIABLE
        f"{var2_name}\n{_BLANK}\n{fs(var2_fs)}{mt(var2_mt)}"
        f"{_BLOCK_END}\n"       # BLOCK_END → save var2, break
        f"{_SEGMENT_END}\n{_BLOCK_END}\n\n"
    )


def _fip_block_two_vars_no_mt(xc_id, formula, var1_name, var1_fs, var2_name, var2_fs):
    """Two variables with no movement types."""
    def fs(accts): return "".join(f"| 1 | 2 | 3 | FS Account | {a} |\n" for a in accts)

    return (
        f"{xc_id} {xc_id} {xc_id} test description\n"
        f"{_FORMULA_HDR}\n{formula}\n{_BLOCK_END}\n"
        f"{_VAR_HDR}\n"
        f"{var1_name}\n{_BLANK}\n{fs(var1_fs)}"
        f"{_BLANK}\n"           # BLANK in FS_ACCOUNT → save var1, back to VARIABLE
        f"{var2_name}\n{_BLANK}\n{fs(var2_fs)}"
        f"{_BLOCK_END}\n"       # BLOCK_END → save var2, break
        f"{_SEGMENT_END}\n{_BLOCK_END}\n\n"
    )


def _make_fip_xc():
    blocks = []

    # ── Standard match/mismatch ───────────────────────────────────────────────

    blocks.append(_fip_block_single("XC_ALL_MATCH",
        formula="VAL_YTD(ACC001)<=0", var_name="ACC001", fs_accounts=["ACC001"]))

    # XC_ALL_MISMATCH: formula wrong (ACC999) AND FS account wrong (ACC999) → all MisMatch
    blocks.append(_fip_block_single("XC_ALL_MISMATCH",
        formula="VAL_YTD(ACC999)<=0", var_name="ACC999", fs_accounts=["ACC999"]))

    # XC_FORMULA_MISMATCH: operator differs (>=0 vs <=0); account same → vars Match
    blocks.append(_fip_block_single("XC_FORMULA_MISMATCH",
        formula="VAL_YTD(ACC001)>=0", var_name="ACC001", fs_accounts=["ACC001"]))

    # XC_VARIABLE_MISMATCH: formula matches EBX but FS account wrong → vars MisMatch
    blocks.append(_fip_block_single("XC_VARIABLE_MISMATCH",
        formula="VAL_YTD(ACC001)<=0", var_name="ACC_WRONG", fs_accounts=["ACC_WRONG"]))

    # XC_NOT_IN_EBX: FIP block present, no EBX Account No row → Not Found (EBX side)
    blocks.append(_fip_block_single("XC_NOT_IN_EBX",
        formula="VAL_YTD(ACC001)<=0", var_name="ACC001", fs_accounts=["ACC001"]))

    # ── FIP text normalisation ────────────────────────────────────────────────

    # XC_TOM_CORRECTION: FIP uses TOM in var name/formula; parser replaces TOM→ToM
    # EBX SubA="AA" → EBX var=ACC001ToMAA. FIP: ACC001TOMAA → parsed as ACC001ToMAA → Match
    blocks.append(_fip_block_single("XC_TOM_CORRECTION",
        formula="VAL_YTD(ACC001TOMAA)<=0", var_name="ACC001TOMAA",
        fs_accounts=["ACC001"], movement_types=["AA"]))

    # XC_REX_CORRECTION: FIP uses REX in var name/formula; parser replaces REX→ToM
    # EBX SubA="CC" → EBX var=ACC001ToMCC. FIP: ACC001REXCC → parsed as ACC001ToMCC → Match
    blocks.append(_fip_block_single("XC_REX_CORRECTION",
        formula="VAL_YTD(ACC001REXCC)<=0", var_name="ACC001REXCC",
        fs_accounts=["ACC001"], movement_types=["CC"]))

    # XC_THOUSANDS_CORR: FIP has 1.000; _clean_text re.sub(\d.\d)→\d\d strips → 1000
    # EBX Limit 1=1000 → formula VAL_YTD(ACC001)<=CONST(1000,'USD','E') → Match
    blocks.append(_fip_block_single("XC_THOUSANDS_CORR",
        formula="VAL_YTD(ACC001)<=CONST(1.000,'USD','E')",
        var_name="ACC001", fs_accounts=["ACC001"]))

    # XC_REORDER_MATCH: FIP has addition terms in reverse order
    # _compare_formulas detects addition-only with same sorted vars → attempts reorder
    # Known edge case: reorder logic produces invalid formula for simple 2-var addition
    # Expected: Formula Match = MisMatch (not a bug we're fixing here)
    blocks.append(_fip_block_two_vars("XC_REORDER_MATCH",
        formula="VAL_YTD(B_ACCToMBB)+VAL_YTD(A_ACCToMAA)<=0",
        var1_name="A_ACCToMAA", var1_fs=["A_ACC"], var1_mt="AA",
        var2_name="B_ACCToMBB", var2_fs=["B_ACC"], var2_mt="BB"))

    # ── EBX formula-building paths ────────────────────────────────────────────

    # XC_ABS_FORMULA: Operator 2 set → ABS wrapping; EBX → ABS(VAL_YTD(ACC001))<=CONST(5,'USD','E')
    blocks.append(_fip_block_single("XC_ABS_FORMULA",
        formula="ABS(VAL_YTD(ACC001))<=CONST(5,'USD','E')",
        var_name="ACC001", fs_accounts=["ACC001"]))

    # XC_LC_YTD: Category=Shareholders' Equity, Limit 1=0 → LC_YTD(ACC001)<=0
    blocks.append(_fip_block_single("XC_LC_YTD",
        formula="LC_YTD(ACC001)<=0", var_name="ACC001", fs_accounts=["ACC001"]))

    # XC_LC_CONST: Category=Shareholders' Equity, Limit 1=100 → LC_YTD(ACC001)<=CONST_LC(100,'USD','E')
    blocks.append(_fip_block_single("XC_LC_CONST",
        formula="LC_YTD(ACC001)<=CONST_LC(100,'USD','E')",
        var_name="ACC001", fs_accounts=["ACC001"]))

    # XC_PCT_FORMAT: % column=X, Limit 1=1.5, op1=< → VAL_YTD(ACC001)<'1,500000%'
    blocks.append(_fip_block_single("XC_PCT_FORMAT",
        formula="VAL_YTD(ACC001)<'1,500000%'",
        var_name="ACC001", fs_accounts=["ACC001"]))

    # XC_FF_SUFFIX: two accounts ACC001+ACC002, no SubA → variable name ACC001ff
    # FIP: single var ACC001ff with FS accounts [ACC001, ACC002] → Match
    blocks.append(_fip_block_single("XC_FF_SUFFIX",
        formula="VAL_YTD(ACC001ff)<=0",
        var_name="ACC001ff", fs_accounts=["ACC001", "ACC002"]))

    # XC_SUBTRACT: VAL_YTD(ACC_POS)-VAL_YTD(ACC_NEG)<=0 — two separate variables
    blocks.append(_fip_block_two_vars_no_mt("XC_SUBTRACT",
        formula="VAL_YTD(ACC_POS)-VAL_YTD(ACC_NEG)<=0",
        var1_name="ACC_POS", var1_fs=["ACC_POS"],
        var2_name="ACC_NEG", var2_fs=["ACC_NEG"]))

    # XC_NONZERO_LIMIT: Limit 1=100 → CONST(100,'USD','E')
    blocks.append(_fip_block_single("XC_NONZERO_LIMIT",
        formula="VAL_YTD(ACC001)<=CONST(100,'USD','E')",
        var_name="ACC001", fs_accounts=["ACC001"]))

    # XC_GTE_OPERATOR: Operator 1 = >= → VAL_YTD(ACC001)>=0
    blocks.append(_fip_block_single("XC_GTE_OPERATOR",
        formula="VAL_YTD(ACC001)>=0", var_name="ACC001", fs_accounts=["ACC001"]))

    # XC_EXCL_MATCH: FIP has @2A@ Account Type 2 → FIP Formula (Excl) includes excl.acc.type=2
    # EBX Exclude Account Type=2 → EBX Formula (Excl) also includes excl.acc.type=2 → Match
    blocks.append(_fip_block_single("XC_EXCL_MATCH",
        formula="VAL_YTD(ACC001)<=0", var_name="ACC001",
        fs_accounts=["ACC001"], excl_types=["2"]))

    # XC_EXCL_MISMATCH: FIP has NO @2A@ → FIP Formula (Excl) = plain formula
    # EBX Exclude Account Type=2 → EBX Formula (Excl) has excl.acc.type=2 → Excl=MisMatch
    blocks.append(_fip_block_single("XC_EXCL_MISMATCH",
        formula="VAL_YTD(ACC001)<=0", var_name="ACC001", fs_accounts=["ACC001"]))

    # XC_KEL_MISMATCH: formula MisMatch (ACC999); KEL entry annotates with Reason
    blocks.append(_fip_block_single("XC_KEL_MISMATCH",
        formula="VAL_YTD(ACC999)<=0", var_name="ACC999", fs_accounts=["ACC999"]))

    # XC_KEL_NO_MATCH: formula MisMatch (ACC888); KEL has an entry keyed to this
    # X-Check No. but with the WRONG fingerprint values → key does not match →
    # output must remain MisMatch with NO Known Exception annotation.
    blocks.append(_fip_block_single("XC_KEL_NO_MATCH",
        formula="VAL_YTD(ACC888)<=0", var_name="ACC888", fs_accounts=["ACC888"]))

    # ── Differences mode ──────────────────────────────────────────────────────

    # XC_DIFF_IN_SCOPE: Type of change=Changed (yellow); in scope; FIP matches → Match
    blocks.append(_fip_block_single("XC_DIFF_IN_SCOPE",
        formula="VAL_YTD(ACC001)<=0", var_name="ACC001", fs_accounts=["ACC001"]))

    # XC_DIFF_YELLOW: Type of change=Changed (yellow fill); in scope; FIP matches → Match
    blocks.append(_fip_block_single("XC_DIFF_YELLOW",
        formula="VAL_YTD(ACC001)<=0", var_name="ACC001", fs_accounts=["ACC001"]))

    # XC_DIFF_GREEN: Type of change=New x-check (green fill); in scope; FIP matches → Match
    blocks.append(_fip_block_single("XC_DIFF_GREEN",
        formula="VAL_YTD(ACC001)<=0", var_name="ACC001", fs_accounts=["ACC001"]))

    # XC_DIFF_ORANGE (Removed) and XC_DIFF_NO_TOC (blank) are excluded by filter;
    # no FIP blocks needed.

    # XC_DIFF_EXCLUDED, XC_DIFF_INACTIVE, XC_DIFF_YELLOW_CAT:
    # All have Account No so extract_ebx produces a formula.
    # No FIP block → Not Found in Comparison (filter exclusion is independent of comparison).
    # XC_NOT_IN_FIP already tests this path; we don't need duplicate FIP blocks.

    (OUT / "fip_xc.txt").write_text("".join(blocks), encoding="utf-8")
    print("  wrote fip_xc.txt")


# ---------------------------------------------------------------------------
# 3. FIP Grouping By — ZQ9_VALFLDGR format
# ---------------------------------------------------------------------------

def _make_fip_valfldgr():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    _write_rows(ws,
        ["ValidRule", "Long Text", "Field name"],
        [["GB_MATCHED", "Test rule", "GB_FIP_FIELD"]],
    )
    wb.save(OUT / "fip_ZQ9_VALFLDGR.xlsx")
    print("  wrote fip_ZQ9_VALFLDGR.xlsx")


# ---------------------------------------------------------------------------
# 4. Mapping file
# ---------------------------------------------------------------------------

def _make_mapping():
    (OUT / "mapping.txt").write_text(
        "FIP Data,EBX item\nGB_FIP_FIELD,GB_GROUPING_ITEM\n",
        encoding="utf-8",
    )
    print("  wrote mapping.txt")


# ---------------------------------------------------------------------------
# 5. Validation Methods (AP)
# ---------------------------------------------------------------------------
# The real validation_methods.xlsx must be copied into test_data/fixtures/
# manually. Event: 'IFRS New RFD', method V900W (Warning, black font).
def _make_validation_methods():
    print("  skipped validation_methods.xlsx — copy the real file into fixtures/")


# ---------------------------------------------------------------------------
# 6. FIP Accounting Principles — ZQ9_VALMSG raw format
# ---------------------------------------------------------------------------

def _make_fip_valmsg():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FIP Methods Rules and Condition"
    _write_rows(ws,
        ["MethC", "MK", "Medium Text", "ValidRule", "Long Text",
         "UCFV20G-TRUE_BRANCH", "Message class", "Msg.", "MT", "Message Text"],
        [
            ["1", "V900W", "Warning method", "AP_MATCH",    "AP Match",    "X", "CLS", "001", "w", "Test"],
            ["1", "V900W", "Warning method", "AP_MISMATCH", "AP MisMatch", "X", "CLS", "002", "e", "Test"],
        ],
    )
    wb.save(OUT / "fip_ZQ9_VALMSG.xlsx")
    print("  wrote fip_ZQ9_VALMSG.xlsx")


# ---------------------------------------------------------------------------
# 7. FIP Conditions — ZQ9_VALMETH raw format
# ---------------------------------------------------------------------------

def _make_fip_valmeth():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FIP Conditions"
    _write_rows(ws,
        ["MethC", "MK", "Medium Text", "ValidRule",
         "Medium Text", "UCFV20G-TRUE_BRANCH", "ValidRule", "Medium Text"],
        [
            ["1", "MK1", "Test MK", "COND_MATCHED",     "Matched test",     "X", "Q1", "Quarter 1"],
            ["1", "MK1", "Test MK", "COND_BASE",        "Ref XC test",      "X", "Q1", "Quarter 1"],
            ["1", "MK1", "Test MK", "COND_DIFF_YELLOW", "Diff yellow test", "X", "Q1", "Quarter 1"],
            ["1", "MK1", "Test MK", "COND_DIFF_GREEN",  "Diff green test",  "X", "Q1", "Quarter 1"],
        ],
    )
    wb.save(OUT / "fip_ZQ9_VALMETH.xlsx")
    print("  wrote fip_ZQ9_VALMETH.xlsx")


# ---------------------------------------------------------------------------
# 8. Known Exception List
# ---------------------------------------------------------------------------

def _make_known_exception_list():
    """
    Build known_exception_list.xlsx keyed to XC_KEL_MISMATCH.
    Runs the comparison after the other fixtures are written to extract
    the exact 8-column fingerprint values.
    """
    import sys as _sys
    _sys.path.insert(0, str(OUT.parent.parent))
    import pandas as pd
    from strategies.x_checks.ebx_extraction import extract_ebx
    from strategies.x_checks.fip_extraction import extract_fip
    from strategies.x_checks.compare import compare as xc_compare

    ebx_df = pd.read_excel(OUT / "xc_pub.xlsx", sheet_name="cross checks all")
    ebx_results = extract_ebx(ebx_df)
    xc_list = sorted(set(str(x) for x in ebx_df["X-Check No."].tolist()
                         if str(x) not in ("nan", "", "None")))
    fip_text = (OUT / "fip_xc.txt").read_text(encoding="utf-8")
    fip_results = extract_fip(fip_text, xc_list)
    xc_df = pd.DataFrame(xc_compare(ebx_results, fip_results))

    kel_row = xc_df[xc_df["X-Check No."] == "XC_KEL_MISMATCH"].iloc[0]

    fp_cols = [
        "X-Check No.", "EBX Formula", "FIP Formula",
        "EBX Formula (Excl)", "FIP Formula (Excl)",
        "EBX Variables", "FIP Variables", "FIP Variable (Builder)",
    ]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name in ["X-Checks", "Grouping By", "Accounting Principles", "Conditions"]:
        ws = wb.create_sheet(sheet_name)
        if sheet_name == "X-Checks":
            headers = fp_cols + ["Reason", "Added By", "Date Added",
                                  "Resolution Status", "Resolution Notes"]
            ws.append(headers)
            ws.append(["Guidance: do not delete this row"] + [""] * (len(headers) - 1))

            # Row 3: XC_KEL_MISMATCH — correct fingerprint → annotation fires
            data_row = [kel_row[c] for c in fp_cols]
            data_row += ["Test fixture — expected mismatch", "fixture_generator",
                         "2026-07-30", "Open", ""]
            ws.append(data_row)

            # Row 4: XC_KEL_NO_MATCH — correct X-Check No. but WRONG FIP Formula
            # fingerprint will not match the actual comparison row → no annotation
            no_match_row = xc_df[xc_df["X-Check No."] == "XC_KEL_NO_MATCH"].iloc[0]
            wrong_row = [no_match_row[c] for c in fp_cols]
            wrong_row[fp_cols.index("FIP Formula")] = "VAL_YTD(WRONG_ACCOUNT)<=0"
            wrong_row[fp_cols.index("FIP Formula (Excl)")] = "VAL_YTD(WRONG_ACCOUNT)<=0"
            wrong_row += ["Test fixture — wrong fingerprint (should not annotate)",
                          "fixture_generator", "2026-07-30", "Open", ""]
            ws.append(wrong_row)
        else:
            ws.append(["(no entries)"])

    ws_inst = wb.create_sheet("Instructions")
    ws_inst.append(["Row 2 of each strategy sheet is a guidance row skipped by the app."])

    wb.save(OUT / "known_exception_list.xlsx")
    print("  wrote known_exception_list.xlsx")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print(f"Writing fixtures to {OUT}/")
    _make_xc_pub()
    _make_fip_xc()
    _make_fip_valfldgr()
    _make_mapping()
    _make_validation_methods()
    _make_fip_valmsg()
    _make_fip_valmeth()
    _make_known_exception_list()   # must run after xc_pub and fip_xc are written
    print("Done.")
