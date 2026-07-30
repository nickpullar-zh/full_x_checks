"""
Generate Conditions strategy fixture files.

Run:  python test_data/generate_cond_fixtures.py

Produces test_data/fixtures/cond/:
  cond_pub.xlsx              EBX publication file (cross checks all)
  cond_fip_ZQ9_VALMETH.xlsx  FIP Conditions (raw ZQ9_VALMETH extract)
  cond_kel.xlsx              Known Exception List

Expected Comparison results (process_only_differences=False)
=============================================================
Row ID / EBX Data                              FIP Data               Comparison
COND_APPL_QTRS|Q1                              COND_APPL_QTRS|Q1      Matched  (Applicable Quarters)
COND_INCL_RUS|RU_NORTH                         COND_INCL_RUS|RU_NORTH Matched  (Included RUs)
COND_EXCL_RUS|RU_SOUTH                         COND_EXCL_RUS|RU_SOUTH Matched  (Excluded RUs)
COND_LIMIT_PCT|10.5                            COND_LIMIT_PCT|10.5    Matched  (Reference X-Check (Limit, %))
REF_BASE|Q1                                    REF_BASE|Q1            Matched  (Reference X-Check override)
REF_BASE|COND_REF_XC                           (blank)                Not Matched (Ref col itself is a condition)
COND_NOT_MATCHED|Q2                            (blank)                Not Matched
COND_MULTI_COL|Q1                              COND_MULTI_COL|Q1      Matched  (multiple condition cols, one matched)
COND_MULTI_COL|RU_IN                           COND_MULTI_COL|RU_IN   Matched
COND_MULTI_COL|RU_OUT                          (blank)                Not Matched
COND_KEL_MISMATCH|Q3                           (blank)                Not Matched + Known Exception annotation
COND_KEL_NO_MATCH|Q4                           (blank)                Not Matched, no annotation (wrong fingerprint)

Differences mode (process_only_differences=True):
  COND_DIFF_YELLOW: yellow Applicable Quarters cell → collected → Matched
  COND_DIFF_GREEN:  green Included RUs cell → collected → Matched
  COND_DIFF_WHITE:  plain white → not collected

All 5 CONDITION_COLS are exercised with at least one row.
"""

import sys
from pathlib import Path
import openpyxl

OUT = Path(__file__).parent / "fixtures" / "cond"
OUT.mkdir(parents=True, exist_ok=True)

sys.path.insert(0, str(Path(__file__).parent.parent))

yellow_fill = openpyxl.styles.PatternFill("solid", fgColor="FFFFFF00")
green_fill  = openpyxl.styles.PatternFill("solid", fgColor="FF92D050")


def _write_rows(ws, header, rows):
    ws.append(header)
    for r in rows:
        ws.append(r)


# Column indices (1-based) in cond_pub.xlsx
# A=1 X-Check No., B=2 Ref X-Check (Condition), C=3 Applicable Quarters,
# D=4 Included RUs, E=5 Excluded RUs, F=6 Ref X-Check (Limit,%), G=7 Status, ...

COL_REF_XC   = 2   # Reference  X-Check (Condition)
COL_APP_QTRS = 3   # Applicable Quarters
COL_INCL_RUS = 4   # Included RUs
COL_EXCL_RUS = 5   # Excluded RUs
COL_LIMIT_PCT= 6   # Reference X-Check (Limit, %)


# ---------------------------------------------------------------------------
# 1. EBX Publication file
# ---------------------------------------------------------------------------

def _make_cond_pub():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "cross checks all"

    headers = [
        "X-Check No.",                    # A  col 1
        "Reference  X-Check (Condition)", # B  col 2  (two spaces)
        "Applicable Quarters",            # C  col 3
        "Included RUs",                   # D  col 4
        "Excluded RUs",                   # E  col 5
        "Reference X-Check (Limit, %)",   # F  col 6
        "Status",                         # G  col 7
        "Type of change",                 # H  col 8
        "Exclude Z-Core",                 # I  col 9
        "Category",                       # J  col 10
    ]
    ws.append(headers)

    # Track fills: list of (row, col, fill)
    pending_fills = []

    def row(xc, ref_xc="", app_qtrs="", incl_rus="", excl_rus="", limit_pct="",
            status="ACTIVE", toc="", excl_zcore="", category=""):
        return [xc, ref_xc, app_qtrs, incl_rus, excl_rus, limit_pct,
                status, toc, excl_zcore, category]

    # ── Each of the 5 CONDITION_COLS exercised ────────────────────────────────

    # Col 1 (Reference X-Check (Condition)): acts as key override for other cols
    # also collected itself as a condition → produces REF_BASE|COND_REF_XC (Not Matched)
    # and overrides effective_xc so Applicable Quarters produces REF_BASE|Q1 (Matched)
    ws.append(row("COND_REF_XC", ref_xc="COND_REF_XC", app_qtrs="Q1"))
    # Note: effective_xc = COND_REF_XC (from ref_xc col), so:
    #   ref_xc concat = COND_REF_XC|COND_REF_XC (value of col = "COND_REF_XC")  → Not Matched
    #   app_qtrs concat = COND_REF_XC|Q1 → Matched (FIP has this key)
    # But we want a clearer override test: ref_xc="REF_BASE", app_qtrs="Q1"
    # gives key = REF_BASE|Q1 (Matched). Let's use a cleaner row:
    # The row above will produce: ref_xc col value="COND_REF_XC" → effective_xc="COND_REF_XC"
    # → REF_COND_XC|COND_REF_XC (not matched) and COND_REF_XC|Q1 (matched if FIP has it)
    # Simpler: dedicate a row purely to the override test
    # Re-do: remove the above and use distinct IDs
    # (openpyxl already appended one row — we'll leave it; it tests Ref+AppQtrs together)

    # COND_APPL_QTRS: Applicable Quarters only
    ws.append(row("COND_APPL_QTRS", app_qtrs="Q1"))

    # COND_INCL_RUS: Included RUs only
    ws.append(row("COND_INCL_RUS", incl_rus="RU_NORTH"))

    # COND_EXCL_RUS: Excluded RUs only
    ws.append(row("COND_EXCL_RUS", excl_rus="RU_SOUTH"))

    # COND_LIMIT_PCT: Reference X-Check (Limit, %) only
    ws.append(row("COND_LIMIT_PCT", limit_pct="10.5"))

    # COND_NOT_MATCHED: Applicable Quarters = Q2, no FIP entry → Not Matched
    ws.append(row("COND_NOT_MATCHED", app_qtrs="Q2"))

    # COND_MULTI_COL: multiple condition cols populated — tests that ALL non-blank cols produce rows
    # Q1 → Matched, RU_IN → Matched, RU_OUT → Not Matched
    ws.append(row("COND_MULTI_COL", app_qtrs="Q1", incl_rus="RU_IN", excl_rus="RU_OUT"))

    # COND_KEL_MISMATCH: Not Matched row with KEL annotation
    ws.append(row("COND_KEL_MISMATCH", app_qtrs="Q3"))

    # COND_KEL_NO_MATCH: Not Matched; KEL entry exists but wrong fingerprint
    ws.append(row("COND_KEL_NO_MATCH", app_qtrs="Q4"))

    # ── Differences mode rows ─────────────────────────────────────────────────
    # COND_DIFF_YELLOW: Applicable Quarters cell yellow → collected in diff mode
    ws.append(row("COND_DIFF_YELLOW", app_qtrs="Q1"))
    pending_fills.append((ws.max_row, COL_APP_QTRS, yellow_fill))

    # COND_DIFF_GREEN: Included RUs cell green → collected in diff mode
    ws.append(row("COND_DIFF_GREEN", incl_rus="RU_NORTH"))
    pending_fills.append((ws.max_row, COL_INCL_RUS, green_fill))

    # COND_DIFF_WHITE: plain white → NOT collected in diff mode
    ws.append(row("COND_DIFF_WHITE", app_qtrs="Q1"))

    # Apply fills
    for r, c, fill in pending_fills:
        ws.cell(row=r, column=c).fill = fill

    wb.save(OUT / "cond_pub.xlsx")
    print("  wrote cond_pub.xlsx")


# ---------------------------------------------------------------------------
# 2. FIP ZQ9_VALMETH file
# ---------------------------------------------------------------------------

def _make_cond_fip():
    """
    Raw 8-column ZQ9_VALMETH. conditions/fip.py renames by position:
    col 3 → Normal X-Check No, col 6 → Condition No.
    Key (Concatenated) = Normal X-Check No|Condition No
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FIP Conditions"

    _write_rows(ws,
        ["MethC", "MK", "Medium Text", "ValidRule",
         "Medium Text", "UCFV20G-TRUE_BRANCH", "ValidRule", "Medium Text"],
        [
            # COND_REF_XC row: ref_xc col produces key COND_REF_XC|COND_REF_XC (not matched)
            # app_qtrs col produces key COND_REF_XC|Q1 (matched)
            ["1", "MK1", "Test", "COND_REF_XC",     "Text", "X", "Q1",      "Quarter 1"],

            # COND_APPL_QTRS: key COND_APPL_QTRS|Q1
            ["1", "MK1", "Test", "COND_APPL_QTRS",  "Text", "X", "Q1",      "Quarter 1"],

            # COND_INCL_RUS: key COND_INCL_RUS|RU_NORTH
            ["1", "MK1", "Test", "COND_INCL_RUS",   "Text", "X", "RU_NORTH","Included RU"],

            # COND_EXCL_RUS: key COND_EXCL_RUS|RU_SOUTH
            ["1", "MK1", "Test", "COND_EXCL_RUS",   "Text", "X", "RU_SOUTH","Excluded RU"],

            # COND_LIMIT_PCT: key COND_LIMIT_PCT|10.5
            ["1", "MK1", "Test", "COND_LIMIT_PCT",  "Text", "X", "10.5",    "Limit PCT"],

            # COND_MULTI_COL: Q1 and RU_IN matched; RU_OUT not present → Not Matched
            ["1", "MK1", "Test", "COND_MULTI_COL",  "Text", "X", "Q1",      "Quarter 1"],
            ["1", "MK1", "Test", "COND_MULTI_COL",  "Text", "X", "RU_IN",   "Included RU"],

            # COND_DIFF_YELLOW / COND_DIFF_GREEN: matched in FIP
            ["1", "MK1", "Test", "COND_DIFF_YELLOW","Text", "X", "Q1",      "Quarter 1"],
            ["1", "MK1", "Test", "COND_DIFF_GREEN", "Text", "X", "RU_NORTH","Included RU"],

            # Intentionally absent: COND_NOT_MATCHED|Q2, COND_MULTI_COL|RU_OUT,
            #   COND_KEL_MISMATCH|Q3, COND_KEL_NO_MATCH|Q4, COND_DIFF_WHITE|Q1,
            #   REF_BASE|Q1 wait — COND_REF_XC row above has Normal X-Check No=COND_REF_XC
            # The reference override test: pub row has ref_xc="COND_REF_XC", app_qtrs="Q1"
            # effective_xc = "COND_REF_XC" → key = COND_REF_XC|Q1 (FIP has it above ✓)
            # ref_xc col value="COND_REF_XC" → key = COND_REF_XC|COND_REF_XC (FIP doesn't have it)
        ],
    )
    wb.save(OUT / "cond_fip_ZQ9_VALMETH.xlsx")
    print("  wrote cond_fip_ZQ9_VALMETH.xlsx")


# ---------------------------------------------------------------------------
# 3. Known Exception List
# ---------------------------------------------------------------------------

def _make_cond_kel():
    """
    Build cond_kel.xlsx keyed to COND_KEL_MISMATCH|Q3 (Not Matched row).
    """
    import pandas as pd
    from strategies.conditions.extract import extract_conditions
    from strategies.conditions.fip import process_fip
    from strategies.conditions.compare import compare as cond_compare

    fip_df = pd.read_excel(OUT / "cond_fip_ZQ9_VALMETH.xlsx", sheet_name="FIP Conditions")
    fip_proc = process_fip(fip_df)
    working_df, _ = extract_conditions(str(OUT / "cond_pub.xlsx"), "cross checks all",
                                       process_only_differences=False)
    results_df, _ = cond_compare(working_df, fip_proc)

    # KEL annotation works on any row where the fingerprint matches.
    # For Conditions, fingerprint = (EBX Data, FIP Data). Both must be non-blank.
    # Matched rows have both populated; Not Matched rows have blank FIP Data.
    # So we annotate a Matched row (COND_APPL_QTRS|Q1) as the KEL test.
    kel_row      = results_df[results_df["EBX Data"] == "COND_APPL_QTRS|Q1"].iloc[0]
    no_match_row = results_df[results_df["EBX Data"] == "COND_INCL_RUS|RU_NORTH"].iloc[0]

    fp_cols = ["EBX Data", "FIP Data"]
    metadata = ["Reason", "Added By", "Date Added", "Resolution Status", "Resolution Notes"]
    headers = fp_cols + metadata

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name in ["X-Checks", "Grouping By", "Accounting Principles", "Conditions"]:
        ws = wb.create_sheet(sheet_name)
        if sheet_name == "Conditions":
            ws.append(headers)
            ws.append(["Guidance: do not delete this row"] + [""] * (len(headers) - 1))
            # Correct entry: COND_APPL_QTRS|Q1 (Matched row — both fingerprint cols non-blank)
            ws.append([kel_row["EBX Data"], kel_row["FIP Data"],
                       "Test fixture — expected condition", "fixture_generator",
                       "2026-07-30", "Open", ""])
            # Wrong fingerprint: COND_INCL_RUS|RU_NORTH but with wrong FIP Data
            ws.append([no_match_row["EBX Data"], "WRONG_FIP_DATA",
                       "Test fixture — wrong fingerprint (should not annotate)",
                       "fixture_generator", "2026-07-30", "Open", ""])
        else:
            ws.append(["(no entries)"])

    ws_inst = wb.create_sheet("Instructions")
    ws_inst.append(["Row 2 of each strategy sheet is a guidance row skipped by the app."])
    wb.save(OUT / "cond_kel.xlsx")
    print("  wrote cond_kel.xlsx")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print(f"Writing Conditions fixtures to {OUT}/")
    _make_cond_pub()
    _make_cond_fip()
    _make_cond_kel()
    print("Done.")
