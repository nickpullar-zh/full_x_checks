"""
Generate Grouping By strategy fixture files.

Run:  python test_data/generate_gb_fixtures.py

Produces test_data/fixtures/gb/:
  gb_pub.xlsx               EBX publication file (cross checks all)
  gb_fip_ZQ9_VALFLDGR.xlsx  FIP Grouping By (ZQ9_VALFLDGR extract)
  gb_mapping.txt            Field mapping file
  gb_kel.xlsx               Known Exception List

Expected Comparison results
===========================
Row ID                  EBX Key                              Result
GB_MATCHED              GB_MATCHED|ITEM_A                    Matched
GB_NOT_IN_FIP           GB_NOT_IN_FIP|ITEM_A                 Not in FIP
GB_REF_XC_KEY           REF_BASE|ITEM_A                      Matched  (base key from Reference X-Check (Condition))
GB_MULTI_GROUPING_1     GB_MULTI|ITEM_A                      Matched  (first of two split values)
GB_MULTI_GROUPING_2     GB_MULTI|ITEM_B                      Not in FIP  (second split value, no FIP entry)
GB_DEDUP                GB_DEDUP|ITEM_A                      Not in FIP  (only first EBX row used; second row ignored)
GB_KEL_MATCH            GB_KEL_MATCH|ITEM_A                  Not in FIP + Known Exception annotation
GB_KEL_NO_MATCH         GB_KEL_NO_MATCH|ITEM_A               Not in FIP, no annotation (wrong KEL fingerprint)

Paths not producing Comparison rows (FIP-side filtering):
  IGNORE_FIELD_ROW  — FIP Field name mapped to "ignore" → FIP row dropped
  UNMAPPED_FIELD    — FIP Field name absent from mapping → FIP row dropped
  BLANK_VALIDRULE   — FIP row has blank ValidRule → FIP row dropped
These all produce "Not in FIP" for their corresponding EBX keys (EBX rows still present).

GB_DEDUP: Two EBX rows share X-Check No. GB_DEDUP — only the first is kept.
The second row's key (GB_DEDUP|ITEM_B) never appears in the Comparison sheet.
"""

import sys
from pathlib import Path
import openpyxl

OUT = Path(__file__).parent / "fixtures" / "gb"
OUT.mkdir(parents=True, exist_ok=True)

sys.path.insert(0, str(Path(__file__).parent.parent))


def _write_rows(ws, header, rows):
    ws.append(header)
    for r in rows:
        ws.append(r)


# ---------------------------------------------------------------------------
# 1. EBX Publication file
# ---------------------------------------------------------------------------

def _make_gb_pub():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "cross checks all"

    headers = [
        "X-Check No.",                    # A
        "Grouping By",                    # B
        "Reference  X-Check (Condition)", # C  (two spaces — real file header)
        "Status",                         # D
        "Type of change",                 # E
        "Exclude Z-Core",                 # F
        "Category",                       # G
    ]
    ws.append(headers)

    def row(xc, gb="", ref_xc="", status="ACTIVE", toc="", excl_zcore="", category=""):
        return [xc, gb, ref_xc, status, toc, excl_zcore, category]

    # GB_MATCHED: standard match — key GB_MATCHED|ITEM_A found in FIP
    ws.append(row("GB_MATCHED",     gb="ITEM_A"))

    # GB_NOT_IN_FIP: EBX key present but no FIP row for it
    ws.append(row("GB_NOT_IN_FIP",  gb="ITEM_A"))

    # GB_REF_XC_KEY: Reference X-Check (Condition) overrides base key → key = REF_BASE|ITEM_A
    ws.append(row("GB_REF_XC_KEY",  gb="ITEM_A", ref_xc="REF_BASE"))

    # GB_MULTI_GROUPING: two comma-separated values → two rows in Comparison
    # ITEM_A has FIP entry (Matched); ITEM_B does not (Not in FIP)
    ws.append(row("GB_MULTI",       gb="ITEM_A, ITEM_B"))

    # GB_DEDUP: two rows with same X-Check No. — only first is used by strategy
    # First row: Grouping By=ITEM_A  → produces GB_DEDUP|ITEM_A in Comparison
    # Second row: Grouping By=ITEM_B → IGNORED (deduplication keeps first only)
    ws.append(row("GB_DEDUP",       gb="ITEM_A"))
    ws.append(row("GB_DEDUP",       gb="ITEM_B"))  # deduplicated out — never appears

    # Rows to test FIP-side filtering — EBX keys will be Not in FIP because
    # their corresponding FIP rows are dropped during FIP processing
    ws.append(row("GB_IGNORE_FIELD",  gb="ITEM_A"))  # FIP field mapped to "ignore"
    ws.append(row("GB_UNMAPPED",      gb="ITEM_A"))  # FIP field not in mapping
    ws.append(row("GB_BLANK_VR",      gb="ITEM_A"))  # FIP row has blank ValidRule

    # KEL tests
    ws.append(row("GB_KEL_MATCH",   gb="ITEM_A"))  # KEL entry matches → annotated
    ws.append(row("GB_KEL_NO_MATCH",gb="ITEM_A"))  # KEL entry exists but wrong key → no annotation

    # Process only differences: colour on Grouping By cell drives inclusion
    ws.append(row("GB_DIFF_YELLOW", gb="ITEM_A"))  # yellow Grouping By cell → in scope
    ws.append(row("GB_DIFF_GREEN",  gb="ITEM_A"))  # green Grouping By cell → in scope
    ws.append(row("GB_DIFF_WHITE",  gb="ITEM_A"))  # white Grouping By cell → excluded

    # Apply fills to Grouping By column (col B = index 2)
    yellow_fill = openpyxl.styles.PatternFill("solid", fgColor="FFFFFF00")
    green_fill  = openpyxl.styles.PatternFill("solid", fgColor="FF92D050")
    total_rows = ws.max_row
    ws.cell(row=total_rows - 2, column=2).fill = yellow_fill  # GB_DIFF_YELLOW
    ws.cell(row=total_rows - 1, column=2).fill = green_fill   # GB_DIFF_GREEN

    wb.save(OUT / "gb_pub.xlsx")
    print("  wrote gb_pub.xlsx")


# ---------------------------------------------------------------------------
# 2. FIP ZQ9_VALFLDGR file
# ---------------------------------------------------------------------------

def _make_gb_fip():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # FIP Key = ValidRule|EBX Item (after mapping Field name → EBX Item)
    # Row structure: ValidRule, Long Text, Field name
    _write_rows(ws,
        ["ValidRule", "Long Text", "Field name"],
        [
            # GB_MATCHED: ValidRule=GB_MATCHED + Field name=GB_FIP_FIELD → maps to ITEM_A
            # FIP Key = GB_MATCHED|ITEM_A → matches EBX key → Matched
            ["GB_MATCHED",    "Standard match",     "GB_FIP_FIELD"],

            # GB_REF_XC_KEY: ValidRule=REF_BASE (the reference key) + ITEM_A
            # FIP Key = REF_BASE|ITEM_A → matches EBX key REF_BASE|ITEM_A → Matched
            ["REF_BASE",      "Ref XC base key",    "GB_FIP_FIELD"],

            # GB_MULTI: only ITEM_A has a FIP entry, ITEM_B does not
            ["GB_MULTI",      "Multi grouping A",   "GB_FIP_FIELD"],

            # GB_DIFF_YELLOW / GB_DIFF_GREEN: in scope when diff=True; FIP matches → Matched
            ["GB_DIFF_YELLOW", "Diff yellow match",  "GB_FIP_FIELD"],
            ["GB_DIFF_GREEN",  "Diff green match",   "GB_FIP_FIELD"],
            # GB_DIFF_WHITE: no FIP entry needed (excluded by diff filter)

            # GB_IGNORE_FIELD: Field name maps to "ignore" → this row dropped during FIP processing
            ["GB_IGNORE_FIELD","Ignore field test",  "GB_IGNORE_FIELD_FIP"],

            # GB_UNMAPPED: Field name not in mapping → row dropped during FIP processing
            ["GB_UNMAPPED",   "Unmapped field test", "UNMAPPED_FIELD"],

            # GB_BLANK_VR: blank ValidRule → row dropped during FIP processing
            ["",              "Blank ValidRule test","GB_FIP_FIELD"],
        ],
    )
    wb.save(OUT / "gb_fip_ZQ9_VALFLDGR.xlsx")
    print("  wrote gb_fip_ZQ9_VALFLDGR.xlsx")


# ---------------------------------------------------------------------------
# 3. Mapping file
# ---------------------------------------------------------------------------

def _make_gb_mapping():
    lines = [
        "FIP Data,EBX item",
        "GB_FIP_FIELD,ITEM_A",          # standard mapping
        "GB_IGNORE_FIELD_FIP,ignore",   # maps to "ignore" → FIP row dropped
        # UNMAPPED_FIELD is deliberately absent → FIP row dropped
    ]
    (OUT / "gb_mapping.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")
    print("  wrote gb_mapping.txt")


# ---------------------------------------------------------------------------
# 4. Known Exception List
# ---------------------------------------------------------------------------

def _make_gb_kel():
    """
    Build gb_kel.xlsx with:
    - One correct entry for GB_KEL_MATCH (EBX Key = GB_KEL_MATCH|ITEM_A)
    - One wrong-fingerprint entry for GB_KEL_NO_MATCH (EBX Key deliberately wrong)
    """
    import pandas as pd
    from strategies.grouping_by.grouping_by import GroupingBy
    from task_configs import GROUPING_BY_UPLOAD_CONFIG

    gb = GroupingBy(GROUPING_BY_UPLOAD_CONFIG)
    gb.log = []
    mapping_txt = (OUT / "gb_mapping.txt").read_text()
    fip_gb = pd.read_excel(OUT / "gb_fip_ZQ9_VALFLDGR.xlsx", sheet_name="Sheet1")
    ebx_gb = pd.read_excel(OUT / "gb_pub.xlsx", sheet_name="cross checks all")

    loaded = {
        GROUPING_BY_UPLOAD_CONFIG.file_fields[0].label: fip_gb,
        GROUPING_BY_UPLOAD_CONFIG.file_fields[1].label: ebx_gb,
        GROUPING_BY_UPLOAD_CONFIG.file_fields[2].label: mapping_txt,
    }
    _, _, df_fip_proc = gb._process_fip(loaded)
    _, df_ebx_proc = gb._process_ebx(loaded)
    df_cmp = gb._process_compare(df_fip_proc, df_ebx_proc)

    kel_row = df_cmp[df_cmp["EBX Key"] == "GB_KEL_MATCH|ITEM_A"].iloc[0]
    no_match_row = df_cmp[df_cmp["EBX Key"] == "GB_KEL_NO_MATCH|ITEM_A"].iloc[0]

    fp_cols = ["EBX Key"]
    metadata_cols = ["Reason", "Added By", "Date Added", "Resolution Status", "Resolution Notes"]
    headers = fp_cols + metadata_cols

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name in ["X-Checks", "Grouping By", "Accounting Principles", "Conditions"]:
        ws = wb.create_sheet(sheet_name)
        if sheet_name == "Grouping By":
            ws.append(headers)
            ws.append(["Guidance: do not delete this row"] + [""] * (len(headers) - 1))
            # Correct entry for GB_KEL_MATCH
            ws.append([kel_row["EBX Key"],
                       "Test fixture — expected Not in FIP",
                       "fixture_generator", "2026-07-30", "Open", ""])
            # Wrong-fingerprint entry for GB_KEL_NO_MATCH — EBX Key deliberately wrong
            ws.append(["GB_KEL_NO_MATCH|WRONG_KEY",
                       "Test fixture — wrong fingerprint (should not annotate)",
                       "fixture_generator", "2026-07-30", "Open", ""])
        else:
            ws.append(["(no entries)"])

    ws_inst = wb.create_sheet("Instructions")
    ws_inst.append(["Row 2 of each strategy sheet is a guidance row skipped by the app."])

    wb.save(OUT / "gb_kel.xlsx")
    print("  wrote gb_kel.xlsx")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print(f"Writing Grouping By fixtures to {OUT}/")
    _make_gb_pub()
    _make_gb_fip()
    _make_gb_mapping()
    _make_gb_kel()
    print("Done.")
