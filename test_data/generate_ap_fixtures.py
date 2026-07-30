"""
Generate Accounting Principles strategy fixture files.

Run:  python test_data/generate_ap_fixtures.py

Produces test_data/fixtures/ap/:
  ap_pub.xlsx               EBX publication file (cross checks all)
  ap_fip_ZQ9_VALMSG.xlsx    FIP VALMSG (raw ZQ9_VALMSG format — Key built at load time)
  ap_kel.xlsx               Known Exception List

Validation Methods: uses the shared test_data/fixtures/validation_methods.xlsx
  which must be the real Zurich file. Event used: 'IFRS New RFD',
  methods V900W (Warning, black font) and V900A (Error, black font).

Expected Comparison results
===========================
Row ID          Event           FIP  Actual  Match
AP_MATCH_W      IFRS New RFD    w    w       Match    (FIP=Warning, actual=Warning)
AP_MISMATCH     IFRS New RFD    e    w       MisMatch (FIP=Error, actual=Warning)
AP_BOTH_W       IFRS New SFD    w    w       Match    (severity=Both, FIP=w matches)
AP_BOTH_E       IFRS New CFD    e    e       Match    (severity=Both, FIP=e matches)
AP_GREY_WINS    SST RFD         w    w       Match    (black binding col empty → grey wins)

Rows NOT producing Comparison output (skipped):
  AP_NO_BINDING       — FIP method (V_UNKNOWN) not in Validation Methods
  AP_NO_ACTUAL        — EBX column for event is blank → no winning binding
  AP_NOT_IN_SCOPE_TOC — Type of change blank (process_only_differences=True)
  AP_NOT_IN_SCOPE_INA — Status=INACTIVE
  AP_EXCL_ZCORE       — Exclude Z-Core=X
  AP_YELLOW_CAT       — Category cell yellow
"""

import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill

OUT = Path(__file__).parent / "fixtures" / "ap"
OUT.mkdir(parents=True, exist_ok=True)

sys.path.insert(0, str(Path(__file__).parent.parent))

VM_PATH = str(Path(__file__).parent / "fixtures" / "validation_methods.xlsx")


def _probe_bindings():
    """Return first few bindings from the real validation_methods.xlsx for reference."""
    from strategies.accounting_principles.validation_methods import (
        parse_method_bindings, list_all_event_names
    )
    events = list_all_event_names(VM_PATH)
    bindings = parse_method_bindings(VM_PATH, events)
    # Index by (method, event, severity, font)
    return bindings, events


def _write_rows(ws, header, rows):
    ws.append(header)
    for r in rows:
        ws.append(r)


# ---------------------------------------------------------------------------
# Discover real method codes from validation_methods.xlsx
# ---------------------------------------------------------------------------

def _get_event_methods():
    """
    Return a dict of {event: {severity: [methods]}} for the events we need.
    We need events with:
    - A black Warning binding (for AP_MATCH_W, AP_MISMATCH)
    - A Both binding (for AP_BOTH_W, AP_BOTH_E)
    - A grey binding (for AP_GREY_WINS) — event where black col is empty in pub
    """
    from strategies.accounting_principles.validation_methods import (
        parse_method_bindings, list_all_event_names
    )
    events = list_all_event_names(VM_PATH)
    bindings = parse_method_bindings(VM_PATH, events)

    result = {}
    for b in bindings:
        result.setdefault(b.event, {}).setdefault((b.severity, b.font), []).append(b.method)
    return result, bindings, events


# ---------------------------------------------------------------------------
# 1. EBX Publication file
# ---------------------------------------------------------------------------

def _make_ap_pub(events_info):
    """
    We need columns for several Validation Events from the real file.
    Events used:
      IFRS New RFD  — has black Warning (V900W) and black Error (V900A, V900B)
      IFRS New SFD  — has Both severity
      IFRS New CFD  — has Both severity
      SST RFD       — has a grey binding we can exercise by leaving black col blank

    EBX rows:
      AP_MATCH_W     — IFRS New RFD = 'w'  (Warning match)
      AP_MISMATCH    — IFRS New RFD = 'w'  (Warning actual; FIP sends 'e' → MisMatch)
      AP_BOTH_W      — IFRS New SFD = 'w'  (Both severity, FIP sends 'w')
      AP_BOTH_E      — IFRS New CFD = 'e'  (Both severity, FIP sends 'e')
      AP_GREY_WINS   — SST RFD = 'w'  but IFRS New RFD = '' (black col for V900W empty)
                       → black binding finds no actual → grey binding wins
      AP_NO_BINDING  — IFRS New RFD = 'w'  (FIP sends unknown method → skipped)
      AP_NO_ACTUAL   — IFRS New RFD = ''   (event col blank → no winning binding → skipped)
      AP_NOT_IN_SCOPE_TOC — IFRS New RFD = 'w', Type of change = '' → excluded
      AP_NOT_IN_SCOPE_INA — IFRS New RFD = 'w', Status = INACTIVE → excluded
      AP_EXCL_ZCORE  — IFRS New RFD = 'w', Exclude Z-Core = X → excluded
      AP_YELLOW_CAT  — IFRS New RFD = 'w', Category = yellow → excluded
    """
    _, all_binding_list, events = _get_event_methods()

    # Pick events: use first 4 from DEFAULT_EVENTS that exist in real file
    from strategies.accounting_principles.accounting_principles import DEFAULT_EVENTS
    available = [e for e in DEFAULT_EVENTS if e in [b.event for b in all_binding_list]]

    # We need: a Warning-only event, a Both event, an event where we can test grey
    warning_event = "IFRS New RFD"
    both_events = []
    grey_event = None

    from strategies.accounting_principles.validation_methods import parse_method_bindings
    all_bindings = all_binding_list
    for b in all_bindings:
        if b.severity == "Both" and len(both_events) < 2:
            if b.event not in both_events:
                both_events.append(b.event)
        if b.font == "grey" and grey_event is None:
            grey_event = b.event

    both_event_w = both_events[0] if both_events else "IFRS New SFD"
    both_event_e = both_events[1] if len(both_events) > 1 else "IFRS New CFD"
    if grey_event is None:
        grey_event = "SST RFD"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "cross checks all"

    # Build headers: standard filtering columns + event columns we need
    event_cols = [warning_event, both_event_w, both_event_e, grey_event]
    # Deduplicate preserving order
    seen = set()
    event_cols_dedup = []
    for e in event_cols:
        if e not in seen:
            event_cols_dedup.append(e)
            seen.add(e)

    fixed_headers = [
        "X-Check No.",      # A
        "Status",           # B
        "Type of change",   # C
        "Exclude Z-Core",   # D
        "Category",         # E
    ]
    headers = fixed_headers + event_cols_dedup
    ws.append(headers)

    def row(xc, status="ACTIVE", toc="Changed", excl_zcore="", category="",
            warn_val="", both_w_val="", both_e_val="", grey_val=""):
        vals = {
            warning_event:  warn_val,
            both_event_w:   both_w_val,
            both_event_e:   both_e_val,
            grey_event:     grey_val,
        }
        return [xc, status, toc, excl_zcore, category] + [vals.get(e, "") for e in event_cols_dedup]

    ws.append(row("AP_MATCH_W",    warn_val="w"))
    ws.append(row("AP_MISMATCH",   warn_val="w"))
    ws.append(row("AP_BOTH_W",     both_w_val="w"))
    ws.append(row("AP_BOTH_E",     both_e_val="e"))

    # AP_GREY_WINS: leave warning_event blank, set grey_event = 'w'
    # The grey binding fires for this row
    ws.append(row("AP_GREY_WINS",  warn_val="", grey_val="w"))

    ws.append(row("AP_NO_BINDING", warn_val="w"))  # FIP sends unknown method
    ws.append(row("AP_NO_ACTUAL",  warn_val=""))   # event col blank → skipped

    # Selection filter tests (process_only_differences=True)
    ws.append(row("AP_NOT_SCOPE_TOC", warn_val="w", toc=""))        # blank Type of change
    ws.append(row("AP_NOT_SCOPE_INA", warn_val="w", status="INACTIVE"))
    ws.append(row("AP_EXCL_ZCORE",    warn_val="w", excl_zcore="X"))

    # Yellow category — must be applied as fill after row is appended
    ws.append(row("AP_YELLOW_CAT",    warn_val="w"))
    yellow_row = ws.max_row
    category_col_idx = headers.index("Category") + 1
    ws.cell(row=yellow_row, column=category_col_idx).fill = PatternFill("solid", fgColor="FFFFFF00")

    # Process only differences: event column colour filter
    # AP_DIFF_YELLOW: warning_event cell yellow → in scope; FIP matches → Match
    ws.append(row("AP_DIFF_YELLOW", warn_val="w"))
    diff_yellow_row = ws.max_row
    warn_col_idx = headers.index(warning_event) + 1
    ws.cell(row=diff_yellow_row, column=warn_col_idx).fill = PatternFill("solid", fgColor="FFFFFF00")

    # AP_DIFF_GREEN: warning_event cell green → in scope; FIP matches → Match
    ws.append(row("AP_DIFF_GREEN",  warn_val="w"))
    diff_green_row = ws.max_row
    ws.cell(row=diff_green_row, column=warn_col_idx).fill = PatternFill("solid", fgColor="FF92D050")

    # AP_DIFF_WHITE: all event columns white → excluded when diff=True
    ws.append(row("AP_DIFF_WHITE",  warn_val="w"))

    wb.save(OUT / "ap_pub.xlsx")

    # Save event names for use in FIP generator
    return {
        "warning_event": warning_event,
        "both_event_w":  both_event_w,
        "both_event_e":  both_event_e,
        "grey_event":    grey_event,
    }


# ---------------------------------------------------------------------------
# 2. FIP ZQ9_VALMSG file
# ---------------------------------------------------------------------------

def _make_ap_fip(event_names):
    """
    Raw ZQ9_VALMSG format. Strategy builds Key = MK|ValidRule at load time.
    MK = validation method code, ValidRule = X-Check No., MT = W/E letter.

    We need:
    - A Warning method for warning_event (V900W for IFRS New RFD)
    - A Both method for both events
    - A grey method for grey_event
    - An unknown method (V_UNKNOWN) for AP_NO_BINDING
    """
    from strategies.accounting_principles.validation_methods import parse_method_bindings
    from strategies.accounting_principles.accounting_principles import DEFAULT_EVENTS
    all_bindings = parse_method_bindings(VM_PATH, DEFAULT_EVENTS)

    warning_event = event_names["warning_event"]
    both_event_w  = event_names["both_event_w"]
    both_event_e  = event_names["both_event_e"]
    grey_event    = event_names["grey_event"]

    # Find the black Warning method for warning_event
    warn_method = next((b.method for b in all_bindings
                        if b.event == warning_event and b.severity == "Warning" and b.font == "black"), "V900W")

    # Find a Both method for both_event_w and both_event_e
    both_method_w = next((b.method for b in all_bindings
                          if b.event == both_event_w and b.severity == "Both"), None)
    both_method_e = next((b.method for b in all_bindings
                          if b.event == both_event_e and b.severity == "Both"), None)

    # Find a grey method for grey_event
    grey_method = next((b.method for b in all_bindings
                        if b.event == grey_event and b.font == "grey"), None)
    # Also find a black method for grey_event (will be checked first but EBX col is blank)
    grey_black_method = next((b.method for b in all_bindings
                              if b.event == grey_event and b.font == "black"), None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FIP Methods Rules and Condition"

    raw_headers = ["MethC", "MK", "Medium Text", "ValidRule", "Long Text",
                   "UCFV20G-TRUE_BRANCH", "Message class", "Msg.", "MT", "Message Text"]
    ws.append(raw_headers)

    def fip_row(mk, xc, mt):
        return ["1", mk, "Test method", xc, f"{mk}|{xc}", "X", "CLS", "001", mt, "Test"]

    # AP_MATCH_W: Warning method, FIP=w, actual=w → Match
    ws.append(fip_row(warn_method, "AP_MATCH_W", "w"))

    # AP_MISMATCH: Warning method, FIP=e, actual=w → MisMatch
    ws.append(fip_row(warn_method, "AP_MISMATCH", "e"))

    # AP_BOTH_W: Both method for both_event_w, FIP=w → Match (Both accepts w)
    if both_method_w:
        ws.append(fip_row(both_method_w, "AP_BOTH_W", "w"))

    # AP_BOTH_E: Both method for both_event_e, FIP=e → Match (Both accepts e)
    if both_method_e:
        ws.append(fip_row(both_method_e, "AP_BOTH_E", "e"))

    # AP_GREY_WINS: send grey method's V-code; black binding will fail (EBX col blank)
    # → grey binding fires with grey_event actual='w'
    if grey_method:
        ws.append(fip_row(grey_method, "AP_GREY_WINS", "w"))
    elif grey_black_method:
        # Fallback: use the black method but leave warning_event blank in pub
        ws.append(fip_row(grey_black_method, "AP_GREY_WINS", "w"))

    # AP_NO_BINDING: unknown method → strategy skips this row
    ws.append(fip_row("V_UNKNOWN_METHOD", "AP_NO_BINDING", "w"))

    # AP_NO_ACTUAL: warning method, but EBX pub has blank event col → skipped
    ws.append(fip_row(warn_method, "AP_NO_ACTUAL", "w"))

    # Selection filter tests — FIP rows present, but X-Checks are out of scope
    ws.append(fip_row(warn_method, "AP_NOT_SCOPE_TOC", "w"))
    ws.append(fip_row(warn_method, "AP_NOT_SCOPE_INA", "w"))
    ws.append(fip_row(warn_method, "AP_EXCL_ZCORE", "w"))
    ws.append(fip_row(warn_method, "AP_YELLOW_CAT", "w"))

    # AP_DIFF_YELLOW/GREEN: event col yellow/green → in scope; FIP matches → Match
    ws.append(fip_row(warn_method, "AP_DIFF_YELLOW", "w"))
    ws.append(fip_row(warn_method, "AP_DIFF_GREEN",  "w"))
    # AP_DIFF_WHITE: excluded by diff filter; FIP row present but won't appear in Comparison

    wb.save(OUT / "ap_fip_ZQ9_VALMSG.xlsx")
    print("  wrote ap_fip_ZQ9_VALMSG.xlsx")

    return {
        "warn_method":       warn_method,
        "both_method_w":     both_method_w,
        "both_method_e":     both_method_e,
        "grey_method":       grey_method or grey_black_method,
    }


# ---------------------------------------------------------------------------
# 3. Known Exception List
# ---------------------------------------------------------------------------

def _make_ap_kel(event_names, method_names):
    """
    Build ap_kel.xlsx keyed to AP_MISMATCH (the MisMatch row).
    Runs the comparison to extract exact 6-column fingerprint.
    """
    import pandas as pd
    from strategies.accounting_principles.validation_methods import parse_method_bindings
    from strategies.accounting_principles.compare import compare_with_bindings
    from strategies.accounting_principles.accounting_principles import DEFAULT_EVENTS

    bindings = parse_method_bindings(VM_PATH, DEFAULT_EVENTS)
    cc_df = pd.read_excel(OUT / "ap_pub.xlsx", sheet_name="cross checks all")
    fip_df = pd.read_excel(OUT / "ap_fip_ZQ9_VALMSG.xlsx",
                           sheet_name="FIP Methods Rules and Condition")
    fip_df["Key"] = (fip_df["MK"].astype(str).str.strip()
                     + "|" + fip_df["ValidRule"].astype(str).str.strip())

    xchecks = [str(x).strip() for x in cc_df["X-Check No."].tolist()
               if str(x).strip() not in ("nan", "", "None")]
    rows = compare_with_bindings(bindings, cc_df, xchecks, fip_df)
    df = pd.DataFrame(rows)

    fp_cols = ["X-Check No.", "Event", "Expected", "FIP", "Actual", "Method"]
    metadata = ["Reason", "Added By", "Date Added", "Resolution Status", "Resolution Notes"]
    headers = fp_cols + metadata

    kel_row = df[df["X-Check No."] == "AP_MISMATCH"].iloc[0]
    no_match_row = df[df["X-Check No."] == "AP_MATCH_W"].iloc[0]  # correct row, wrong fingerprint

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name in ["X-Checks", "Grouping By", "Accounting Principles", "Conditions"]:
        ws = wb.create_sheet(sheet_name)
        if sheet_name == "Accounting Principles":
            ws.append(headers)
            ws.append(["Guidance: do not delete this row"] + [""] * (len(headers) - 1))
            # Correct entry for AP_MISMATCH
            data = [kel_row[c] for c in fp_cols]
            data += ["Test fixture — expected mismatch", "fixture_generator", "2026-07-30", "Open", ""]
            ws.append(data)
            # Wrong fingerprint for AP_KEL_NO_MATCH — use AP_MATCH_W values but wrong FIP letter
            wrong = [no_match_row[c] for c in fp_cols]
            wrong[fp_cols.index("FIP")] = "e"   # wrong FIP value
            wrong[0] = "AP_KEL_NO_MATCH"         # point to non-existent XC
            wrong += ["Test fixture — wrong fingerprint", "fixture_generator", "2026-07-30", "Open", ""]
            ws.append(wrong)
        else:
            ws.append(["(no entries)"])

    ws_inst = wb.create_sheet("Instructions")
    ws_inst.append(["Row 2 of each strategy sheet is a guidance row skipped by the app."])
    wb.save(OUT / "ap_kel.xlsx")
    print("  wrote ap_kel.xlsx")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print(f"Writing Accounting Principles fixtures to {OUT}/")
    events_info = _make_ap_pub({})
    print(f"  wrote ap_pub.xlsx  (events: {events_info})")
    method_info = _make_ap_fip(events_info)
    print(f"  method codes: {method_info}")
    _make_ap_kel(events_info, method_info)
    print("Done.")
