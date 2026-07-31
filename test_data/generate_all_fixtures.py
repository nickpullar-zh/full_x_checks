"""
Generate ALL fixture files into test_data/fixtures/ (flat layout, no sub-folders).

Run:  python test_data/generate_all_fixtures.py

File naming convention:
  Shared across strategies  → plain name         (e.g. xc_pub.xlsx, validation_methods.xlsx)
  X-Checks only             → xc_ prefix         (e.g. xc_fip.txt, xc_gcoa.xlsx, xc_kel.xlsx)
  Grouping By only          → gb_ prefix         (e.g. gb_fip_ZQ9_VALFLDGR.xlsx, gb_mapping.txt)
  Accounting Principles     → ap_ prefix         (e.g. ap_fip_ZQ9_VALMSG.xlsx, ap_kel.xlsx)
  Conditions                → cond_ prefix       (e.g. cond_fip_ZQ9_VALMETH.xlsx, cond_kel.xlsx)

validation_methods.xlsx is NOT generated — copy the real Zurich reference file manually.

Complete file list
==================
  xc_pub.xlsx                — shared EBX publication (all strategies, all rows merged)
  xc_fip.txt                 — FIP X-Checks text (SAP Validation Rule export)
  xc_gcoa.xlsx               — GCoA Publication File (one QU-type account row)
  xc_kel.xlsx                — Known Exception List for X-Checks
  gb_fip_ZQ9_VALFLDGR.xlsx   — FIP Grouping By
  gb_mapping.txt             — Grouping By field mapping
  gb_kel.xlsx                — Known Exception List for Grouping By
  ap_fip_ZQ9_VALMSG.xlsx     — FIP Accounting Principles (raw ZQ9_VALMSG)
  ap_kel.xlsx                — Known Exception List for Accounting Principles
  validation_methods.xlsx    — AP Validation Methods (copy real file manually)
  cond_fip_ZQ9_VALMETH.xlsx  — FIP Conditions (raw ZQ9_VALMETH)
  cond_kel.xlsx              — Known Exception List for Conditions
"""

import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill

OUT = Path(__file__).parent / "fixtures"
OUT.mkdir(exist_ok=True)
sys.path.insert(0, str(OUT.parent.parent))

yellow_fill = PatternFill("solid", fgColor="FFFFFF00")
green_fill  = PatternFill("solid", fgColor="FF92D050")
orange_fill = PatternFill("solid", fgColor="FFFF9900")


def _write_rows(ws, header, rows):
    ws.append(header)
    for r in rows:
        ws.append(r)


# ===========================================================================
# 1. Shared EBX Publication file  (xc_pub.xlsx)
# ===========================================================================

def _make_xc_pub():
    """
    Single merged publication file for ALL strategies.
    Columns span every strategy requirement.
    All fills applied at end for stable row indexing.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "cross checks all"

    # Column layout — indices are 1-based
    # A=1  X-Check No.
    # B=2  Account No.
    # C=3  SubA No.
    # D=4  Operator (X-Check Term)
    # E=5  Absolute (result)
    # F=6  Category
    # G=7  Operator 1
    # H=8  Operator 2
    # I=9  Limit 1
    # J=10 Limit 2
    # K=11 %
    # L=12 Exclude Account Type
    # M=13 Version Spanning Validation
    # N=14 Ending Balance Prior Year
    # O=15 Grouping By
    # P=16 Reference  X-Check (Condition)   (two spaces)
    # Q=17 IFRS New RFD                     (AP Warning event)
    # R=18 Applicable Quarters
    # S=19 Included RUs
    # T=20 Excluded RUs
    # U=21 Reference X-Check (Limit, %)
    # V=22 Status
    # W=23 Type of change
    # X=24 Exclude Z-Core
    # Y=25 Sense Check
    # Z=26 In Scope
    # AA=27 Stammhaus SLST RFD              (AP Both-severity event 1)
    # AB=28 Stammhaus SLST SFD              (AP Both-severity event 2)
    # AC=29 IFRS New SFD                    (AP grey binding event)

    headers = [
        "X-Check No.", "Account No.", "SubA No.", "Operator (X-Check Term)",
        "Absolute (result)", "Category", "Operator 1", "Operator 2",
        "Limit 1", "Limit 2", "%", "Exclude Account Type",
        "Version Spanning Validation", "Ending Balance Prior Year",
        "Grouping By", "Reference  X-Check (Condition)",
        "IFRS New RFD",
        "Applicable Quarters", "Included RUs", "Excluded RUs", "Reference X-Check (Limit, %)",
        "Status", "Type of change", "Exclude Z-Core", "Sense Check", "In Scope",
        "Stammhaus SLST RFD", "Stammhaus SLST SFD", "IFRS New SFD",
    ]
    ws.append(headers)

    pending_fills = []   # (row, col, fill)

    def row(xc, acct="", suba="", op_term="+", absolute="", category="",
            op1="<=", op2="", lim1="0", lim2="", pct="",
            excl_acc="", vsv="", ebpy="",
            grouping_by="", ref_xc_cond="",
            ifrs_new_rfd="",
            app_qtrs="", incl_rus="", excl_rus="", ref_xc_lim="",
            status="ACTIVE", toc="", excl_zcore="",
            slst_rfd="", slst_sfd="", ifrs_new_sfd=""):
        return [
            xc, acct, suba, op_term, absolute, category,
            op1, op2, lim1, lim2, pct,
            excl_acc, vsv, ebpy,
            grouping_by, ref_xc_cond,
            ifrs_new_rfd,
            app_qtrs, incl_rus, excl_rus, ref_xc_lim,
            status, toc, excl_zcore, "", "",
            slst_rfd, slst_sfd, ifrs_new_sfd,
        ]

    # ── X-Checks: standard match/mismatch ─────────────────────────────────
    ws.append(row("XC_ALL_MATCH",         acct="ACC001", op1="<=", lim1="0"))
    ws.append(row("XC_ALL_MISMATCH",      acct="ACC001", op1="<=", lim1="0"))
    ws.append(row("XC_FORMULA_MISMATCH",  acct="ACC001", op1="<=", lim1="0"))
    ws.append(row("XC_VARIABLE_MISMATCH", acct="ACC001", op1="<=", lim1="0"))
    ws.append(row("XC_NOT_IN_FIP",        acct="ACC001", op1="<=", lim1="0"))
    ws.append(row("XC_NOT_IN_EBX"))

    # ── X-Checks: FIP text normalisation ──────────────────────────────────
    ws.append(row("XC_TOM_CORRECTION",    acct="ACC001", suba="AA", op_term="+", op1="<=", lim1="0"))
    ws.append(row("XC_REX_CORRECTION",    acct="ACC001", suba="CC", op_term="+", op1="<=", lim1="0"))
    ws.append(row("XC_THOUSANDS_CORR",    acct="ACC001", op1="<=", lim1="1000"))

    # ── X-Checks: addition reorder ─────────────────────────────────────────
    ws.append(row("XC_REORDER_MATCH", acct="A_ACC", suba="AA", op_term="+", op1="<=", lim1="0"))
    ws.append(row("XC_REORDER_MATCH", acct="B_ACC", suba="BB", op_term="+", op1="<=", lim1="0"))

    # ── X-Checks: EBX formula paths ───────────────────────────────────────
    ws.append(row("XC_ABS_FORMULA",    acct="ACC001", op1="<=", op2="<=", lim1="0", lim2="5"))
    ws.append(row("XC_LC_YTD",         acct="ACC001", category="Shareholders' Equity", op1="<=", lim1="0"))
    ws.append(row("XC_LC_CONST",       acct="ACC001", category="Shareholders' Equity", op1="<=", lim1="100"))
    ws.append(row("XC_PCT_FORMAT",     acct="ACC001", op1="<",  lim1="1.5", pct="X"))
    ws.append(row("XC_FF_SUFFIX",      acct="ACC001", op_term="+", op1="<=", lim1="0"))
    ws.append(row("XC_FF_SUFFIX",      acct="ACC002", op_term="+", op1="<=", lim1="0"))
    ws.append(row("XC_SUBTRACT",       acct="ACC_POS", op_term="+", op1="<=", lim1="0"))
    ws.append(row("XC_SUBTRACT",       acct="ACC_NEG", op_term="-", op1="<=", lim1="0"))
    ws.append(row("XC_NONZERO_LIMIT",  acct="ACC001", op1="<=", lim1="100"))
    ws.append(row("XC_GTE_OPERATOR",   acct="ACC001", op1=">=", lim1="0"))
    ws.append(row("XC_EXCL_MATCH",     acct="ACC001", excl_acc="2", op1="<=", lim1="0"))
    ws.append(row("XC_EXCL_MISMATCH",  acct="ACC001", excl_acc="2", op1="<=", lim1="0"))

    # ── X-Checks: GCoA QU_YTD path ────────────────────────────────────────
    # XC_QU_YTD: Account=ACC_QU appears in xc_gcoa.xlsx with Data type=QU
    # → extract_ebx uses QU_YTD instead of VAL_YTD → Match
    ws.append(row("XC_QU_YTD", acct="ACC_QU", op1="<=", lim1="0"))

    # ── X-Checks: KEL ─────────────────────────────────────────────────────
    ws.append(row("XC_KEL_MISMATCH",   acct="ACC001", op1="<=", lim1="0"))
    ws.append(row("XC_KEL_NO_MATCH",   acct="ACC001", op1="<=", lim1="0"))

    # ── X-Checks: differences mode ────────────────────────────────────────
    ws.append(row("XC_DIFF_IN_SCOPE",    acct="ACC001", op1="<=", lim1="0", toc="Changed"))
    pending_fills.append((ws.max_row, 23, yellow_fill))

    ws.append(row("XC_DIFF_EXCLUDED",    acct="ACC001", op1="<=", lim1="0", toc="Changed", excl_zcore="X"))
    pending_fills.append((ws.max_row, 23, yellow_fill))

    ws.append(row("XC_DIFF_INACTIVE",    acct="ACC001", op1="<=", lim1="0", status="INACTIVE", toc="Changed"))
    pending_fills.append((ws.max_row, 23, yellow_fill))

    ws.append(row("XC_DIFF_YELLOW_CAT",  acct="ACC001", op1="<=", lim1="0", category="Test Cat", toc="Changed"))
    pending_fills.append((ws.max_row, 6,  yellow_fill))
    pending_fills.append((ws.max_row, 23, yellow_fill))

    ws.append(row("XC_DIFF_YELLOW",      acct="ACC001", op1="<=", lim1="0", toc="Changed"))
    pending_fills.append((ws.max_row, 23, yellow_fill))

    ws.append(row("XC_DIFF_GREEN",       acct="ACC001", op1="<=", lim1="0", toc="New x-check or association"))
    pending_fills.append((ws.max_row, 23, green_fill))

    ws.append(row("XC_DIFF_ORANGE",      acct="ACC001", op1="<=", lim1="0", toc="Removed"))
    pending_fills.append((ws.max_row, 23, orange_fill))

    ws.append(row("XC_DIFF_NO_TOC",      acct="ACC001", op1="<=", lim1="0"))

    # ── Grouping By rows ──────────────────────────────────────────────────
    ws.append(row("GB_MATCHED",       grouping_by="ITEM_A"))
    ws.append(row("GB_NOT_IN_FIP",    grouping_by="ITEM_A"))
    ws.append(row("GB_REF_XC_KEY",    grouping_by="ITEM_A", ref_xc_cond="REF_BASE"))
    ws.append(row("GB_MULTI",         grouping_by="ITEM_A, ITEM_B"))
    ws.append(row("GB_DEDUP",         grouping_by="ITEM_A"))
    ws.append(row("GB_DEDUP",         grouping_by="ITEM_B"))
    ws.append(row("GB_IGNORE_FIELD",  grouping_by="ITEM_A"))
    ws.append(row("GB_UNMAPPED",      grouping_by="ITEM_A"))
    ws.append(row("GB_BLANK_VR",      grouping_by="ITEM_A"))
    ws.append(row("GB_KEL_MATCH",     grouping_by="ITEM_A"))
    ws.append(row("GB_KEL_NO_MATCH",  grouping_by="ITEM_A"))

    ws.append(row("GB_DIFF_YELLOW",   grouping_by="ITEM_A"))
    pending_fills.append((ws.max_row, 15, yellow_fill))  # col 15 = Grouping By

    ws.append(row("GB_DIFF_GREEN",    grouping_by="ITEM_A"))
    pending_fills.append((ws.max_row, 15, green_fill))

    ws.append(row("GB_DIFF_WHITE",    grouping_by="ITEM_A"))

    # ── Accounting Principles rows ────────────────────────────────────────
    ws.append(row("AP_MATCH_W",    ifrs_new_rfd="w"))
    ws.append(row("AP_MISMATCH",   ifrs_new_rfd="w"))
    ws.append(row("AP_BOTH_W",     slst_rfd="w"))
    ws.append(row("AP_BOTH_E",     slst_sfd="e"))
    ws.append(row("AP_GREY_WINS",  ifrs_new_rfd="", ifrs_new_sfd="w"))
    ws.append(row("AP_NO_BINDING", ifrs_new_rfd="w"))
    ws.append(row("AP_NO_ACTUAL",  ifrs_new_rfd=""))
    ws.append(row("AP_NOT_SCOPE_TOC", ifrs_new_rfd="w", toc=""))
    ws.append(row("AP_NOT_SCOPE_INA", ifrs_new_rfd="w", status="INACTIVE"))
    ws.append(row("AP_EXCL_ZCORE",    ifrs_new_rfd="w", excl_zcore="X"))
    ws.append(row("AP_YELLOW_CAT",    ifrs_new_rfd="w"))
    pending_fills.append((ws.max_row, 6, yellow_fill))   # Category yellow

    # AP diff rows also need Type of change so they pass the text-based pre-filter
    ws.append(row("AP_DIFF_YELLOW", ifrs_new_rfd="w", toc="Changed"))
    pending_fills.append((ws.max_row, 17, yellow_fill))  # IFRS New RFD cell yellow
    pending_fills.append((ws.max_row, 23, yellow_fill))  # Type of change cell yellow

    ws.append(row("AP_DIFF_GREEN",  ifrs_new_rfd="w", toc="New x-check or association"))
    pending_fills.append((ws.max_row, 17, green_fill))
    pending_fills.append((ws.max_row, 23, green_fill))   # Type of change cell green

    ws.append(row("AP_DIFF_WHITE",  ifrs_new_rfd="w"))

    # ── Conditions rows ───────────────────────────────────────────────────
    ws.append(row("COND_APPL_QTRS",     app_qtrs="Q1"))
    ws.append(row("COND_INCL_RUS",      incl_rus="RU_NORTH"))
    ws.append(row("COND_EXCL_RUS",      excl_rus="RU_SOUTH"))
    ws.append(row("COND_LIMIT_PCT",     ref_xc_lim="10.5"))
    ws.append(row("COND_NOT_MATCHED",   app_qtrs="Q2"))
    ws.append(row("COND_MULTI_COL",     app_qtrs="Q1", incl_rus="RU_IN", excl_rus="RU_OUT"))
    ws.append(row("COND_KEL_MISMATCH",  app_qtrs="Q3"))
    ws.append(row("COND_KEL_NO_MATCH",  app_qtrs="Q4"))
    ws.append(row("COND_REF_XC",        ref_xc_cond="COND_REF_XC", app_qtrs="Q1"))

    ws.append(row("COND_DIFF_YELLOW",   app_qtrs="Q1"))
    pending_fills.append((ws.max_row, 18, yellow_fill))  # Applicable Quarters yellow

    ws.append(row("COND_DIFF_GREEN",    incl_rus="RU_NORTH"))
    pending_fills.append((ws.max_row, 19, green_fill))   # Included RUs green

    ws.append(row("COND_DIFF_WHITE",    app_qtrs="Q1"))

    # Apply all fills
    for r, c, fill in pending_fills:
        ws.cell(row=r, column=c).fill = fill

    wb.save(OUT / "xc_pub.xlsx")
    print("  wrote xc_pub.xlsx")


# ===========================================================================
# 2. FIP X-Checks text  (xc_fip.txt)
# ===========================================================================

_SEGMENT_END = "|-Segment @28@ * |"
_BLOCK_END   = "-|"
_BLANK       = "|"
_FORMULA_HDR = "|Formula String |"
_VAR_HDR     = "|-Characteristic Sel Opt Attributes Node Characteristic From To |"


def _fip_block_single(xc_id, formula, var_name, fs_accounts,
                      movement_types=None, excl_types=None):
    if movement_types is None: movement_types = []
    if excl_types     is None: excl_types     = []
    fs_lines   = "".join(f"| 1 | 2 | 3 | FS Account | {a} |\n" for a in fs_accounts)
    excl_lines = "".join(f"| 0 | 1 | @2A@ | Account | Type | {t} |\n" for t in excl_types)
    if movement_types:
        mt_lines = f"| |- Movement Type @20@ {movement_types[0]} desc |\n"
        for mt in movement_types[1:]:
            mt_lines += f"| 1 | 2 | 3 | @20@ | {mt} | desc |\n"
    else:
        mt_lines = ""
    return (
        f"{xc_id} {xc_id} {xc_id} test description\n"
        f"{_FORMULA_HDR}\n{formula}\n{_BLOCK_END}\n"
        f"{_VAR_HDR}\n{var_name}\n{_BLANK}\n"
        f"{fs_lines}{excl_lines}{mt_lines}"
        f"{_BLOCK_END}\n{_SEGMENT_END}\n{_BLOCK_END}\n\n"
    )


def _fip_block_two_vars(xc_id, formula, var1_name, var1_fs, var1_mt, var2_name, var2_fs, var2_mt):
    def fs(a): return "".join(f"| 1 | 2 | 3 | FS Account | {x} |\n" for x in a)
    def mt(m): return f"| |- Movement Type @20@ {m} desc |\n"
    return (
        f"{xc_id} {xc_id} {xc_id} test description\n"
        f"{_FORMULA_HDR}\n{formula}\n{_BLOCK_END}\n{_VAR_HDR}\n"
        f"{var1_name}\n{_BLANK}\n{fs(var1_fs)}{mt(var1_mt)}{_BLANK}\n"
        f"{var2_name}\n{_BLANK}\n{fs(var2_fs)}{mt(var2_mt)}{_BLOCK_END}\n"
        f"{_SEGMENT_END}\n{_BLOCK_END}\n\n"
    )


def _fip_block_two_vars_no_mt(xc_id, formula, var1_name, var1_fs, var2_name, var2_fs):
    def fs(a): return "".join(f"| 1 | 2 | 3 | FS Account | {x} |\n" for x in a)
    return (
        f"{xc_id} {xc_id} {xc_id} test description\n"
        f"{_FORMULA_HDR}\n{formula}\n{_BLOCK_END}\n{_VAR_HDR}\n"
        f"{var1_name}\n{_BLANK}\n{fs(var1_fs)}{_BLANK}\n"
        f"{var2_name}\n{_BLANK}\n{fs(var2_fs)}{_BLOCK_END}\n"
        f"{_SEGMENT_END}\n{_BLOCK_END}\n\n"
    )


def _make_xc_fip():
    blocks = [
        _fip_block_single("XC_ALL_MATCH",        "VAL_YTD(ACC001)<=0", "ACC001", ["ACC001"]),
        _fip_block_single("XC_ALL_MISMATCH",      "VAL_YTD(ACC999)<=0", "ACC999", ["ACC999"]),
        _fip_block_single("XC_FORMULA_MISMATCH",  "VAL_YTD(ACC001)>=0", "ACC001", ["ACC001"]),
        _fip_block_single("XC_VARIABLE_MISMATCH", "VAL_YTD(ACC001)<=0", "ACC_WRONG", ["ACC_WRONG"]),
        _fip_block_single("XC_NOT_IN_EBX",        "VAL_YTD(ACC001)<=0", "ACC001", ["ACC001"]),
        _fip_block_single("XC_TOM_CORRECTION",    "VAL_YTD(ACC001TOMAA)<=0", "ACC001TOMAA", ["ACC001"], movement_types=["AA"]),
        _fip_block_single("XC_REX_CORRECTION",    "VAL_YTD(ACC001REXCC)<=0", "ACC001REXCC", ["ACC001"], movement_types=["CC"]),
        _fip_block_single("XC_THOUSANDS_CORR",    "VAL_YTD(ACC001)<=CONST(1.000,'USD','E')", "ACC001", ["ACC001"]),
        _fip_block_two_vars("XC_REORDER_MATCH",   "VAL_YTD(B_ACCToMBB)+VAL_YTD(A_ACCToMAA)<=0",
                            "A_ACCToMAA", ["A_ACC"], "AA", "B_ACCToMBB", ["B_ACC"], "BB"),
        _fip_block_single("XC_ABS_FORMULA",       "ABS(VAL_YTD(ACC001))<=CONST(5,'USD','E')", "ACC001", ["ACC001"]),
        _fip_block_single("XC_LC_YTD",            "LC_YTD(ACC001)<=0", "ACC001", ["ACC001"]),
        _fip_block_single("XC_LC_CONST",          "LC_YTD(ACC001)<=CONST_LC(100,'USD','E')", "ACC001", ["ACC001"]),
        _fip_block_single("XC_PCT_FORMAT",        "VAL_YTD(ACC001)<'1,500000%'", "ACC001", ["ACC001"]),
        _fip_block_single("XC_FF_SUFFIX",         "VAL_YTD(ACC001ff)<=0", "ACC001ff", ["ACC001", "ACC002"]),
        _fip_block_two_vars_no_mt("XC_SUBTRACT",  "VAL_YTD(ACC_POS)-VAL_YTD(ACC_NEG)<=0",
                                  "ACC_POS", ["ACC_POS"], "ACC_NEG", ["ACC_NEG"]),
        _fip_block_single("XC_NONZERO_LIMIT",     "VAL_YTD(ACC001)<=CONST(100,'USD','E')", "ACC001", ["ACC001"]),
        _fip_block_single("XC_GTE_OPERATOR",      "VAL_YTD(ACC001)>=0", "ACC001", ["ACC001"]),
        _fip_block_single("XC_EXCL_MATCH",        "VAL_YTD(ACC001)<=0", "ACC001", ["ACC001"], excl_types=["2"]),
        _fip_block_single("XC_EXCL_MISMATCH",     "VAL_YTD(ACC001)<=0", "ACC001", ["ACC001"]),
        # XC_QU_YTD: without GCoA EBX produces VAL_YTD; FIP has VAL_YTD → Match
        # The GCoA test (FX-06e/f) verifies that WITH GCoA it switches to QU_YTD
        _fip_block_single("XC_QU_YTD",            "VAL_YTD(ACC_QU)<=0", "ACC_QU", ["ACC_QU"]),
        _fip_block_single("XC_KEL_MISMATCH",      "VAL_YTD(ACC999)<=0", "ACC999", ["ACC999"]),
        _fip_block_single("XC_KEL_NO_MATCH",      "VAL_YTD(ACC888)<=0", "ACC888", ["ACC888"]),
        _fip_block_single("XC_DIFF_IN_SCOPE",     "VAL_YTD(ACC001)<=0", "ACC001", ["ACC001"]),
        _fip_block_single("XC_DIFF_YELLOW",       "VAL_YTD(ACC001)<=0", "ACC001", ["ACC001"]),
        _fip_block_single("XC_DIFF_GREEN",        "VAL_YTD(ACC001)<=0", "ACC001", ["ACC001"]),
        # XC_NOT_IN_FIP, XC_DIFF_ORANGE, XC_DIFF_NO_TOC, XC_DIFF_EXCLUDED etc: intentionally absent
    ]
    (OUT / "xc_fip.txt").write_text("".join(blocks), encoding="utf-8")
    print("  wrote xc_fip.txt")


# ===========================================================================
# 3. GCoA Publication File  (xc_gcoa.xlsx)
# ===========================================================================

def _make_xc_gcoa():
    """
    GCoA file with one QU-type account (ACC_QU).
    When this file is loaded, extract_ebx uses QU_YTD instead of VAL_YTD
    for any X-Check that includes ACC_QU.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GCoA Base account table"
    _write_rows(ws,
        ["Account ID", "Data type", "Description"],
        [
            ["ACC_QU", "QU", "Test QU-type account — triggers QU_YTD formula"],
            ["ACC001", "FS", "Standard FS account"],
        ],
    )
    wb.save(OUT / "xc_gcoa.xlsx")
    print("  wrote xc_gcoa.xlsx")


# ===========================================================================
# 4. X-Checks Known Exception List  (xc_kel.xlsx)
# ===========================================================================

def _make_known_exception_list():
    """
    Single Known Exception List workbook with one sheet per strategy.
    This matches how the app (and Known Exception Builder) expects it.
    """
    import pandas as pd
    from strategies.x_checks.ebx_extraction import extract_ebx
    from strategies.x_checks.fip_extraction import extract_fip
    from strategies.x_checks.compare import compare as xc_compare
    from strategies.grouping_by.grouping_by import GroupingBy
    from task_configs import GROUPING_BY_UPLOAD_CONFIG
    from strategies.accounting_principles.validation_methods import parse_method_bindings
    from strategies.accounting_principles.compare import compare_with_bindings
    from strategies.accounting_principles.accounting_principles import DEFAULT_EVENTS
    from strategies.conditions.extract import extract_conditions
    from strategies.conditions.fip import process_fip
    from strategies.conditions.compare import compare as cond_compare

    meta = ["Reason", "Added By", "Date Added", "Resolution Status", "Resolution Notes"]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── X-Checks sheet ────────────────────────────────────────────────────
    ebx_df = pd.read_excel(OUT / "xc_pub.xlsx", sheet_name="cross checks all")
    ebx_results = extract_ebx(ebx_df)
    xc_list = sorted(set(str(x) for x in ebx_df["X-Check No."].tolist()
                         if str(x) not in ("nan", "", "None")))
    fip_text = (OUT / "xc_fip.txt").read_text(encoding="utf-8")
    fip_results = extract_fip(fip_text, xc_list)
    xc_df = pd.DataFrame(xc_compare(ebx_results, fip_results))

    xc_fp   = ["X-Check No.", "EBX Formula", "FIP Formula", "EBX Formula (Excl)",
               "FIP Formula (Excl)", "EBX Variables", "FIP Variables", "FIP Variable (Builder)"]
    xc_hdrs = xc_fp + meta
    ws = wb.create_sheet("X-Checks")
    ws.append(xc_hdrs)
    ws.append(["Guidance: do not delete this row"] + [""] * (len(xc_hdrs) - 1))
    kel_row = xc_df[xc_df["X-Check No."] == "XC_KEL_MISMATCH"].iloc[0]
    correct = [kel_row[c] for c in xc_fp]
    correct += ["Test fixture — expected mismatch", "fixture_generator", "2026-07-31", "Open", ""]
    ws.append(correct)
    no_match_row = xc_df[xc_df["X-Check No."] == "XC_KEL_NO_MATCH"].iloc[0]
    wrong = [no_match_row[c] for c in xc_fp]
    wrong[xc_fp.index("FIP Formula")] = "VAL_YTD(WRONG_ACCOUNT)<=0"
    wrong[xc_fp.index("FIP Formula (Excl)")] = "VAL_YTD(WRONG_ACCOUNT)<=0"
    wrong += ["Test fixture — wrong fingerprint (should not annotate)",
              "fixture_generator", "2026-07-31", "Open", ""]
    ws.append(wrong)

    # ── Grouping By sheet ─────────────────────────────────────────────────
    gb = GroupingBy(GROUPING_BY_UPLOAD_CONFIG)
    gb.log = []
    mapping_txt = (OUT / "gb_mapping.txt").read_text()
    fip_gb = pd.read_excel(OUT / "gb_fip_ZQ9_VALFLDGR.xlsx", sheet_name="Sheet1")
    loaded = {
        GROUPING_BY_UPLOAD_CONFIG.file_fields[0].label: ebx_df.copy(),
        GROUPING_BY_UPLOAD_CONFIG.file_fields[1].label: fip_gb,
        GROUPING_BY_UPLOAD_CONFIG.file_fields[2].label: mapping_txt,
    }
    _, _, df_fip = gb._process_fip(loaded)
    _, df_ebx   = gb._process_ebx(loaded)
    df_cmp      = gb._process_compare(df_fip, df_ebx)
    gb_fp   = ["EBX Key"]
    gb_hdrs = gb_fp + meta
    ws = wb.create_sheet("Grouping By")
    ws.append(gb_hdrs)
    ws.append(["Guidance: do not delete this row"] + [""] * (len(gb_hdrs) - 1))
    kel_gb = df_cmp[df_cmp["EBX Key"] == "GB_KEL_MATCH|ITEM_A"].iloc[0]
    ws.append([kel_gb["EBX Key"],
               "Test fixture — expected Not in FIP", "fixture_generator", "2026-07-31", "Open", ""])
    ws.append(["GB_KEL_NO_MATCH|WRONG_KEY",
               "Test fixture — wrong fingerprint (should not annotate)",
               "fixture_generator", "2026-07-31", "Open", ""])

    # ── Accounting Principles sheet ───────────────────────────────────────
    vm_path  = str(OUT / "validation_methods.xlsx")
    bindings = parse_method_bindings(vm_path, DEFAULT_EVENTS)
    fip_ap   = pd.read_excel(OUT / "ap_fip_ZQ9_VALMSG.xlsx",
                              sheet_name="FIP Methods Rules and Condition")
    fip_ap["Key"] = fip_ap["MK"].astype(str).str.strip() + "|" + fip_ap["ValidRule"].astype(str).str.strip()
    xchecks  = [str(x).strip() for x in ebx_df["X-Check No."].tolist()
                if str(x).strip() not in ("nan", "", "None")]
    ap_df    = pd.DataFrame(compare_with_bindings(bindings, ebx_df, xchecks, fip_ap))
    ap_fp   = ["X-Check No.", "Event", "Expected", "FIP", "Actual", "Method"]
    ap_hdrs = ap_fp + meta
    ws = wb.create_sheet("Accounting Principles")
    ws.append(ap_hdrs)
    ws.append(["Guidance: do not delete this row"] + [""] * (len(ap_hdrs) - 1))
    kel_ap = ap_df[ap_df["X-Check No."] == "AP_MISMATCH"].iloc[0]
    ap_data = [kel_ap[c] for c in ap_fp]
    ap_data += ["Test fixture — expected mismatch", "fixture_generator", "2026-07-31", "Open", ""]
    ws.append(ap_data)

    # ── Conditions sheet ──────────────────────────────────────────────────
    fip_cond = pd.read_excel(OUT / "cond_fip_ZQ9_VALMETH.xlsx", sheet_name="FIP Conditions")
    fip_proc = process_fip(fip_cond)
    working_df, _ = extract_conditions(str(OUT / "xc_pub.xlsx"), "cross checks all",
                                        process_only_differences=False)
    results_df, _ = cond_compare(working_df, fip_proc)
    cond_fp   = ["EBX Data", "FIP Data"]
    cond_hdrs = cond_fp + meta
    ws = wb.create_sheet("Conditions")
    ws.append(cond_hdrs)
    ws.append(["Guidance: do not delete this row"] + [""] * (len(cond_hdrs) - 1))
    kel_cond  = results_df[results_df["EBX Data"] == "COND_APPL_QTRS|Q1"].iloc[0]
    no_match_cond = results_df[results_df["EBX Data"] == "COND_INCL_RUS|RU_NORTH"].iloc[0]
    ws.append([kel_cond["EBX Data"], kel_cond["FIP Data"],
               "Test fixture — expected condition", "fixture_generator", "2026-07-31", "Open", ""])
    ws.append([no_match_cond["EBX Data"], "WRONG_FIP_DATA",
               "Test fixture — wrong fingerprint", "fixture_generator", "2026-07-31", "Open", ""])

    # ── Instructions sheet ────────────────────────────────────────────────
    ws_inst = wb.create_sheet("Instructions")
    ws_inst.append(["One workbook, one sheet per strategy. Row 2 of each strategy sheet is a guidance row skipped by the app."])

    wb.save(OUT / "known_exception_list.xlsx")
    print("  wrote known_exception_list.xlsx")


# ===========================================================================
# 5. Grouping By FIP  (gb_fip_ZQ9_VALFLDGR.xlsx)
# ===========================================================================

def _make_gb_fip():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    _write_rows(ws,
        ["ValidRule", "Long Text", "Field name"],
        [
            ["GB_MATCHED",    "Standard match",      "GB_FIP_FIELD"],
            ["REF_BASE",      "Ref XC base key",     "GB_FIP_FIELD"],
            ["GB_MULTI",      "Multi grouping A",    "GB_FIP_FIELD"],
            ["GB_DIFF_YELLOW","Diff yellow match",   "GB_FIP_FIELD"],
            ["GB_DIFF_GREEN", "Diff green match",    "GB_FIP_FIELD"],
            ["GB_IGNORE_FIELD","Ignore field test",  "GB_IGNORE_FIELD_FIP"],
            ["GB_UNMAPPED",   "Unmapped field test", "UNMAPPED_FIELD"],
            ["",              "Blank ValidRule test","GB_FIP_FIELD"],
        ],
    )
    wb.save(OUT / "gb_fip_ZQ9_VALFLDGR.xlsx")
    print("  wrote gb_fip_ZQ9_VALFLDGR.xlsx")


# ===========================================================================
# 6. Grouping By Mapping  (gb_mapping.txt)
# ===========================================================================

def _make_gb_mapping():
    (OUT / "gb_mapping.txt").write_text(
        "FIP Data,EBX item\nGB_FIP_FIELD,ITEM_A\nGB_IGNORE_FIELD_FIP,ignore\n",
        encoding="utf-8",
    )
    print("  wrote gb_mapping.txt")


# ===========================================================================
# (gb_kel merged into known_exception_list.xlsx — see _make_known_exception_list)
# ===========================================================================

def _make_gb_kel():
    import pandas as pd
    from strategies.grouping_by.grouping_by import GroupingBy
    from task_configs import GROUPING_BY_UPLOAD_CONFIG

    gb = GroupingBy(GROUPING_BY_UPLOAD_CONFIG)
    gb.log = []
    mapping_txt = (OUT / "gb_mapping.txt").read_text()
    fip_gb = pd.read_excel(OUT / "gb_fip_ZQ9_VALFLDGR.xlsx", sheet_name="Sheet1")
    ebx_gb = pd.read_excel(OUT / "xc_pub.xlsx", sheet_name="cross checks all")
    loaded = {
        GROUPING_BY_UPLOAD_CONFIG.file_fields[0].label: ebx_gb,
        GROUPING_BY_UPLOAD_CONFIG.file_fields[1].label: fip_gb,
        GROUPING_BY_UPLOAD_CONFIG.file_fields[2].label: mapping_txt,
    }
    _, _, df_fip = gb._process_fip(loaded)
    _, df_ebx    = gb._process_ebx(loaded)
    df_cmp       = gb._process_compare(df_fip, df_ebx)

    kel_row      = df_cmp[df_cmp["EBX Key"] == "GB_KEL_MATCH|ITEM_A"].iloc[0]
    fp_cols      = ["EBX Key"]
    meta         = ["Reason", "Added By", "Date Added", "Resolution Status", "Resolution Notes"]
    headers      = fp_cols + meta

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name in ["X-Checks", "Grouping By", "Accounting Principles", "Conditions"]:
        ws = wb.create_sheet(sheet_name)
        if sheet_name == "Grouping By":
            ws.append(headers)
            ws.append(["Guidance: do not delete this row"] + [""] * (len(headers) - 1))
            ws.append([kel_row["EBX Key"],
                       "Test fixture — expected Not in FIP", "fixture_generator", "2026-07-31", "Open", ""])
            ws.append(["GB_KEL_NO_MATCH|WRONG_KEY",
                       "Test fixture — wrong fingerprint (should not annotate)",
                       "fixture_generator", "2026-07-31", "Open", ""])
        else:
            ws.append(["(no entries)"])
    ws_inst = wb.create_sheet("Instructions")
    ws_inst.append(["Row 2 of each strategy sheet is a guidance row skipped by the app."])
    wb.save(OUT / "gb_kel.xlsx")
    print("  wrote gb_kel.xlsx")


# ===========================================================================
# 8. AP FIP  (ap_fip_ZQ9_VALMSG.xlsx)
# ===========================================================================

def _make_ap_fip():
    """Discover real method codes from validation_methods.xlsx."""
    vm_path = str(OUT / "validation_methods.xlsx")
    from strategies.accounting_principles.validation_methods import (
        parse_method_bindings, list_all_event_names
    )
    events   = list_all_event_names(vm_path)
    bindings = parse_method_bindings(vm_path, events)

    def find(event, severity=None, font=None):
        for b in bindings:
            if b.event == event:
                if severity and b.severity != severity: continue
                if font     and b.font     != font:     continue
                return b.method
        return None

    warn_method  = find("IFRS New RFD", severity="Warning", font="black") or "V900W"
    both_w       = find("Stammhaus SLST RFD", severity="Both")            or "V600A"
    both_e       = find("Stammhaus SLST SFD", severity="Both")            or "V600S"
    grey_method  = find("IFRS New SFD", font="grey")                      or "V900W"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FIP Methods Rules and Condition"
    raw_hdrs = ["MethC", "MK", "Medium Text", "ValidRule", "Long Text",
                "UCFV20G-TRUE_BRANCH", "Message class", "Msg.", "MT", "Message Text"]
    ws.append(raw_hdrs)

    def fip_row(mk, xc, mt):
        return ["1", mk, "Test method", xc, f"{mk}|{xc}", "X", "CLS", "001", mt, "Test"]

    for mk, xc, mt in [
        (warn_method,  "AP_MATCH_W",       "w"),
        (warn_method,  "AP_MISMATCH",       "e"),
        (both_w,       "AP_BOTH_W",         "w"),
        (both_e,       "AP_BOTH_E",         "e"),
        (grey_method,  "AP_GREY_WINS",      "w"),
        ("V_UNKNOWN",  "AP_NO_BINDING",     "w"),
        (warn_method,  "AP_NO_ACTUAL",      "w"),
        (warn_method,  "AP_NOT_SCOPE_TOC",  "w"),
        (warn_method,  "AP_NOT_SCOPE_INA",  "w"),
        (warn_method,  "AP_EXCL_ZCORE",     "w"),
        (warn_method,  "AP_YELLOW_CAT",     "w"),
        (warn_method,  "AP_DIFF_YELLOW",    "w"),
        (warn_method,  "AP_DIFF_GREEN",     "w"),
    ]:
        ws.append(fip_row(mk, xc, mt))

    wb.save(OUT / "ap_fip_ZQ9_VALMSG.xlsx")
    print("  wrote ap_fip_ZQ9_VALMSG.xlsx")


# ===========================================================================
# 9. AP KEL  (ap_kel.xlsx)
# ===========================================================================

def _make_ap_kel():
    import pandas as pd
    from strategies.accounting_principles.validation_methods import parse_method_bindings
    from strategies.accounting_principles.compare import compare_with_bindings
    from strategies.accounting_principles.accounting_principles import DEFAULT_EVENTS

    vm_path  = str(OUT / "validation_methods.xlsx")
    bindings = parse_method_bindings(vm_path, DEFAULT_EVENTS)
    cc_df    = pd.read_excel(OUT / "xc_pub.xlsx", sheet_name="cross checks all")
    fip_df   = pd.read_excel(OUT / "ap_fip_ZQ9_VALMSG.xlsx", sheet_name="FIP Methods Rules and Condition")
    fip_df["Key"] = fip_df["MK"].astype(str).str.strip() + "|" + fip_df["ValidRule"].astype(str).str.strip()
    xchecks  = [str(x).strip() for x in cc_df["X-Check No."].tolist()
                if str(x).strip() not in ("nan", "", "None")]
    df       = pd.DataFrame(compare_with_bindings(bindings, cc_df, xchecks, fip_df))

    kel_row  = df[df["X-Check No."] == "AP_MISMATCH"].iloc[0]
    fp_cols  = ["X-Check No.", "Event", "Expected", "FIP", "Actual", "Method"]
    meta     = ["Reason", "Added By", "Date Added", "Resolution Status", "Resolution Notes"]
    headers  = fp_cols + meta

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name in ["X-Checks", "Grouping By", "Accounting Principles", "Conditions"]:
        ws = wb.create_sheet(sheet_name)
        if sheet_name == "Accounting Principles":
            ws.append(headers)
            ws.append(["Guidance: do not delete this row"] + [""] * (len(headers) - 1))
            data = [kel_row[c] for c in fp_cols]
            data += ["Test fixture — expected mismatch", "fixture_generator", "2026-07-31", "Open", ""]
            ws.append(data)
        else:
            ws.append(["(no entries)"])
    ws_inst = wb.create_sheet("Instructions")
    ws_inst.append(["Row 2 of each strategy sheet is a guidance row skipped by the app."])
    wb.save(OUT / "ap_kel.xlsx")
    print("  wrote ap_kel.xlsx")


# ===========================================================================
# 10. Conditions FIP  (cond_fip_ZQ9_VALMETH.xlsx)
# ===========================================================================

def _make_cond_fip():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FIP Conditions"
    _write_rows(ws,
        ["MethC", "MK", "Medium Text", "ValidRule",
         "Medium Text", "UCFV20G-TRUE_BRANCH", "ValidRule", "Medium Text"],
        [
            ["1", "MK1", "Test", "COND_APPL_QTRS",     "Text", "X", "Q1",       "Q"],
            ["1", "MK1", "Test", "COND_INCL_RUS",       "Text", "X", "RU_NORTH", "R"],
            ["1", "MK1", "Test", "COND_EXCL_RUS",       "Text", "X", "RU_SOUTH", "R"],
            ["1", "MK1", "Test", "COND_LIMIT_PCT",      "Text", "X", "10.5",     "L"],
            ["1", "MK1", "Test", "COND_MULTI_COL",      "Text", "X", "Q1",       "Q"],
            ["1", "MK1", "Test", "COND_MULTI_COL",      "Text", "X", "RU_IN",    "R"],
            ["1", "MK1", "Test", "COND_REF_XC",         "Text", "X", "Q1",       "Q"],
            ["1", "MK1", "Test", "COND_DIFF_YELLOW",    "Text", "X", "Q1",       "Q"],
            ["1", "MK1", "Test", "COND_DIFF_GREEN",     "Text", "X", "RU_NORTH", "R"],
            # COND_NOT_MATCHED|Q2, COND_MULTI_COL|RU_OUT, COND_KEL_*, COND_DIFF_WHITE absent
        ],
    )
    wb.save(OUT / "cond_fip_ZQ9_VALMETH.xlsx")
    print("  wrote cond_fip_ZQ9_VALMETH.xlsx")


# ===========================================================================
# 11. Conditions KEL  (cond_kel.xlsx)
# ===========================================================================

def _make_cond_kel():
    import pandas as pd
    from strategies.conditions.extract import extract_conditions
    from strategies.conditions.fip import process_fip
    from strategies.conditions.compare import compare as cond_compare

    fip_df   = pd.read_excel(OUT / "cond_fip_ZQ9_VALMETH.xlsx", sheet_name="FIP Conditions")
    fip_proc = process_fip(fip_df)
    working_df, _ = extract_conditions(str(OUT / "xc_pub.xlsx"), "cross checks all",
                                        process_only_differences=False)
    results_df, _ = cond_compare(working_df, fip_proc)

    kel_row  = results_df[results_df["EBX Data"] == "COND_APPL_QTRS|Q1"].iloc[0]
    no_match = results_df[results_df["EBX Data"] == "COND_INCL_RUS|RU_NORTH"].iloc[0]
    fp_cols  = ["EBX Data", "FIP Data"]
    meta     = ["Reason", "Added By", "Date Added", "Resolution Status", "Resolution Notes"]
    headers  = fp_cols + meta

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name in ["X-Checks", "Grouping By", "Accounting Principles", "Conditions"]:
        ws = wb.create_sheet(sheet_name)
        if sheet_name == "Conditions":
            ws.append(headers)
            ws.append(["Guidance: do not delete this row"] + [""] * (len(headers) - 1))
            ws.append([kel_row["EBX Data"], kel_row["FIP Data"],
                       "Test fixture — expected condition", "fixture_generator", "2026-07-31", "Open", ""])
            ws.append([no_match["EBX Data"], "WRONG_FIP_DATA",
                       "Test fixture — wrong fingerprint", "fixture_generator", "2026-07-31", "Open", ""])
        else:
            ws.append(["(no entries)"])
    ws_inst = wb.create_sheet("Instructions")
    ws_inst.append(["Row 2 of each strategy sheet is a guidance row skipped by the app."])
    wb.save(OUT / "cond_kel.xlsx")
    print("  wrote cond_kel.xlsx")


# ===========================================================================
# 12. Validation Methods — skip (copy real file manually)
# ===========================================================================

def _make_validation_methods():
    print("  skipped validation_methods.xlsx — copy the real file into fixtures/")


# ===========================================================================
# main
# ===========================================================================

if __name__ == "__main__":
    print(f"Writing all fixtures to {OUT}/")
    _make_xc_pub()
    _make_xc_fip()
    _make_xc_gcoa()
    _make_gb_fip()
    _make_gb_mapping()
    _make_validation_methods()
    _make_ap_fip()
    _make_cond_fip()
    # Single combined KEL — must run last (needs other files already written)
    _make_known_exception_list()
    print("Done.")
    print()
    print("Files in fixtures/:")
    for f in sorted(OUT.iterdir()):
        if f.is_file():
            print(f"  {f.name}")
